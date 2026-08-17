#!/usr/bin/env python3
"""K5B.2 offline candidate generation and blind annotation workbook tooling.

This script consumes only the three frozen K5B.1b artifacts.  Candidate
generation is local YAKE plus the existing production extractor; workbook
validation never consults curation sessions, checkpoints, telemetry, or any
mutable live corpus.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
import unicodedata
from collections import Counter
from datetime import datetime, timezone
from importlib.metadata import version as package_version
from pathlib import Path
from random import Random
from typing import Any, Callable

import yake
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import TYPE_ERROR, TYPE_FORMULA
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from research_agent import keywords as production_keywords


FROZEN_SAMPLE_SCHEMA_VERSION = "k5b1b-bounded-sample-v1"
FROZEN_SELECTION_RULE_VERSION = "k5b1b-bounded-v1"
FROZEN_SOURCE_SCHEMA_VERSION = "k5b1b-bounded-source-v1"

ANNOTATION_SCHEMA_VERSION = "k5b2-primary-annotation-v1"
MAPPING_SCHEMA_VERSION = "k5b2-candidate-mapping-v1"
MANIFEST_SCHEMA_VERSION = "k5b2-annotation-manifest-v1"

REFERENCE_METHOD_ID = "yake_reference-v1"
PRODUCTION_METHOD_ID = "production_yake-v2"
METHOD_ORDER = (REFERENCE_METHOD_ID, PRODUCTION_METHOD_ID)
ANNOTATION_ORDER_SEED = 5202610

REFERENCE_PARAMETERS: dict[str, Any] = {
    "language": "en",
    "n": 3,
    "dedupLim": 0.85,
    "top": 6,
    "input": "title + two newlines + abstract as one YAKE input",
    "normalization": "NFKC, trim, collapse whitespace",
}
PRODUCTION_PARAMETERS: dict[str, Any] = {
    "callable": "research_agent.keywords.extract_keywords",
    "maximum_candidates": 6,
    "abstract_extraction": {"language": "en", "n": 3, "dedupLim": 0.85, "top": 25},
    "title_extraction": {"language": "en", "n": 3, "dedupLim": 0.85, "top": 8},
    "behavior": "existing production cleanup, redundancy resolution, and ordering unchanged",
}

EVAL_WORKING_DIR = Path(__file__).resolve().parent.parent / "eval_working" / "paper_keywords"
SAMPLE_PATH = EVAL_WORKING_DIR / "proposed_sample.jsonl"
RULES_PATH = EVAL_WORKING_DIR / "selection_rules.json"
SOURCE_SNAPSHOT_PATH = EVAL_WORKING_DIR / "bounded_source_snapshot.jsonl"
WORKBOOK_PATH = EVAL_WORKING_DIR / "keyword_annotation_primary.xlsx"
MAPPING_PATH = EVAL_WORKING_DIR / "keyword_candidate_mapping.json"
MANIFEST_PATH = EVAL_WORKING_DIR / "keyword_annotation_manifest.json"

PAPER_HEADERS = (
    "paper_code",
    "review_stage",
    "title",
    "abstract",
    "concept_1",
    "concept_2",
    "concept_3",
    "concept_4",
    "concept_5",
    "reviewer_notes",
)
CANDIDATE_HEADERS = (
    "candidate_id",
    "paper_code",
    "review_stage",
    "candidate_phrase",
    "decision",
    "rejection_reason",
    "matched_concept_ids",
    "confidence",
    "reviewer_notes",
)
PAPER_EVIDENCE_HEADERS = PAPER_HEADERS[:4]
CANDIDATE_EVIDENCE_HEADERS = CANDIDATE_HEADERS[:4]

DECISION_VALUES = ("accept", "reject", "uncertain")
REJECTION_REASON_VALUES = (
    "fragment",
    "affiliation_or_entity",
    "redundant",
    "too_broad",
    "malformed",
    "unrelated",
    "other",
)
CONFIDENCE_VALUES = ("high", "medium", "low")
EXPECTED_SHEETS = ("Instructions", "Paper concepts", "Candidate review")

HEADER_FILL = PatternFill("solid", fgColor="17365D")
HEADER_FONT = Font(color="FFFFFF", bold=True)
EVIDENCE_FILL = PatternFill("solid", fgColor="E7EEF8")
EDITABLE_FILL = PatternFill("solid", fgColor="FFF2CC")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
THIN_GRAY = "D9E2F3"


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _canonical_bytes(payload: Any) -> bytes:
    return json.dumps(payload, sort_keys=True, separators=(",", ":"), ensure_ascii=False).encode("utf-8")


def _sha256_payload(payload: Any) -> str:
    return hashlib.sha256(_canonical_bytes(payload)).hexdigest()


def _file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _source_hash(title: str, abstract: str | None) -> str:
    return hashlib.sha256(((title or "") + "\x00" + (abstract or "")).encode("utf-8")).hexdigest()


def _load_jsonl(path: Path) -> list[dict[str, Any]]:
    return [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]


def _json_bytes(payload: dict[str, Any]) -> bytes:
    return (json.dumps(payload, sort_keys=True, indent=2, ensure_ascii=False) + "\n").encode("utf-8")


def _write_atomic(path: Path, data: bytes) -> None:
    temporary = path.with_name(f".{path.name}.tmp")
    temporary.write_bytes(data)
    os.replace(temporary, path)


def _manifest_hash_valid(payload: dict[str, Any], field: str) -> bool:
    copy = dict(payload)
    claimed = copy.pop(field, None)
    return isinstance(claimed, str) and claimed == _sha256_payload(copy)


def validate_frozen_inputs(
    sample_path: Path = SAMPLE_PATH,
    rules_path: Path = RULES_PATH,
    source_snapshot_path: Path = SOURCE_SNAPSHOT_PATH,
) -> tuple[list[str], dict[str, Any] | None]:
    """Validate only the immutable K5B.1b inputs; never inspect live state."""
    violations: list[str] = []
    for path, label in (
        (sample_path, "frozen sample"),
        (rules_path, "selection rules"),
        (source_snapshot_path, "bounded source snapshot"),
    ):
        if not path.exists():
            violations.append(f"{label} is missing")
    if violations:
        return violations, None

    try:
        sample = _load_jsonl(sample_path)
        rules = json.loads(rules_path.read_text(encoding="utf-8"))
        sources = _load_jsonl(source_snapshot_path)
    except (OSError, json.JSONDecodeError) as exc:
        return [f"frozen input could not be read: {type(exc).__name__}"], None

    if rules.get("schema_version") != FROZEN_SAMPLE_SCHEMA_VERSION:
        violations.append("selection rules use an unexpected sample schema")
    if rules.get("selection_rule_version") != FROZEN_SELECTION_RULE_VERSION:
        violations.append("selection rules use an unexpected selection version")
    if rules.get("status") != "frozen_bounded_sample":
        violations.append("selection rules are not frozen_bounded_sample")
    if not _manifest_hash_valid(rules, "manifest_sha256"):
        violations.append("selection rules manifest_sha256 mismatch")
    if rules.get("sample_sha256") != _file_sha256(sample_path):
        violations.append("frozen sample hash does not match selection rules")
    if rules.get("bounded_source_snapshot_sha256") != _file_sha256(source_snapshot_path):
        violations.append("bounded source snapshot hash does not match selection rules")

    sample_ids = [row.get("paper_id") for row in sample]
    source_ids = [row.get("paper_id") for row in sources]
    expected_ids = rules.get("paper_ids")
    if len(sample) != 10 or len(set(sample_ids)) != 10:
        violations.append("frozen sample must contain exactly 10 distinct papers")
    if len(sources) != 10 or len(set(source_ids)) != 10:
        violations.append("source snapshot must contain exactly 10 distinct papers")
    if sorted(sample_ids) != sorted(source_ids) or sorted(sample_ids) != expected_ids:
        violations.append("sample, source snapshot, and rules paper IDs differ")

    sample_by_id = {row.get("paper_id"): row for row in sample}
    source_by_id = {row.get("paper_id"): row for row in sources}
    for paper_id in sorted(set(sample_ids) | set(source_ids)):
        sample_row = sample_by_id.get(paper_id, {})
        source_row = source_by_id.get(paper_id, {})
        if sample_row.get("schema_version") != FROZEN_SAMPLE_SCHEMA_VERSION:
            violations.append(f"{paper_id}: frozen sample row schema mismatch")
        if sample_row.get("selection_rule_version") != FROZEN_SELECTION_RULE_VERSION:
            violations.append(f"{paper_id}: frozen selection rule version mismatch")
        if sample_row.get("status") != "frozen_bounded_sample":
            violations.append(f"{paper_id}: frozen sample row status mismatch")
        if source_row.get("schema_version") != FROZEN_SOURCE_SCHEMA_VERSION:
            violations.append(f"{paper_id}: source snapshot row schema mismatch")
        actual_source_hash = _source_hash(str(source_row.get("title") or ""), source_row.get("abstract"))
        if actual_source_hash != source_row.get("source_hash"):
            violations.append(f"{paper_id}: source snapshot content hash mismatch")
        if sample_row.get("source_hash") != actual_source_hash:
            violations.append(f"{paper_id}: sample and source snapshot hashes differ")

    pilot_ids = rules.get("pilot_ids")
    if not isinstance(pilot_ids, list) or len(pilot_ids) != 2 or not set(pilot_ids) <= set(sample_ids):
        violations.append("selection rules must identify exactly two sample pilot papers")
        pilot_ids = []
    for paper_id, sample_row in sample_by_id.items():
        expected_role = "pilot_only" if paper_id in set(pilot_ids) else "headline"
        if sample_row.get("metrics_role") != expected_role:
            violations.append(f"{paper_id}: frozen metrics role is inconsistent")

    if violations:
        return violations, None
    return [], {
        "sample": sample,
        "rules": rules,
        "sources": sources,
        "sample_by_id": sample_by_id,
        "source_by_id": source_by_id,
    }


def _readable_reference_phrase(phrase: str) -> str:
    return re.sub(r"\s+", " ", unicodedata.normalize("NFKC", phrase)).strip()


def _canonical_key(phrase: str) -> str:
    return " ".join(production_keywords._canonical_tokens(phrase))


def _candidate_id(paper_id: str, canonical_key: str, seed: int) -> str:
    digest = hashlib.sha256(f"{seed}\x00{paper_id}\x00{canonical_key}".encode("utf-8")).hexdigest()
    return f"C-{digest[:12].upper()}"


def _paper_shuffle_seed(paper_id: str, seed: int) -> int:
    digest = hashlib.sha256(f"{seed}\x00{paper_id}".encode("utf-8")).digest()
    return int.from_bytes(digest[:16], "big")


def generate_candidate_bundle(
    frozen: dict[str, Any],
    seed: int = ANNOTATION_ORDER_SEED,
    reference_extractor_factory: Callable[[], Any] | None = None,
) -> dict[str, Any]:
    """Generate the two frozen methods and deterministic blind union."""
    if production_keywords.KEYWORD_EXTRACTOR_VERSION != "yake-v2":
        raise RuntimeError("production keyword extractor is no longer the frozen yake-v2 implementation")

    if reference_extractor_factory is None:
        reference_extractor_factory = lambda: yake.KeywordExtractor(lan="en", n=3, dedupLim=0.85, top=6)
    reference_extractor = reference_extractor_factory()

    rules = frozen["rules"]
    source_by_id = frozen["source_by_id"]
    pilot_ids = set(rules["pilot_ids"])
    ordered_ids = sorted(rules["paper_ids"])
    paper_codes = {paper_id: f"P{index:02d}" for index, paper_id in enumerate(ordered_ids, start=1)}
    method_counts: Counter[str] = Counter()
    candidates: list[dict[str, Any]] = []
    papers: list[dict[str, Any]] = []
    seen_candidate_ids: set[str] = set()

    for paper_id in ordered_ids:
        source = source_by_id[paper_id]
        title = str(source.get("title") or "")
        abstract = source.get("abstract")
        combined_text = f"{title}\n\n{abstract or ''}"
        reference_outputs = [
            _readable_reference_phrase(phrase)
            for phrase, _score in reference_extractor.extract_keywords(combined_text)
            if isinstance(phrase, str) and _readable_reference_phrase(phrase)
        ]
        production_outputs = list(production_keywords.extract_keywords(title, abstract))[:6]
        outputs_by_method = {
            REFERENCE_METHOD_ID: reference_outputs,
            PRODUCTION_METHOD_ID: production_outputs,
        }
        for method_id, outputs in outputs_by_method.items():
            method_counts[method_id] += len(outputs)

        union: dict[str, dict[str, Any]] = {}
        for method_id in METHOD_ORDER:
            for rank, display_phrase in enumerate(outputs_by_method[method_id], start=1):
                key = _canonical_key(display_phrase)
                if not key:
                    continue
                entry = union.setdefault(key, {"display_phrase": display_phrase, "provenance": []})
                if not any(item["method_id"] == method_id for item in entry["provenance"]):
                    entry["provenance"].append({"method_id": method_id, "rank": rank})

        paper_candidates: list[dict[str, Any]] = []
        for key, entry in union.items():
            candidate_id = _candidate_id(paper_id, key, seed)
            if candidate_id in seen_candidate_ids:
                raise RuntimeError("opaque candidate ID collision")
            seen_candidate_ids.add(candidate_id)
            paper_candidates.append({
                "candidate_id": candidate_id,
                "paper_id": paper_id,
                "paper_code": paper_codes[paper_id],
                "canonical_key": key,
                "display_phrase": entry["display_phrase"],
                "provenance": entry["provenance"],
            })
        Random(_paper_shuffle_seed(paper_id, seed)).shuffle(paper_candidates)
        candidates.extend(paper_candidates)
        papers.append({
            "paper_id": paper_id,
            "paper_code": paper_codes[paper_id],
            "review_stage": "pilot" if paper_id in pilot_ids else "headline",
            "title": title,
            "abstract": abstract or "",
            "source_hash": source["source_hash"],
        })

    method_definitions = {
        REFERENCE_METHOD_ID: {
            "method_version": REFERENCE_METHOD_ID,
            "package": "yake",
            "package_version": package_version("yake"),
            "parameters": REFERENCE_PARAMETERS,
        },
        PRODUCTION_METHOD_ID: {
            "method_version": production_keywords.KEYWORD_EXTRACTOR_VERSION,
            "evaluation_method_id": PRODUCTION_METHOD_ID,
            "package": "yake",
            "package_version": package_version("yake"),
            "parameters": PRODUCTION_PARAMETERS,
            "implementation_source_sha256": _file_sha256(Path(production_keywords.__file__)),
        },
    }
    return {
        "seed": seed,
        "papers": papers,
        "candidates": candidates,
        "method_counts": dict(method_counts),
        "method_definitions": method_definitions,
    }


def _evidence_hash(headers: tuple[str, ...], values: list[Any]) -> str:
    return _sha256_payload(dict(zip(headers, values)))


def _build_protected_evidence(bundle: dict[str, Any]) -> dict[str, Any]:
    paper_rows = []
    for row_number, paper in enumerate(bundle["papers"], start=2):
        values = [paper["paper_code"], paper["review_stage"], paper["title"], paper["abstract"]]
        paper_rows.append({
            "row_number": row_number,
            "paper_code": paper["paper_code"],
            "sha256": _evidence_hash(PAPER_EVIDENCE_HEADERS, values),
        })
    candidate_rows = []
    paper_stage = {paper["paper_code"]: paper["review_stage"] for paper in bundle["papers"]}
    for row_number, candidate in enumerate(bundle["candidates"], start=2):
        values = [
            candidate["candidate_id"],
            candidate["paper_code"],
            paper_stage[candidate["paper_code"]],
            candidate["display_phrase"],
        ]
        candidate_rows.append({
            "row_number": row_number,
            "candidate_id": candidate["candidate_id"],
            "sha256": _evidence_hash(CANDIDATE_EVIDENCE_HEADERS, values),
        })
    payload = {"paper_rows": paper_rows, "candidate_rows": candidate_rows}
    payload["overall_sha256"] = _sha256_payload(payload)
    return payload


def _style_header(sheet, cell_range: str) -> None:
    for row in sheet[cell_range]:
        for cell in row:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _add_list_validation(sheet, cell_range: str, values: tuple[str, ...], prompt: str) -> None:
    validation = DataValidation(type="list", formula1='"' + ",".join(values) + '"', allow_blank=True)
    validation.error = "Choose a value from the dropdown."
    validation.errorTitle = "Invalid value"
    validation.prompt = prompt
    validation.promptTitle = "Allowed values"
    validation.showErrorMessage = True
    validation.showInputMessage = True
    sheet.add_data_validation(validation)
    validation.add(cell_range)


def create_annotation_workbook(bundle: dict[str, Any], destination: Path) -> None:
    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "Instructions"
    concepts = workbook.create_sheet("Paper concepts")
    review = workbook.create_sheet("Candidate review")

    instruction_rows = [
        ("K5B.2 blind keyword annotation", "Use only the title and abstract shown in this workbook."),
        ("Order of work", "First complete 3–5 important scientific concepts for a paper. Then review its candidates."),
        ("Pilot first", "Label only the two papers marked pilot before beginning any headline paper."),
        ("Blinding", "Do not inspect the separate candidate mapping JSON while annotating."),
        ("Concept IDs", "The five concept cells correspond to C1, C2, C3, C4, and C5. Enter matched IDs as comma-separated values, for example C1, C3."),
        ("Decision", "accept = useful scientific keyword; reject = not useful; uncertain = cannot confidently decide."),
        ("Rejection reason", "For reject only: fragment, affiliation_or_entity, redundant, too_broad, malformed, unrelated, or other."),
        ("Confidence", "For every completed candidate decision choose high, medium, or low."),
        ("Editable cells", "Pale yellow cells are for human entry. Blue-gray cells are frozen evidence and hash-checked."),
        ("Notes", "Use notes for concise rationale or ambiguity. Do not add, delete, sort, or reorder rows."),
    ]
    instructions.append(["Topic", "Guidance"])
    for row in instruction_rows:
        instructions.append(list(row))
    _style_header(instructions, "A1:B1")
    for row_number, row in enumerate(
        instructions.iter_rows(min_row=2, max_row=instructions.max_row, min_col=1, max_col=2),
        start=2,
    ):
        row[0].fill = SECTION_FILL
        row[0].font = Font(bold=True, color="17365D", size=10)
        row[1].font = Font(size=10)
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        guidance_length = len(str(row[1].value or ""))
        instructions.row_dimensions[row_number].height = min(72, 30 + 12 * (guidance_length // 70))
    instructions.column_dimensions["A"].width = 21
    instructions.column_dimensions["B"].width = 58
    instructions.freeze_panes = "A2"
    instructions.sheet_view.showGridLines = False
    instructions.print_area = f"A1:B{instructions.max_row}"
    instructions.sheet_properties.pageSetUpPr.fitToPage = True
    instructions.page_setup.orientation = "landscape"
    instructions.page_setup.fitToWidth = 1
    instructions.page_setup.fitToHeight = 1
    instructions.protection.sheet = True

    concepts.append(list(PAPER_HEADERS))
    for paper in bundle["papers"]:
        concepts.append([
            paper["paper_code"], paper["review_stage"], paper["title"], paper["abstract"],
            None, None, None, None, None, None,
        ])
    _style_header(concepts, f"A1:J1")
    concepts.freeze_panes = "A2"
    concepts.auto_filter.ref = f"A1:J{concepts.max_row}"
    concepts.sheet_view.showGridLines = False
    for row in concepts.iter_rows(min_row=2, max_row=concepts.max_row, min_col=1, max_col=10):
        for cell in row[:4]:
            cell.fill = EVIDENCE_FILL
            cell.protection = Protection(locked=True)
        for cell in row[4:]:
            cell.fill = EDITABLE_FILL
            cell.protection = Protection(locked=False)
        for index in (2, 3, 4, 5, 6, 7, 8, 9):
            row[index].alignment = Alignment(vertical="top", wrap_text=True)
    concepts.column_dimensions["A"].width = 12
    concepts.column_dimensions["B"].width = 14
    concepts.column_dimensions["C"].width = 44
    concepts.column_dimensions["D"].width = 80
    for column in ("E", "F", "G", "H", "I"):
        concepts.column_dimensions[column].width = 24
    concepts.column_dimensions["J"].width = 36
    concepts.row_dimensions[1].height = 32
    for row_number in range(2, concepts.max_row + 1):
        concepts.row_dimensions[row_number].height = 105
    concepts.protection.sheet = True
    concepts.protection.autoFilter = False
    concepts.protection.sort = False

    review.append(list(CANDIDATE_HEADERS))
    paper_stage = {paper["paper_code"]: paper["review_stage"] for paper in bundle["papers"]}
    for candidate in bundle["candidates"]:
        review.append([
            candidate["candidate_id"], candidate["paper_code"], paper_stage[candidate["paper_code"]],
            candidate["display_phrase"], None, None, None, None, None,
        ])
    _style_header(review, f"A1:I1")
    review.freeze_panes = "A2"
    review.auto_filter.ref = f"A1:I{review.max_row}"
    review.sheet_view.showGridLines = False
    for row in review.iter_rows(min_row=2, max_row=review.max_row, min_col=1, max_col=9):
        for cell in row[:4]:
            cell.fill = EVIDENCE_FILL
            cell.protection = Protection(locked=True)
        for cell in row[4:]:
            cell.fill = EDITABLE_FILL
            cell.protection = Protection(locked=False)
        for index in (3, 6, 8):
            row[index].alignment = Alignment(vertical="top", wrap_text=True)
    review.column_dimensions["A"].width = 18
    review.column_dimensions["B"].width = 12
    review.column_dimensions["C"].width = 14
    review.column_dimensions["D"].width = 42
    review.column_dimensions["E"].width = 14
    review.column_dimensions["F"].width = 24
    review.column_dimensions["G"].width = 24
    review.column_dimensions["H"].width = 14
    review.column_dimensions["I"].width = 38
    review.row_dimensions[1].height = 32
    for row_number in range(2, review.max_row + 1):
        review.row_dimensions[row_number].height = 42
    _add_list_validation(review, f"E2:E{review.max_row}", DECISION_VALUES, "accept, reject, or uncertain")
    _add_list_validation(review, f"F2:F{review.max_row}", REJECTION_REASON_VALUES, "Required for reject only")
    _add_list_validation(review, f"H2:H{review.max_row}", CONFIDENCE_VALUES, "high, medium, or low")
    review.protection.sheet = True
    review.protection.autoFilter = False
    review.protection.sort = False

    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)


def _build_mapping(bundle: dict[str, Any]) -> dict[str, Any]:
    mapping = {
        "schema_version": MAPPING_SCHEMA_VERSION,
        "candidate_order_seed": bundle["seed"],
        "method_order": list(METHOD_ORDER),
        "methods": bundle["method_definitions"],
        "method_candidate_counts": bundle["method_counts"],
        "paper_count": len(bundle["papers"]),
        "candidate_count": len(bundle["candidates"]),
        "papers": [
            {
                "paper_id": paper["paper_id"],
                "paper_code": paper["paper_code"],
                "review_stage": paper["review_stage"],
                "source_hash": paper["source_hash"],
            }
            for paper in bundle["papers"]
        ],
        "candidates": [
            {
                "candidate_id": candidate["candidate_id"],
                "paper_id": candidate["paper_id"],
                "paper_code": candidate["paper_code"],
                "canonical_key": candidate["canonical_key"],
                "display_phrase": candidate["display_phrase"],
                "methods": candidate["provenance"],
            }
            for candidate in bundle["candidates"]
        ],
    }
    mapping["mapping_sha256"] = _sha256_payload(mapping)
    return mapping


def _build_manifest(bundle: dict[str, Any], mapping_path: Path, protected: dict[str, Any]) -> dict[str, Any]:
    pilot_papers = [paper for paper in bundle["papers"] if paper["review_stage"] == "pilot"]
    headline_papers = [paper for paper in bundle["papers"] if paper["review_stage"] == "headline"]
    manifest = {
        "schema_version": MANIFEST_SCHEMA_VERSION,
        "annotation_schema_version": ANNOTATION_SCHEMA_VERSION,
        "status": "awaiting_human_annotation",
        "creation_timestamp": _now_iso(),
        "workbook_filename": WORKBOOK_PATH.name,
        "mapping_filename": MAPPING_PATH.name,
        "frozen_inputs": {
            "sample_path": str(SAMPLE_PATH),
            "sample_sha256": _file_sha256(SAMPLE_PATH),
            "selection_rules_path": str(RULES_PATH),
            "selection_rules_sha256": _file_sha256(RULES_PATH),
            "source_snapshot_path": str(SOURCE_SNAPSHOT_PATH),
            "source_snapshot_sha256": _file_sha256(SOURCE_SNAPSHOT_PATH),
        },
        "candidate_mapping_sha256": _file_sha256(mapping_path),
        "candidate_order_seed": bundle["seed"],
        "paper_count": len(bundle["papers"]),
        "candidate_count": len(bundle["candidates"]),
        "method_candidate_counts": bundle["method_counts"],
        "paper_row_order": [paper["paper_code"] for paper in bundle["papers"]],
        "candidate_row_order": [candidate["candidate_id"] for candidate in bundle["candidates"]],
        "pilot_ids": [paper["paper_id"] for paper in pilot_papers],
        "pilot_codes": [paper["paper_code"] for paper in pilot_papers],
        "headline_ids": [paper["paper_id"] for paper in headline_papers],
        "headline_codes": [paper["paper_code"] for paper in headline_papers],
        "protected_evidence": protected,
        "editable_columns_excluded_from_protected_hashes": {
            "Paper concepts": list(PAPER_HEADERS[4:]),
            "Candidate review": list(CANDIDATE_HEADERS[4:]),
        },
    }
    manifest["manifest_sha256"] = _sha256_payload(manifest)
    return manifest


def _cell_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _parse_concept_ids(value: Any) -> tuple[list[str], bool]:
    text = _cell_text(value)
    if not text:
        return [], True
    tokens = [token.upper() for token in re.split(r"[\s,;]+", text) if token]
    valid = bool(tokens) and all(re.fullmatch(r"C[1-5]", token) for token in tokens) and len(tokens) == len(set(tokens))
    return tokens, valid


def _validation_specs(candidate_row_count: int) -> set[tuple[str, str]]:
    return {
        (f"E2:E{candidate_row_count + 1}", '"' + ",".join(DECISION_VALUES) + '"'),
        (f"F2:F{candidate_row_count + 1}", '"' + ",".join(REJECTION_REASON_VALUES) + '"'),
        (f"H2:H{candidate_row_count + 1}", '"' + ",".join(CONFIDENCE_VALUES) + '"'),
    }


def validate_annotation_workbook(
    require_pilots_complete: bool = False,
    require_all_complete: bool = False,
    workbook_path: Path = WORKBOOK_PATH,
    mapping_path: Path = MAPPING_PATH,
    manifest_path: Path = MANIFEST_PATH,
    sample_path: Path = SAMPLE_PATH,
    rules_path: Path = RULES_PATH,
    source_snapshot_path: Path = SOURCE_SNAPSHOT_PATH,
) -> list[str]:
    violations: list[str] = []
    for path, label in ((workbook_path, "annotation workbook"), (mapping_path, "candidate mapping"), (manifest_path, "annotation manifest")):
        if not path.exists():
            violations.append(f"{label} is missing")
    if violations:
        return violations

    frozen_violations, frozen = validate_frozen_inputs(sample_path, rules_path, source_snapshot_path)
    violations.extend(f"frozen input: {item}" for item in frozen_violations)
    try:
        mapping = json.loads(mapping_path.read_text(encoding="utf-8"))
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        workbook = load_workbook(workbook_path, data_only=False)
    except (OSError, json.JSONDecodeError, ValueError) as exc:
        return violations + [f"annotation evidence could not be read: {type(exc).__name__}"]

    if mapping.get("schema_version") != MAPPING_SCHEMA_VERSION:
        violations.append("candidate mapping schema is not current")
    if not _manifest_hash_valid(mapping, "mapping_sha256"):
        violations.append("candidate mapping self-hash mismatch")
    if manifest.get("schema_version") != MANIFEST_SCHEMA_VERSION:
        violations.append("annotation manifest schema is not current")
    if manifest.get("annotation_schema_version") != ANNOTATION_SCHEMA_VERSION:
        violations.append("annotation workbook schema is not current")
    if not _manifest_hash_valid(manifest, "manifest_sha256"):
        violations.append("annotation manifest self-hash mismatch")
    if manifest.get("candidate_mapping_sha256") != _file_sha256(mapping_path):
        violations.append("candidate mapping file hash mismatch")

    input_paths = {
        "sample_sha256": sample_path,
        "selection_rules_sha256": rules_path,
        "source_snapshot_sha256": source_snapshot_path,
    }
    frozen_bindings = manifest.get("frozen_inputs", {})
    for field, path in input_paths.items():
        if frozen_bindings.get(field) != _file_sha256(path):
            violations.append(f"manifest {field} binding mismatch")

    if workbook.sheetnames != list(EXPECTED_SHEETS):
        violations.append("workbook sheet names or order changed")
        return violations
    concepts = workbook["Paper concepts"]
    review = workbook["Candidate review"]
    paper_count = manifest.get("paper_count")
    candidate_count = manifest.get("candidate_count")
    if paper_count != 10 or mapping.get("paper_count") != 10:
        violations.append("annotation evidence does not describe exactly 10 papers")
    if candidate_count != mapping.get("candidate_count"):
        violations.append("candidate count differs between mapping and manifest")
    if tuple(cell.value for cell in concepts[1]) != PAPER_HEADERS:
        violations.append("Paper concepts headers changed")
    if tuple(cell.value for cell in review[1]) != CANDIDATE_HEADERS:
        violations.append("Candidate review headers changed")
    if concepts.max_row != paper_count + 1 or concepts.max_column != len(PAPER_HEADERS):
        violations.append("Paper concepts rows or columns were added or removed")
    if review.max_row != candidate_count + 1 or review.max_column != len(CANDIDATE_HEADERS):
        violations.append("Candidate review rows or columns were added or removed")

    paper_order = [_cell_text(concepts.cell(row=row, column=1).value) for row in range(2, concepts.max_row + 1)]
    candidate_order = [_cell_text(review.cell(row=row, column=1).value) for row in range(2, review.max_row + 1)]
    if paper_order != manifest.get("paper_row_order"):
        violations.append("Paper concepts rows are missing, added, or reordered")
    if candidate_order != manifest.get("candidate_row_order"):
        violations.append("Candidate review rows are missing, added, or reordered")

    protected = manifest.get("protected_evidence", {})
    paper_hash_records = protected.get("paper_rows", [])
    candidate_hash_records = protected.get("candidate_rows", [])
    for record in paper_hash_records:
        row_number = record.get("row_number")
        if not isinstance(row_number, int) or row_number > concepts.max_row:
            violations.append("Paper concepts protected row binding is invalid")
            continue
        values = [concepts.cell(row=row_number, column=column).value for column in range(1, 5)]
        if record.get("sha256") != _evidence_hash(PAPER_EVIDENCE_HEADERS, values):
            violations.append(f"{record.get('paper_code')}: protected paper evidence hash mismatch")
    for record in candidate_hash_records:
        row_number = record.get("row_number")
        if not isinstance(row_number, int) or row_number > review.max_row:
            violations.append("Candidate review protected row binding is invalid")
            continue
        values = [review.cell(row=row_number, column=column).value for column in range(1, 5)]
        if record.get("sha256") != _evidence_hash(CANDIDATE_EVIDENCE_HEADERS, values):
            violations.append(f"{record.get('candidate_id')}: protected candidate evidence hash mismatch")
    protected_copy = {"paper_rows": paper_hash_records, "candidate_rows": candidate_hash_records}
    if protected.get("overall_sha256") != _sha256_payload(protected_copy):
        violations.append("protected evidence aggregate hash mismatch")

    mapping_candidates = mapping.get("candidates", [])
    mapped_order = [row.get("candidate_id") for row in mapping_candidates]
    if mapped_order != candidate_order:
        violations.append("candidate mapping and workbook row order differ")
    mapped_paper_order = [row.get("paper_code") for row in mapping.get("papers", [])]
    if mapped_paper_order != paper_order:
        violations.append("paper mapping and workbook row order differ")

    if concepts.freeze_panes != "A2" or review.freeze_panes != "A2":
        violations.append("required frozen header rows changed")
    if concepts.auto_filter.ref != f"A1:J{paper_count + 1}" or review.auto_filter.ref != f"A1:I{candidate_count + 1}":
        violations.append("required worksheet filters changed")
    observed_validations = {(str(item.sqref), item.formula1) for item in review.data_validations.dataValidation}
    if observed_validations != _validation_specs(candidate_count):
        violations.append("candidate dropdown validations changed")

    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == TYPE_FORMULA:
                    violations.append(f"unexpected formula in {sheet.title}!{cell.coordinate}")
                elif cell.data_type == TYPE_ERROR:
                    violations.append(f"spreadsheet error in {sheet.title}!{cell.coordinate}")

    stages_by_code = {
        _cell_text(concepts.cell(row=row, column=1).value): _cell_text(concepts.cell(row=row, column=2).value)
        for row in range(2, concepts.max_row + 1)
    }
    concept_values_by_code: dict[str, list[str]] = {}
    for row_number in range(2, concepts.max_row + 1):
        code = _cell_text(concepts.cell(row=row_number, column=1).value)
        concepts_entered = [_cell_text(concepts.cell(row=row_number, column=column).value) for column in range(5, 10)]
        concept_values_by_code[code] = concepts_entered
        nonblank = [value for value in concepts_entered if value]
        if len(nonblank) != len({value.casefold() for value in nonblank}):
            violations.append(f"{code}: concept entries must be distinct")
        completion_required = require_all_complete or (
            require_pilots_complete and stages_by_code.get(code) == "pilot"
        )
        if completion_required and not 3 <= len(nonblank) <= 5:
            label = "paper" if require_all_complete else "pilot"
            violations.append(f"{code}: {label} requires 3–5 completed concepts")

    completed_candidate_codes: set[str] = set()
    for row_number in range(2, review.max_row + 1):
        candidate_id = _cell_text(review.cell(row=row_number, column=1).value)
        code = _cell_text(review.cell(row=row_number, column=2).value)
        decision = _cell_text(review.cell(row=row_number, column=5).value).casefold()
        reason = _cell_text(review.cell(row=row_number, column=6).value).casefold()
        matched, matched_valid = _parse_concept_ids(review.cell(row=row_number, column=7).value)
        confidence = _cell_text(review.cell(row=row_number, column=8).value).casefold()
        if decision and decision not in DECISION_VALUES:
            violations.append(f"{candidate_id}: invalid decision")
        if reason and reason not in REJECTION_REASON_VALUES:
            violations.append(f"{candidate_id}: invalid rejection reason")
        if confidence and confidence not in CONFIDENCE_VALUES:
            violations.append(f"{candidate_id}: invalid confidence")
        if not matched_valid:
            violations.append(f"{candidate_id}: invalid or duplicate matched concept IDs")
        if not decision:
            if reason or matched or confidence:
                violations.append(f"{candidate_id}: annotation fields entered without a decision")
            completion_required = require_all_complete or (
                require_pilots_complete and stages_by_code.get(code) == "pilot"
            )
            if completion_required:
                label = "candidate" if require_all_complete else "pilot candidate"
                violations.append(f"{candidate_id}: {label} decision is incomplete")
            continue
        completed_candidate_codes.add(code)
        if not confidence:
            violations.append(f"{candidate_id}: completed decision requires confidence")
        if decision == "reject" and not reason:
            violations.append(f"{candidate_id}: reject requires a rejection reason")
        if decision != "reject" and reason:
            violations.append(f"{candidate_id}: rejection reason is only valid for reject")
        if decision == "accept" and not matched:
            violations.append(f"{candidate_id}: accept requires at least one matched concept ID")
        if decision == "reject" and matched:
            violations.append(f"{candidate_id}: reject must not contain matched concept IDs")
        entered = concept_values_by_code.get(code, [])
        for concept_id in matched:
            if not entered[int(concept_id[1]) - 1]:
                violations.append(f"{candidate_id}: {concept_id} does not identify a completed concept")

    for code in completed_candidate_codes:
        concept_count = sum(bool(value) for value in concept_values_by_code.get(code, []))
        if concept_count < 3:
            violations.append(f"{code}: complete 3–5 concepts before candidate review")

    if frozen is not None:
        expected_pilot_ids = sorted(frozen["rules"]["pilot_ids"])
        if sorted(manifest.get("pilot_ids", [])) != expected_pilot_ids:
            violations.append("manifest pilot IDs differ from frozen selection rules")
        expected_headline = sorted(set(frozen["rules"]["paper_ids"]) - set(expected_pilot_ids))
        if sorted(manifest.get("headline_ids", [])) != expected_headline:
            violations.append("manifest headline IDs differ from frozen selection rules")
    return violations


def generate_annotation_artifacts(replace: bool = False) -> int:
    outputs = (WORKBOOK_PATH, MAPPING_PATH, MANIFEST_PATH)
    if any(path.exists() for path in outputs) and not replace:
        print("refused: annotation output exists; pass --replace to overwrite deliberately.", file=sys.stderr)
        return 3

    violations, frozen = validate_frozen_inputs(SAMPLE_PATH, RULES_PATH, SOURCE_SNAPSHOT_PATH)
    if violations or frozen is None:
        for violation in violations:
            print(f"VIOLATION: {violation}", file=sys.stderr)
        return 4

    bundle = generate_candidate_bundle(frozen)
    mapping = _build_mapping(bundle)
    protected = _build_protected_evidence(bundle)
    EVAL_WORKING_DIR.mkdir(parents=True, exist_ok=True)
    temporary_workbook = WORKBOOK_PATH.with_name(f".{WORKBOOK_PATH.name}.tmp.xlsx")
    temporary_mapping = MAPPING_PATH.with_name(f".{MAPPING_PATH.name}.tmp")
    temporary_manifest = MANIFEST_PATH.with_name(f".{MANIFEST_PATH.name}.tmp")
    try:
        create_annotation_workbook(bundle, temporary_workbook)
        temporary_mapping.write_bytes(_json_bytes(mapping))
        manifest = _build_manifest(bundle, temporary_mapping, protected)
        temporary_manifest.write_bytes(_json_bytes(manifest))
        validation = validate_annotation_workbook(
            workbook_path=temporary_workbook,
            mapping_path=temporary_mapping,
            manifest_path=temporary_manifest,
            sample_path=SAMPLE_PATH,
            rules_path=RULES_PATH,
            source_snapshot_path=SOURCE_SNAPSHOT_PATH,
        )
        if validation:
            for violation in validation:
                print(f"VIOLATION: {violation}", file=sys.stderr)
            return 4
        os.replace(temporary_workbook, WORKBOOK_PATH)
        os.replace(temporary_mapping, MAPPING_PATH)
        os.replace(temporary_manifest, MANIFEST_PATH)
    finally:
        for temporary in (temporary_workbook, temporary_mapping, temporary_manifest):
            if temporary.exists():
                temporary.unlink()

    counts = bundle["method_counts"]
    print(
        "created blind annotation workbook: "
        f"papers={len(bundle['papers'])}, reference={counts.get(REFERENCE_METHOD_ID, 0)}, "
        f"production={counts.get(PRODUCTION_METHOD_ID, 0)}, union={len(bundle['candidates'])}"
    )
    return 0


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="K5B.2 bounded keyword annotation tooling.")
    subparsers = parser.add_subparsers(dest="command", required=True)
    generate_parser = subparsers.add_parser("generate")
    generate_parser.add_argument("--replace", action="store_true")
    validate_parser = subparsers.add_parser("validate")
    validate_parser.add_argument("--require-pilots-complete", action="store_true")
    validate_parser.add_argument("--require-all-complete", action="store_true")
    args = parser.parse_args(argv)

    if args.command == "generate":
        return generate_annotation_artifacts(replace=args.replace)
    if args.command == "validate":
        violations = validate_annotation_workbook(
            require_pilots_complete=args.require_pilots_complete,
            require_all_complete=args.require_all_complete,
            workbook_path=WORKBOOK_PATH,
            mapping_path=MAPPING_PATH,
            manifest_path=MANIFEST_PATH,
            sample_path=SAMPLE_PATH,
            rules_path=RULES_PATH,
            source_snapshot_path=SOURCE_SNAPSHOT_PATH,
        )
        if violations:
            for violation in violations:
                print(f"VIOLATION: {violation}", file=sys.stderr)
            return 1
        print("valid: annotation workbook, mapping, manifest, and frozen inputs pass the K5B.2 contract.")
        return 0
    return 2


if __name__ == "__main__":
    raise SystemExit(main())
