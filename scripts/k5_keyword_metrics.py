#!/usr/bin/env python3
"""K5B.3 offline annotation freeze and bounded descriptive metrics."""
from __future__ import annotations

import argparse, csv, hashlib, io, json, os, sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from scripts import k5_keyword_annotation as annotation

BASE = annotation.EVAL_WORKING_DIR
FROZEN_PATH = BASE / "keyword_annotation_frozen.json"
RESULTS_PATH = BASE / "keyword_method_results.json"
CSV_PATH = BASE / "keyword_method_results.csv"
SUMMARY_PATH = BASE / "keyword_evaluation_summary.md"
FROZEN_SCHEMA = "k5b3-frozen-annotation-v1"
RESULTS_SCHEMA = "k5b3-bounded-results-v1"
REVIEWER_TYPE = "ai_assisted_human_approved"
METHODS = (annotation.REFERENCE_METHOD_ID, annotation.PRODUCTION_METHOD_ID)

def now() -> str: return datetime.now(timezone.utc).isoformat()
def canon(value: Any) -> bytes: return json.dumps(value,sort_keys=True,separators=(",",":"),ensure_ascii=False).encode()
def payload_hash(value: Any) -> str: return hashlib.sha256(canon(value)).hexdigest()
def file_hash(path: Path) -> str: return annotation._file_sha256(path)
def atomic(path: Path, data: bytes) -> None:
    tmp=path.with_name("."+path.name+".tmp");tmp.write_bytes(data);os.replace(tmp,path)
def json_bytes(value: Any) -> bytes: return (json.dumps(value,sort_keys=True,indent=2,ensure_ascii=False)+"\n").encode()

def validate_frozen_payload(data: dict[str,Any]) -> list[str]:
    errors=[];copy=dict(data);claimed=copy.pop("frozen_annotation_sha256",None)
    if data.get("schema_version")!=FROZEN_SCHEMA: errors.append("frozen schema mismatch")
    if claimed!=payload_hash(copy): errors.append("frozen annotation self-hash mismatch")
    if data.get("reviewer_type")!=REVIEWER_TYPE: errors.append("reviewer provenance mismatch")
    papers=data.get("papers",[]);candidates=data.get("candidates",[])
    if len(papers)!=10 or len({x.get('paper_code') for x in papers})!=10: errors.append("frozen paper count/identity mismatch")
    if len(candidates)!=92 or len({x.get('candidate_id') for x in candidates})!=92: errors.append("frozen candidate count/identity mismatch")
    pilots=[x for x in papers if x.get('review_stage')=='pilot']
    if len(pilots)!=2 or any(not x.get('excluded_from_headline_metrics') for x in pilots): errors.append("pilot exclusion mismatch")
    if any(x.get('excluded_from_headline_metrics') for x in papers if x.get('review_stage')=='headline'): errors.append("headline exclusion mismatch")
    concepts={x['paper_code']:x['concepts'] for x in papers}
    for paper in papers:
        if not 3<=sum(v not in (None,"") for v in paper['concepts'])<=5: errors.append(f"{paper['paper_code']}: concept completeness")
    for row in candidates:
        d=row.get('decision');r=row.get('rejection_reason');ids,valid=annotation._parse_concept_ids(row.get('matched_concept_ids'))
        if d not in annotation.DECISION_VALUES or row.get('confidence') not in annotation.CONFIDENCE_VALUES: errors.append(f"{row.get('candidate_id')}: incomplete")
        if bool(r)!=(d=='reject') or not valid: errors.append(f"{row.get('candidate_id')}: semantic mismatch")
        for cid in ids:
            if not concepts[row['paper_code']][int(cid[1])-1]: errors.append(f"{row['candidate_id']}: unpopulated {cid}")
    return errors

def freeze_annotations(replace: bool=False) -> dict[str,Any]:
    if FROZEN_PATH.exists() and not replace: raise FileExistsError("frozen annotation exists; pass --replace")
    errors=annotation.validate_annotation_workbook(require_all_complete=True)
    if errors: raise ValueError("; ".join(errors))
    wb=load_workbook(annotation.WORKBOOK_PATH,data_only=False);manifest=json.loads(annotation.MANIFEST_PATH.read_text())
    ps=wb['Paper concepts'];cs=wb['Candidate review']
    papers=[]
    for r in range(2,ps.max_row+1):
        stage=ps.cell(r,2).value
        papers.append({'paper_code':ps.cell(r,1).value,'review_stage':stage,'concepts':[ps.cell(r,c).value for c in range(5,10)],
          'reviewer_notes':ps.cell(r,10).value,'excluded_from_headline_metrics':stage=='pilot'})
    candidates=[]
    for r in range(2,cs.max_row+1):
        candidates.append({'candidate_id':cs.cell(r,1).value,'paper_code':cs.cell(r,2).value,'review_stage':cs.cell(r,3).value,
          'decision':cs.cell(r,5).value,'rejection_reason':cs.cell(r,6).value,'matched_concept_ids':cs.cell(r,7).value,
          'confidence':cs.cell(r,8).value,'reviewer_notes':cs.cell(r,9).value})
    data={'schema_version':FROZEN_SCHEMA,'status':'frozen_complete_annotation','frozen_at':now(),'reviewer_type':REVIEWER_TYPE,
      'bindings':{'workbook_sha256':file_hash(annotation.WORKBOOK_PATH),'annotation_manifest_sha256':file_hash(annotation.MANIFEST_PATH),
        'candidate_mapping_sha256':file_hash(annotation.MAPPING_PATH),'sample_sha256':file_hash(annotation.SAMPLE_PATH),
        'selection_rules_sha256':file_hash(annotation.RULES_PATH),'source_snapshot_sha256':file_hash(annotation.SOURCE_SNAPSHOT_PATH),
        'protected_evidence_sha256':manifest['protected_evidence']['overall_sha256']},
      'pilot_codes':manifest['pilot_codes'],'headline_codes':manifest['headline_codes'],'papers':papers,'candidates':candidates}
    data['frozen_annotation_sha256']=payload_hash(data)
    errors=validate_frozen_payload(data)
    if errors: raise ValueError("; ".join(errors))
    atomic(FROZEN_PATH,json_bytes(data));return data

def rate(n:int,d:int)->float|None: return None if not d else round(n/d,6)
def method_stats(rows:list[dict], paper_codes:list[str], concepts:dict[str,list]) -> dict:
    decisions=Counter(x['decision'] for x in rows);reasons=Counter(x['rejection_reason'] for x in rows if x['decision']=='reject')
    per={}
    for code in paper_codes:
        subset=[x for x in rows if x['paper_code']==code];dc=Counter(x['decision'] for x in subset);matched=set()
        for x in subset:
            if x['decision']=='accept': matched.update(annotation._parse_concept_ids(x['matched_concept_ids'])[0])
        populated=sum(v not in (None,"") for v in concepts[code]);accepted=dc['accept'];resolved=accepted+dc['reject']
        per[code]={'emitted':len(subset),'accept':accepted,'reject':dc['reject'],'uncertain':dc['uncertain'],
          'acceptance_rate':rate(accepted,len(subset)),'resolved_precision':rate(accepted,resolved),'uncertain_rate':rate(dc['uncertain'],len(subset)),
          'accepted_candidates':accepted,'matched_concepts':len(matched),'populated_concepts':populated,'concept_coverage':rate(len(matched),populated)}
    total=len(rows);resolved=decisions['accept']+decisions['reject'];rejects=decisions['reject']
    reason_data={reason:{'count':reasons[reason],'rate_among_rejected':rate(reasons[reason],rejects),'rate_over_all_outputs':rate(reasons[reason],total)} for reason in annotation.REJECTION_REASON_VALUES}
    return {'total_emitted_candidates':total,'accept':decisions['accept'],'reject':decisions['reject'],'uncertain':decisions['uncertain'],
      'acceptance_rate':rate(decisions['accept'],total),'resolved_precision':rate(decisions['accept'],resolved),'uncertain_rate':rate(decisions['uncertain'],total),
      'rejection_reasons':reason_data,'mean_accepted_candidates_per_paper':round(decisions['accept']/len(paper_codes),6),
      'concept_coverage_per_paper':{c:per[c]['concept_coverage'] for c in paper_codes},
      'macro_average_concept_coverage':round(sum(per[c]['concept_coverage'] for c in paper_codes)/len(paper_codes),6),'per_paper':per}

def calculate(data:dict[str,Any],replace:bool=False)->dict[str,Any]:
    for p in (RESULTS_PATH,CSV_PATH,SUMMARY_PATH):
        if p.exists() and not replace: raise FileExistsError("result exists; pass --replace")
    errors=validate_frozen_payload(data)
    if errors: raise ValueError("; ".join(errors))
    if data['bindings']['candidate_mapping_sha256']!=file_hash(annotation.MAPPING_PATH): raise ValueError('mapping binding mismatch')
    mapping=json.loads(annotation.MAPPING_PATH.read_text()) # unblinding begins here, after frozen artifact exists and validates
    annotations={x['candidate_id']:x for x in data['candidates']};headline=list(data['headline_codes']);concepts={x['paper_code']:x['concepts'] for x in data['papers']}
    method_rows={m:[] for m in METHODS}
    for candidate in mapping['candidates']:
        if candidate['paper_code'] not in headline: continue
        judged=annotations[candidate['candidate_id']]
        for provenance in candidate['methods']:
            if provenance['method_id'] in method_rows: method_rows[provenance['method_id']].append(judged)
    stats={m:method_stats(method_rows[m],headline,concepts) for m in METHODS};ref,prod=stats[METHODS[0]],stats[METHODS[1]]
    per=[]
    for code in headline:
        a,b=ref['per_paper'][code],prod['per_paper'][code]
        per.append({'paper_code':code,'yake_reference-v1':a,'production_yake-v2':b,'delta_production_minus_reference':{k:round((b[k] or 0)-(a[k] or 0),6) for k in ('emitted','accept','reject','uncertain','acceptance_rate','resolved_precision','uncertain_rate','concept_coverage')}})
    if prod['resolved_precision']>=ref['resolved_precision'] and prod['macro_average_concept_coverage']>=ref['macro_average_concept_coverage']:
        choice='keep production YAKE-v2 unchanged';basis='Production matched or exceeded the reference on resolved precision and macro concept coverage in this bounded headline sample.'
    else:
        choice='make a specific later deterministic improvement';basis='Production did not match the reference on both resolved precision and macro concept coverage; inspect its observed rejection mix before a separately scoped deterministic change.'
    recommendation={'choice':choice,'basis':basis,'limitation':'Based only on eight product-local headline papers with AI-assisted, human-approved annotation; not an external benchmark or universal claim.'}
    result={'schema_version':RESULTS_SCHEMA,'status':'bounded_descriptive_results','created_at':now(),'reviewer_type':REVIEWER_TYPE,
      'scope':{'headline_codes':headline,'excluded_pilot_codes':data['pilot_codes'],'paper_count':8,'uncertain_treatment':'neither accepted nor rejected'},
      'bindings':{'frozen_annotation_sha256':file_hash(FROZEN_PATH),'mapping_sha256':file_hash(annotation.MAPPING_PATH)},'methods':stats,'per_paper_comparison':per,
      'recommendation':recommendation,'limitations':['product-local corpus','eight headline papers','pilot papers excluded','AI-assisted human-approved annotation','not an external benchmark','no significance testing or universal claim']}
    result['results_sha256']=payload_hash(result);atomic(RESULTS_PATH,json_bytes(result))
    out=io.StringIO();fields=['paper_code']
    metrics=['emitted','accept','reject','uncertain','acceptance_rate','resolved_precision','uncertain_rate','concept_coverage']
    fields += [f'{m}_{x}' for m in METHODS for x in metrics]+[f'delta_{x}' for x in metrics]
    writer=csv.DictWriter(out,fieldnames=fields);writer.writeheader()
    for row in per:
        flat={'paper_code':row['paper_code']}
        for m in METHODS:
            for x in metrics:flat[f'{m}_{x}']=row[m][x]
        for x in metrics:flat[f'delta_{x}']=row['delta_production_minus_reference'][x]
        writer.writerow(flat)
    atomic(CSV_PATH,out.getvalue().encode())
    lines=['# Bounded keyword-method evaluation','','Reviewer provenance: `ai_assisted_human_approved`. P01–P02 are excluded from headline metrics.','',
      '| Method | Outputs | Accept | Reject | Uncertain | Acceptance | Resolved precision | Macro concept coverage |','|---|---:|---:|---:|---:|---:|---:|---:|']
    for m in METHODS:
        s=stats[m];lines.append(f"| {m} | {s['total_emitted_candidates']} | {s['accept']} | {s['reject']} | {s['uncertain']} | {s['acceptance_rate']:.1%} | {s['resolved_precision']:.1%} | {s['macro_average_concept_coverage']:.1%} |")
    lines += ['','## Recommendation','',f"**{choice}.** {basis}",'',recommendation['limitation'],'','No significance test or general claim is made.']
    atomic(SUMMARY_PATH,("\n".join(lines)+"\n").encode());return result

def main(argv=None)->int:
    p=argparse.ArgumentParser();p.add_argument('command',choices=('run','freeze','validate-frozen','metrics'));p.add_argument('--replace',action='store_true');a=p.parse_args(argv)
    try:
        if a.command in ('run','freeze'): data=freeze_annotations(a.replace);print('frozen annotation validated')
        else: data=json.loads(FROZEN_PATH.read_text())
        if a.command=='validate-frozen':
            errors=validate_frozen_payload(data);print('valid frozen annotation' if not errors else '\n'.join(errors));return bool(errors)
        if a.command in ('run','metrics'):
            result=calculate(data,a.replace);print(json.dumps({m:{k:result['methods'][m][k] for k in ('total_emitted_candidates','accept','reject','uncertain','acceptance_rate','resolved_precision','macro_average_concept_coverage')} for m in METHODS},sort_keys=True))
        return 0
    except Exception as exc: print(f'error: {exc}',file=sys.stderr);return 1
if __name__=='__main__':raise SystemExit(main())
