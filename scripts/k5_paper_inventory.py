#!/usr/bin/env python3
"""K5B.1a offline paper inventory, failure screening, and sample gate.

This module reads already-saved curation sessions.  It never extracts fresh
keywords, calls a provider/search service, or writes a checkpoint.  Its XLSX
output is a preparatory human screen of heuristic stored-keyword hits; it is
not the formal K5 annotation workbook.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
import textwrap
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from random import Random
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from research_agent.curation_session import list_curation_sessions, load_curation_session
from research_agent.keywords import _MIN_ABSTRACT_CHARS, _MIN_ABSTRACT_WORDS
from research_agent.qa import QA_CHECKPOINT_DB_PATH, sqlite_checkpointer
from research_agent.schema import Paper

INVENTORY_SCHEMA_VERSION = "k5b1a-inventory-v2"
SCREENING_SELECTION_RULE_VERSION = "k5b1a-v2"
SAMPLE_SCHEMA_VERSION = "k5b1b-bounded-sample-v1"
SELECTION_RULE_VERSION = "k5b1b-bounded-v1"
SOURCE_SNAPSHOT_SCHEMA_VERSION = "k5b1b-bounded-source-v1"
SCREENING_MANIFEST_SCHEMA_VERSION = "k5b1a-failure-screening-manifest-v1"

K5_RANDOM_SEED = 5202608

EVAL_WORKING_DIR = Path(__file__).resolve().parent.parent / "eval_working" / "paper_keywords"
INVENTORY_PATH = EVAL_WORKING_DIR / "inventory.jsonl"
SAMPLE_PATH = EVAL_WORKING_DIR / "proposed_sample.jsonl"
RULES_PATH = EVAL_WORKING_DIR / "selection_rules.json"
SCREENING_WORKBOOK_PATH = EVAL_WORKING_DIR / "failure_screening.xlsx"
SCREENING_MANIFEST_PATH = EVAL_WORKING_DIR / "failure_screening_manifest.json"
SOURCE_SNAPSHOT_PATH = EVAL_WORKING_DIR / "bounded_source_snapshot.jsonl"

STRATA_TARGETS = {"confirmed_failure": 4, "product_local_diversity": 4, "local_edge_case": 2}
PILOT_COUNT = 2
DOUBLE_REVIEW_COUNT = 3

EXPECTED_SCREENING_ROW_COUNT = 21
EXPECTED_SCREENING_PAPER_COUNT = 16
EXPECTED_SCREENING_DECISIONS = {"yes": 21}
EXPECTED_SCREENING_REASONS = {"fragment": 20, "too_broad": 1}

CONFIRMED_FAILURE_VALUES = ("yes", "no", "uncertain")
REASON_CODES = (
    "fragment",
    "affiliation_or_entity",
    "redundant_fragment",
    "too_broad",
    "malformed",
    "other",
)

SCREENING_HEADERS = (
    "paper_id",
    "title",
    "abstract",
    "source",
    "external_id",
    "candidate_phrase",
    "triggering_heuristic",
    "source_hash",
    "provenance_count",
    "stored_keyword_version",
    "confirmed_failure",
    "reason_code",
    "reviewer_notes",
)
SCREENING_EVIDENCE_HEADERS = SCREENING_HEADERS[:10]

# Previously documented failures.  A listed ID only becomes locally evidenced
# when its actual Paper record is present in the local checkpoint scan; an ID
# in this set alone is never counted or sampled.
DOCUMENTED_FAILURE_EVIDENCE: dict[str, dict[str, str]] = {
    "8a30576ea1aebfbe9dd6e227a5c9427cf3040dff": {
        "evidence_path": "docs/architecture.md",
        "evidence_kind": "previously_documented_affiliation_failure",
    },
    "7bbe04578073b4afebeffaab4bbd42f5132afe6a": {
        "evidence_path": "docs/architecture.md",
        "evidence_kind": "previously_documented_affiliation_failure",
    },
}

_FRAGMENT_BOUNDARY_WORDS = frozenset({
    "causing", "boom", "fields", "introduced", "purely", "remains",
    "based", "from", "for", "with", "of", "the", "and", "unclear",
})
FRAGMENT_HEURISTIC_VERSION = "stored_keyword_fragment_boundary_v1"

# Deliberately narrow topic-string cues.  Bare ``detection`` is excluded:
# "vulnerability detection" is a security topic, not evidence of CV.  All
# matching buckets are exposed, so conflicts are never resolved by dict order.
_DOMAIN_BUCKET_KEYWORDS: dict[str, frozenset[str]] = {
    "security_software": frozenset({
        "security", "vulnerability", "software engineering", "privacy",
        "cyber attack", "malware", "intrusion detection",
    }),
    "computer_vision": frozenset({
        "computer vision", "image", "video", "object detection",
        "image detection", "segmentation", "visual recognition",
    }),
    "ml_nlp": frozenset({
        "language model", "nlp", "llm", "retrieval", "rag",
        "natural language", "transformer",
    }),
    "applied_domain": frozenset({
        "accelerator", "biomolecular", "clinical", "education", "student",
        "physics", "chemistry", "biology",
    }),
}
_DOMAIN_ORDER = tuple(_DOMAIN_BUCKET_KEYWORDS)
_URL_RE = re.compile(r"https?://\S+")


@dataclass(frozen=True)
class ProvenanceEntry:
    session_id: str
    topic: str
    stage: str


@dataclass
class LocalPaperRecord:
    paper: Paper
    provenance: list[ProvenanceEntry]
    stored_keyword_phrases: set[str]
    stored_keyword_count: int


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _canonical_bytes_for_hash(payload: Any) -> bytes:
    return json.dumps(payload, sort_keys=True, separators=(",", ":"), ensure_ascii=False).encode("utf-8")


def _sha256_payload(payload: Any) -> str:
    return hashlib.sha256(_canonical_bytes_for_hash(payload)).hexdigest()


def _file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _source_hash(title: str, abstract: str | None) -> str:
    blob = (title or "") + "\x00" + (abstract or "")
    return hashlib.sha256(blob.encode("utf-8")).hexdigest()


def _abstract_word_char_counts(abstract: str | None) -> tuple[int, int]:
    if not abstract:
        return 0, 0
    stripped = _URL_RE.sub(" ", abstract).strip()
    return len(stripped.split()), len(stripped)


def _usable_abstract(abstract: str | None) -> bool:
    if not abstract:
        return False
    words, chars = _abstract_word_char_counts(abstract)
    return chars >= _MIN_ABSTRACT_CHARS and words >= _MIN_ABSTRACT_WORDS


def _stress_signals(title: str, abstract: str | None, provenance: list[ProvenanceEntry]) -> dict[str, Any]:
    """Measured local-edge signals; the historical key name is retained so
    evidence remains easy to compare, but no absolute stress claim is made."""
    words, chars = _abstract_word_char_counts(abstract)
    text = abstract or ""
    alnum_space = sum(1 for char in text if char.isalnum() or char.isspace())
    punctuation_ratio = round(1 - (alnum_space / len(text)), 4) if text else 0.0
    acronym_tokens = re.findall(r"\b[A-Z]{2,6}\b", f"{title} {text}")
    acronym_density = round(len(acronym_tokens) / words, 4) if words else 0.0
    distinct_topics = {entry.topic.strip().lower() for entry in provenance if entry.topic}
    return {
        "abstract_word_count": words,
        "abstract_char_count": chars,
        "short_abstract": words > 0 and words < 60,
        "noisy_punctuation_ratio": punctuation_ratio,
        "acronym_density": acronym_density,
        "cross_topic_provenance_count": len(distinct_topics),
    }


def _looks_like_fragment(phrase: str) -> bool:
    tokens = phrase.strip().lower().split()
    return bool(tokens) and (tokens[0] in _FRAGMENT_BOUNDARY_WORDS or tokens[-1] in _FRAGMENT_BOUNDARY_WORDS)


def _triggering_heuristics(phrase: str) -> list[str]:
    return [FRAGMENT_HEURISTIC_VERSION] if _looks_like_fragment(phrase) else []


def _domain_bucket_classification(provenance: list[ProvenanceEntry]) -> dict[str, Any]:
    topics = " ".join(entry.topic.casefold() for entry in provenance if entry.topic)

    def contains_cue(cue: str) -> bool:
        return re.search(r"(?<!\w)" + re.escape(cue) + r"(?!\w)", topics) is not None

    matches = [
        bucket
        for bucket in _DOMAIN_ORDER
        if any(contains_cue(cue) for cue in _DOMAIN_BUCKET_KEYWORDS[bucket])
    ]
    if not matches:
        return {"domain_bucket_matches": [], "domain_bucket_guess": None, "domain_bucket_status": "unclassified"}
    if len(matches) == 1:
        return {"domain_bucket_matches": matches, "domain_bucket_guess": matches[0], "domain_bucket_status": "unambiguous"}
    return {"domain_bucket_matches": matches, "domain_bucket_guess": None, "domain_bucket_status": "ambiguous"}


def _domain_bucket_guess(provenance: list[ProvenanceEntry]) -> str | None:
    """Compatibility helper: only returns a genuinely unambiguous guess."""
    return _domain_bucket_classification(provenance)["domain_bucket_guess"]


def _external_id(paper: Paper) -> str:
    return paper.doi or paper.url or paper.paper_id


def _deduplicated_provenance(entries: list[ProvenanceEntry]) -> list[ProvenanceEntry]:
    seen: set[tuple[str, str, str]] = set()
    result: list[ProvenanceEntry] = []
    for entry in entries:
        key = (entry.session_id, entry.topic, entry.stage)
        if key not in seen:
            seen.add(key)
            result.append(entry)
    return result


def _collect_local_papers() -> dict[str, LocalPaperRecord]:
    """Collect local Paper records through production read APIs only."""
    records: dict[str, LocalPaperRecord] = {}

    def record_paper(paper: Paper, provenance: ProvenanceEntry) -> None:
        existing = records.get(paper.paper_id)
        if existing is None:
            existing = LocalPaperRecord(paper=paper, provenance=[], stored_keyword_phrases=set(), stored_keyword_count=0)
            records[paper.paper_id] = existing
        existing.provenance.append(provenance)
        existing.stored_keyword_phrases.update(keyword for keyword in paper.keywords if isinstance(keyword, str))
        existing.stored_keyword_count = max(existing.stored_keyword_count, len(paper.keywords))

    with sqlite_checkpointer(QA_CHECKPOINT_DB_PATH) as checkpointer:
        for summary in list_curation_sessions(checkpointer):
            session = load_curation_session(summary["session_id"], checkpointer)
            if session is None:
                continue
            provenance = ProvenanceEntry(summary["session_id"], session.topic, session.stage)
            for paper, _score in session.reserve:
                record_paper(paper, provenance)
            for paper in session.selected_papers:
                record_paper(paper, provenance)
            for entry in session.turn_history:
                for paper_dict, _score in entry.get("batch", []):
                    if paper_dict.get("paper_id"):
                        record_paper(Paper(**paper_dict), provenance)

    for record in records.values():
        record.provenance = _deduplicated_provenance(record.provenance)
    return records


def _documented_evidence_for(paper_id: str) -> dict[str, str] | None:
    configured = DOCUMENTED_FAILURE_EVIDENCE.get(paper_id)
    if configured is None:
        return None
    path = Path(__file__).resolve().parent.parent / configured["evidence_path"]
    if not path.is_file():
        return None
    return {
        **configured,
        "evidence_sha256": _file_sha256(path),
    }


def build_inventory() -> list[dict[str, Any]]:
    """Build a non-sensitive inventory; abstract and phrase values stay out."""
    records: list[dict[str, Any]] = []
    for paper_id, local in _collect_local_papers().items():
        paper = local.paper
        provenance = local.provenance
        heuristic_hit_count = sum(bool(_triggering_heuristics(phrase)) for phrase in local.stored_keyword_phrases)
        documented_evidence = _documented_evidence_for(paper_id)
        failure_status = (
            "known_failure" if documented_evidence else
            "suspected_failure" if heuristic_hit_count else
            None
        )
        domain = _domain_bucket_classification(provenance)
        records.append({
            "schema_version": INVENTORY_SCHEMA_VERSION,
            "paper_id": paper_id,
            "source": paper.source,
            "external_id": _external_id(paper),
            "title": paper.title,
            "usable_abstract": _usable_abstract(paper.abstract),
            "stored_keywords_present": local.stored_keyword_count > 0,
            "stored_keywords_count": local.stored_keyword_count,
            "stored_keyword_version": None,
            "stored_keyword_version_note": "unknown; no re-extraction performed",
            "source_hash": _source_hash(paper.title, paper.abstract),
            "provenance": [
                {"session_id": item.session_id, "topic": item.topic, "stage": item.stage}
                for item in provenance
            ],
            "stress_signals": _stress_signals(paper.title, paper.abstract, provenance),
            "failure_status": failure_status,
            "failure_evidence": documented_evidence,
            "heuristic_screening_hit_count": heuristic_hit_count,
            **domain,
            "inventoried_at": _now_iso(),
        })
    return sorted(records, key=lambda record: record["paper_id"])


def write_inventory(records: list[dict[str, Any]], path: Path | None = None) -> None:
    resolved = path if path is not None else INVENTORY_PATH
    resolved.parent.mkdir(parents=True, exist_ok=True)
    with resolved.open("w", encoding="utf-8") as handle:
        for record in records:
            handle.write(json.dumps(record, sort_keys=True) + "\n")


def load_inventory(path: Path | None = None) -> list[dict[str, Any]]:
    resolved = path if path is not None else INVENTORY_PATH
    with resolved.open("r", encoding="utf-8") as handle:
        return [json.loads(line) for line in handle if line.strip()]


def _source_record_payload(local: LocalPaperRecord) -> dict[str, Any]:
    paper = local.paper
    return {
        "paper_id": paper.paper_id,
        "title": paper.title,
        "abstract": paper.abstract,
        "source": paper.source,
        "external_id": _external_id(paper),
        "source_hash": _source_hash(paper.title, paper.abstract),
        "provenance_count": len(local.provenance),
        "stored_keyword_version": None,
    }


def build_screening_rows() -> tuple[list[dict[str, Any]], dict[str, str]]:
    """Return sensitive workbook rows and per-paper source-record hashes."""
    rows: list[dict[str, Any]] = []
    source_record_hashes: dict[str, str] = {}
    for paper_id, local in sorted(_collect_local_papers().items()):
        source = _source_record_payload(local)
        triggering_phrases = sorted(
            (
                (phrase, heuristic)
                for phrase in local.stored_keyword_phrases
                for heuristic in _triggering_heuristics(phrase)
            ),
            key=lambda item: (item[0].casefold(), item[0], item[1]),
        )
        if not triggering_phrases:
            continue
        source_record_hashes[paper_id] = _sha256_payload(source)
        for phrase, heuristic in triggering_phrases:
            rows.append({
                **source,
                "candidate_phrase": phrase,
                "triggering_heuristic": heuristic,
                "confirmed_failure": None,
                "reason_code": None,
                "reviewer_notes": None,
            })
    return rows, source_record_hashes


def _screening_evidence_payload(row: dict[str, Any]) -> dict[str, Any]:
    return {header: row.get(header) for header in SCREENING_EVIDENCE_HEADERS}


def _screening_manifest_payload(
    rows: list[dict[str, Any]], source_record_hashes: dict[str, str], generated_at: str,
) -> dict[str, Any]:
    return {
        "schema_version": SCREENING_MANIFEST_SCHEMA_VERSION,
        "selection_rule_version": SCREENING_SELECTION_RULE_VERSION,
        "generated_at": generated_at,
        "workbook_filename": SCREENING_WORKBOOK_PATH.name,
        "screening_sheet": "Failure screening",
        "row_count": len(rows),
        "paper_count": len(source_record_hashes),
        "allowed_confirmed_failure_values": list(CONFIRMED_FAILURE_VALUES),
        "allowed_reason_codes": list(REASON_CODES),
        "source_records": {
            paper_id: {"source_record_sha256": digest}
            for paper_id, digest in sorted(source_record_hashes.items())
        },
        "rows": [
            {
                "worksheet_row": index,
                "paper_id": row["paper_id"],
                "evidence_sha256": _sha256_payload(_screening_evidence_payload(row)),
            }
            for index, row in enumerate(rows, start=2)
        ],
    }


def _configure_screening_workbook(rows: list[dict[str, Any]]) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Failure screening"
    sheet.append(list(SCREENING_HEADERS))
    for row in rows:
        sheet.append([row.get(header) for header in SCREENING_HEADERS])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:M{max(1, sheet.max_row)}"
    widths = {"A": 44, "B": 40, "C": 90, "D": 16, "E": 42, "F": 34, "G": 34,
              "H": 66, "I": 18, "J": 22, "K": 20, "L": 24, "M": 46}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row_number in range(2, sheet.max_row + 1):
        sheet.row_dimensions[row_number].height = 72
        for column in ("B", "C", "F", "M"):
            sheet[f"{column}{row_number}"].alignment = Alignment(vertical="top", wrap_text=True)
        for column in ("K", "L", "M"):
            sheet[f"{column}{row_number}"].fill = PatternFill("solid", fgColor="DDEBF7")

    decision_validation = DataValidation(
        type="list", formula1='"' + ",".join(CONFIRMED_FAILURE_VALUES) + '"', allow_blank=True,
    )
    reason_validation = DataValidation(
        type="list", formula1='"' + ",".join(REASON_CODES) + '"', allow_blank=True,
    )
    decision_validation.error = "Choose yes, no, or uncertain."
    decision_validation.errorTitle = "Invalid decision"
    decision_validation.showErrorMessage = True
    reason_validation.error = "Choose a reason code from the list."
    reason_validation.errorTitle = "Invalid reason code"
    reason_validation.showErrorMessage = True
    sheet.add_data_validation(decision_validation)
    sheet.add_data_validation(reason_validation)
    if rows:
        decision_validation.add(f"K2:K{sheet.max_row}")
        reason_validation.add(f"L2:L{sheet.max_row}")
        sheet.conditional_formatting.add(
            f"K2:K{sheet.max_row}",
            FormulaRule(formula=["ISBLANK(K2)"], fill=PatternFill("solid", fgColor="FFF2CC")),
        )

    instructions = workbook.create_sheet("Instructions", 0)
    for column in "ABCDEFGHIJKLM":
        instructions.column_dimensions[column].width = 10
    instructions.sheet_view.showGridLines = False
    instruction_lines = [
        "K5B.1a heuristic keyword-failure screening",
        "Purpose: confirm whether each stored keyword phrase flagged by the offline heuristic is a genuine keyword failure.",
        "This is preparatory screening, not the formal blind K5 annotation task.",
        "For every row, choose confirmed_failure: yes, no, or uncertain.",
        "If yes, choose the best reason_code. For no/uncertain, reason_code may be left blank; notes are optional.",
        "Do not edit the prefilled evidence columns A:J, add/delete/reorder rows, rename sheets, or replace validation lists.",
        "Save this same workbook at the same path. The freeze gate validates the manifest, source hashes, and every prefilled row.",
    ]
    for row_number, line in enumerate(instruction_lines, start=1):
        displayed_line = line if row_number == 1 else "\n".join(textwrap.wrap(line, width=64))
        instructions.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=13)
        instructions.cell(row=row_number, column=1, value=displayed_line)
        line_count = displayed_line.count("\n") + 1
        instructions.row_dimensions[row_number].height = 30 if row_number == 1 else 20 * line_count
    instructions["A1"].fill = header_fill
    instructions["A1"].font = Font(name="Aptos", size=14, bold=True, color="FFFFFF")
    for row_number in range(1, len(instruction_lines) + 1):
        instructions[f"A{row_number}"].alignment = Alignment(wrap_text=True, vertical="center")
        if row_number != 1:
            instructions[f"A{row_number}"].font = Font(name="Aptos", size=11)
    instructions.freeze_panes = "A2"
    return workbook


def export_failure_screening(replace: bool = False) -> int:
    existing = [path for path in (SCREENING_WORKBOOK_PATH, SCREENING_MANIFEST_PATH) if path.exists()]
    if existing and not replace:
        print("refused: screening output already exists; pass --replace to overwrite deliberately.", file=sys.stderr)
        return 3

    rows, source_record_hashes = build_screening_rows()
    generated_at = _now_iso()
    manifest = _screening_manifest_payload(rows, source_record_hashes, generated_at)
    manifest["manifest_sha256"] = _sha256_payload(manifest)

    EVAL_WORKING_DIR.mkdir(parents=True, exist_ok=True)
    workbook = _configure_screening_workbook(rows)
    workbook.save(SCREENING_WORKBOOK_PATH)
    with SCREENING_MANIFEST_PATH.open("w", encoding="utf-8") as handle:
        json.dump(manifest, handle, sort_keys=True, indent=2)
        handle.write("\n")

    print(f"screening export ready: {len(source_record_hashes)} paper(s), {len(rows)} phrase row(s)")
    print(f"workbook: {SCREENING_WORKBOOK_PATH}")
    print(f"manifest: {SCREENING_MANIFEST_PATH}")
    return 0


def _read_screening_workbook_rows() -> tuple[list[dict[str, Any]], list[str]]:
    violations: list[str] = []
    try:
        workbook = load_workbook(SCREENING_WORKBOOK_PATH, data_only=False)
    except Exception as exc:
        return [], [f"screening workbook could not be opened: {type(exc).__name__}"]
    if "Failure screening" not in workbook.sheetnames:
        return [], ["screening workbook is missing the 'Failure screening' sheet"]
    sheet = workbook["Failure screening"]
    headers = tuple(cell.value for cell in sheet[1])
    if headers != SCREENING_HEADERS:
        violations.append("screening workbook headers do not match the current contract")
        return [], violations
    rows: list[dict[str, Any]] = []
    for values in sheet.iter_rows(min_row=2, max_col=len(SCREENING_HEADERS), values_only=True):
        if all(value is None for value in values):
            continue
        row = dict(zip(SCREENING_HEADERS, values))
        if any(isinstance(value, str) and value.startswith("=") for value in values):
            violations.append("screening workbook contains a formula in a screening row")
        rows.append(row)
    return rows, violations


def _load_screening_manifest() -> tuple[dict[str, Any] | None, list[str]]:
    if not SCREENING_MANIFEST_PATH.exists():
        return None, ["screening manifest is required"]
    try:
        return json.loads(SCREENING_MANIFEST_PATH.read_text(encoding="utf-8")), []
    except (OSError, json.JSONDecodeError) as exc:
        return None, [f"screening manifest could not be read: {type(exc).__name__}"]


def validate_failure_screening(
    *, require_complete: bool = True, inventory: list[dict[str, Any]] | None = None,
) -> tuple[list[str], set[str]]:
    """Validate the frozen screening snapshot without consulting live sessions.

    The workbook evidence is hash-bound by the screening manifest and resolved
    against the frozen inventory.  Current session additions/removals/changes
    are intentionally outside this validity decision; use
    :func:`audit_live_screening_drift` to report them informationally.
    """
    violations: list[str] = []
    confirmed_paper_ids: set[str] = set()
    if not SCREENING_WORKBOOK_PATH.exists():
        return ["screening workbook is required"], confirmed_paper_ids
    manifest, manifest_read_violations = _load_screening_manifest()
    if manifest is None:
        return manifest_read_violations, confirmed_paper_ids
    violations.extend(manifest_read_violations)

    if inventory is None:
        if not INVENTORY_PATH.exists():
            return ["manifest-bound inventory is required"], confirmed_paper_ids
        try:
            inventory = load_inventory()
        except (OSError, json.JSONDecodeError) as exc:
            return [f"manifest-bound inventory could not be read: {type(exc).__name__}"], confirmed_paper_ids
    violations.extend(_inventory_contract_violations(inventory))
    inventory_by_id = {record["paper_id"]: record for record in inventory}

    manifest_copy = dict(manifest)
    claimed_manifest_hash = manifest_copy.pop("manifest_sha256", None)
    if claimed_manifest_hash != _sha256_payload(manifest_copy):
        violations.append("screening manifest_sha256 mismatch")
    if manifest.get("schema_version") != SCREENING_MANIFEST_SCHEMA_VERSION:
        violations.append("screening manifest schema_version is not current")
    if manifest.get("selection_rule_version") != SCREENING_SELECTION_RULE_VERSION:
        violations.append("screening manifest selection_rule_version is not current")
    if manifest.get("workbook_filename") != SCREENING_WORKBOOK_PATH.name:
        violations.append("screening manifest workbook filename does not identify the canonical workbook")
    if manifest.get("screening_sheet") != "Failure screening":
        violations.append("screening manifest sheet name does not match the frozen contract")
    if manifest.get("allowed_confirmed_failure_values") != list(CONFIRMED_FAILURE_VALUES):
        violations.append("screening manifest decision values do not match the frozen contract")
    if manifest.get("allowed_reason_codes") != list(REASON_CODES):
        violations.append("screening manifest reason codes do not match the frozen contract")

    workbook_rows, workbook_violations = _read_screening_workbook_rows()
    violations.extend(workbook_violations)
    manifest_rows = manifest.get("rows", [])
    if len(workbook_rows) != manifest.get("row_count") or len(workbook_rows) != len(manifest_rows):
        violations.append("screening workbook row count does not match the manifest")

    workbook_paper_ids = {str(row.get("paper_id")) for row in workbook_rows}
    manifest_sources = manifest.get("source_records", {})
    if workbook_paper_ids != set(manifest_sources):
        violations.append("screening workbook paper set does not match manifest source records")

    source_payloads: dict[str, dict[str, Any]] = {}
    for index, row in enumerate(workbook_rows, start=2):
        paper_id = str(row.get("paper_id"))
        manifest_row = manifest_rows[index - 2] if index - 2 < len(manifest_rows) else {}
        if manifest_row.get("worksheet_row") != index or manifest_row.get("paper_id") != paper_id:
            violations.append(f"screening row {index} identity does not match the manifest")
        if manifest_row.get("evidence_sha256") != _sha256_payload(_screening_evidence_payload(row)):
            violations.append(f"screening row {index} prefilled evidence hash mismatch")

        if _source_hash(str(row.get("title") or ""), row.get("abstract")) != row.get("source_hash"):
            violations.append(f"screening row {index} source_hash does not match title/abstract evidence")
        inventory_record = inventory_by_id.get(paper_id)
        if inventory_record is None:
            violations.append(f"{paper_id}: screened paper is absent from the manifest-bound inventory")
        else:
            if inventory_record.get("source_hash") != row.get("source_hash"):
                violations.append(f"{paper_id}: screening source_hash does not match manifest-bound inventory")
            for field in ("title", "source", "external_id"):
                if inventory_record.get(field) != row.get(field):
                    violations.append(f"{paper_id}: screening {field} does not match manifest-bound inventory")

        source_payload = {header: row.get(header) for header in (
            "paper_id", "title", "abstract", "source", "external_id",
            "source_hash", "provenance_count", "stored_keyword_version",
        )}
        prior_payload = source_payloads.setdefault(paper_id, source_payload)
        if prior_payload != source_payload:
            violations.append(f"{paper_id}: repeated screening rows disagree on protected source evidence")

        decision = row.get("confirmed_failure")
        reason = row.get("reason_code")
        if require_complete and decision not in CONFIRMED_FAILURE_VALUES:
            violations.append(f"screening row {index} has no valid nonblank human decision")
        elif decision not in (None, "") and decision not in CONFIRMED_FAILURE_VALUES:
            violations.append(f"screening row {index} has an invalid human decision")
        if reason not in (None, "") and reason not in REASON_CODES:
            violations.append(f"screening row {index} has an invalid reason_code")
        if decision == "yes" and reason not in REASON_CODES:
            violations.append(f"screening row {index} confirmed yes without a valid reason_code")
        if decision == "yes":
            confirmed_paper_ids.add(paper_id)

    for paper_id, source_payload in source_payloads.items():
        expected_hash = manifest_sources.get(paper_id, {}).get("source_record_sha256")
        if expected_hash != _sha256_payload(source_payload):
            violations.append(f"{paper_id}: protected source-record hash does not match screening manifest")
    return violations, confirmed_paper_ids


def _screening_summary() -> dict[str, Any]:
    rows, _violations = _read_screening_workbook_rows()
    return {
        "row_count": len(rows),
        "paper_count": len({str(row.get("paper_id")) for row in rows}),
        "decision_counts": dict(Counter(row.get("confirmed_failure") for row in rows)),
        "reason_counts": dict(Counter(row.get("reason_code") for row in rows)),
    }


def _canonical_screening_violations(confirmed_ids: set[str]) -> list[str]:
    summary = _screening_summary()
    violations: list[str] = []
    if summary["row_count"] != EXPECTED_SCREENING_ROW_COUNT:
        violations.append(f"screening row count {summary['row_count']} != {EXPECTED_SCREENING_ROW_COUNT}")
    if summary["paper_count"] != EXPECTED_SCREENING_PAPER_COUNT:
        violations.append(f"screening paper count {summary['paper_count']} != {EXPECTED_SCREENING_PAPER_COUNT}")
    decision_counts = {key: value for key, value in summary["decision_counts"].items() if value}
    reason_counts = {key: value for key, value in summary["reason_counts"].items() if value}
    if decision_counts != EXPECTED_SCREENING_DECISIONS:
        violations.append(f"screening decision totals {decision_counts} != {EXPECTED_SCREENING_DECISIONS}")
    if reason_counts != EXPECTED_SCREENING_REASONS:
        violations.append(f"screening reason totals {reason_counts} != {EXPECTED_SCREENING_REASONS}")
    if reason_counts.get("redundant_fragment", 0):
        violations.append("screening still contains redundant_fragment")
    if len(confirmed_ids) < STRATA_TARGETS["confirmed_failure"]:
        violations.append("fewer than four distinct human-confirmed failure papers are available")
    return violations


def audit_live_screening_drift() -> dict[str, Any]:
    """Compare live heuristic records to the frozen manifest, informationally."""
    manifest, violations = _load_screening_manifest()
    if manifest is None:
        return {"status": "unavailable", "errors": violations}
    try:
        live_rows, live_source_hashes = build_screening_rows()
    except Exception as exc:
        return {"status": "unavailable", "errors": [type(exc).__name__]}
    frozen_sources = {
        paper_id: value.get("source_record_sha256")
        for paper_id, value in manifest.get("source_records", {}).items()
    }
    added = sorted(set(live_source_hashes) - set(frozen_sources))
    removed = sorted(set(frozen_sources) - set(live_source_hashes))
    changed = sorted(
        paper_id for paper_id in set(frozen_sources) & set(live_source_hashes)
        if frozen_sources[paper_id] != live_source_hashes[paper_id]
    )
    return {
        "status": "drift_observed" if added or removed or changed else "no_drift",
        "informational_only": True,
        "added_paper_ids": added,
        "removed_paper_ids": removed,
        "changed_paper_ids": changed,
        "frozen_phrase_row_count": manifest.get("row_count"),
        "live_phrase_row_count": len(live_rows),
    }


def _inventory_contract_violations(inventory: list[dict[str, Any]]) -> list[str]:
    violations: list[str] = []
    for record in inventory:
        paper_id = record.get("paper_id", "<missing>")
        if record.get("schema_version") != INVENTORY_SCHEMA_VERSION:
            violations.append(f"{paper_id}: inventory schema_version is not current")
        status = record.get("failure_status")
        if status == "known_failure" and not record.get("failure_evidence"):
            violations.append(f"{paper_id}: known_failure lacks hash-bound local evidence")
        if status == "suspected_failure" and not record.get("heuristic_screening_hit_count"):
            violations.append(f"{paper_id}: suspected_failure lacks a heuristic hit")
        matches = record.get("domain_bucket_matches")
        domain_status = record.get("domain_bucket_status")
        guess = record.get("domain_bucket_guess")
        if not isinstance(matches, list) or domain_status not in {"unclassified", "unambiguous", "ambiguous"}:
            violations.append(f"{paper_id}: invalid domain classification fields")
        elif domain_status == "unambiguous" and (len(matches) != 1 or guess != matches[0]):
            violations.append(f"{paper_id}: inconsistent unambiguous domain classification")
        elif domain_status != "unambiguous" and guess is not None:
            violations.append(f"{paper_id}: ambiguous/unclassified domain must not have a guess")
    return violations


def _local_diversity_group(record: dict[str, Any]) -> str:
    status = record.get("domain_bucket_status")
    if status == "unambiguous":
        return record.get("domain_bucket_guess") or "unclassified"
    return status or "unclassified"


def _round_robin_local_diversity(pool: list[dict[str, Any]], target: int, seed: int) -> list[dict[str, Any]]:
    groups: dict[str, list[dict[str, Any]]] = {}
    for record in pool:
        groups.setdefault(_local_diversity_group(record), []).append(record)
    group_order = [group for group in (*_DOMAIN_ORDER, "ambiguous", "unclassified") if group in groups]
    rng = Random(seed)
    for group in group_order:
        groups[group].sort(key=lambda record: record["paper_id"])
        rng.shuffle(groups[group])
    picked: list[dict[str, Any]] = []
    while len(picked) < target:
        progressed = False
        for group in group_order:
            if groups[group]:
                picked.append(groups[group].pop())
                progressed = True
                if len(picked) == target:
                    break
        if not progressed:
            break
    return picked


def select_sample(
    inventory: list[dict[str, Any]],
    seed: int = K5_RANDOM_SEED,
    confirmed_screening_ids: set[str] | None = None,
) -> tuple[list[dict[str, Any]], dict[str, list[str]], list[str], list[str]]:
    """Select the deterministic 4/4/2 bounded product-evaluation sample."""
    contract_violations = _inventory_contract_violations(inventory)
    if contract_violations:
        raise ValueError("inventory does not satisfy current contract: " + "; ".join(contract_violations[:3]))
    by_id = {record["paper_id"]: record for record in inventory}
    confirmed_screening_ids = confirmed_screening_ids or set()
    shortfalls: dict[str, list[str]] = {}
    confirmed_pool = sorted(
        paper_id for paper_id in confirmed_screening_ids
        if paper_id in by_id and by_id[paper_id].get("usable_abstract")
    )
    confirmed_target = STRATA_TARGETS["confirmed_failure"]
    if len(confirmed_pool) < confirmed_target:
        shortfalls["confirmed_failure"] = [
            f"requested {confirmed_target}, found {len(confirmed_pool)} human-confirmed usable paper(s)"
        ]
        confirmed_failure_ids = confirmed_pool
    else:
        confirmed_failure_ids = sorted(Random(seed).sample(confirmed_pool, confirmed_target))

    # Papers screened as genuine failures belong only to the failure pool,
    # including confirmed papers not drawn into the bounded four.
    used_ids = set(confirmed_screening_ids)
    eligible_edges = [
        record for record in inventory
        if record["usable_abstract"] and record["paper_id"] not in used_ids
    ]
    local_edge_ids: list[str] = []
    edge_subtypes: dict[str, str] = {}

    def pick_max(signal: str, minimum_exclusive: float) -> dict[str, Any] | None:
        pool = [record for record in eligible_edges if record["paper_id"] not in local_edge_ids]
        if not pool:
            return None
        winner = sorted(pool, key=lambda record: (-record["stress_signals"][signal], record["paper_id"]))[0]
        return winner if winner["stress_signals"][signal] > minimum_exclusive else None

    edge_pickers = (
        ("local_acronym_density_max", lambda: pick_max("acronym_density", 0.0)),
        ("local_punctuation_ratio_max", lambda: pick_max("noisy_punctuation_ratio", 0.0)),
        ("local_short_abstract", lambda: next((
            record for record in sorted(eligible_edges, key=lambda item: item["paper_id"])
            if record["paper_id"] not in local_edge_ids and record["stress_signals"]["short_abstract"]
        ), None)),
        ("local_cross_topic_provenance_max", lambda: pick_max("cross_topic_provenance_count", 1.0)),
    )
    edge_notes: list[str] = []
    for subtype, picker in edge_pickers:
        if len(local_edge_ids) >= STRATA_TARGETS["local_edge_case"]:
            break
        selected = picker()
        if selected is None:
            edge_notes.append(f"local edge subtype {subtype!r}: no qualifying local paper")
            continue
        local_edge_ids.append(selected["paper_id"])
        edge_subtypes[selected["paper_id"]] = subtype
    if len(local_edge_ids) < STRATA_TARGETS["local_edge_case"]:
        shortfalls["local_edge_case"] = edge_notes
    used_ids.update(local_edge_ids)

    diversity_pool = [
        record for record in inventory
        if record["usable_abstract"] and record["paper_id"] not in used_ids
    ]
    diversity_target = STRATA_TARGETS["product_local_diversity"]
    diversity_records = (
        sorted(Random(seed + 1).sample(sorted(diversity_pool, key=lambda item: item["paper_id"]), diversity_target),
               key=lambda item: item["paper_id"])
        if len(diversity_pool) >= diversity_target else sorted(diversity_pool, key=lambda item: item["paper_id"])
    )
    diversity_ids = [record["paper_id"] for record in diversity_records]
    if len(diversity_ids) < diversity_target:
        shortfalls["product_local_diversity"] = [
            f"requested {diversity_target}, found {len(diversity_ids)} usable unused local paper(s)"
        ]

    selected_rows: list[dict[str, Any]] = []

    def append_row(paper_id: str, stratum: str) -> None:
        source = by_id[paper_id]
        selected_rows.append({
            "paper_id": paper_id,
            "stratum": stratum,
            "local_edge_subtype": edge_subtypes.get(paper_id),
            "measured_signals": source["stress_signals"] if stratum == "local_edge_case" else None,
            "selection_basis": (
                "relative extreme within the frozen product-local inventory; not a universal threshold"
                if stratum == "local_edge_case" else None
            ),
            "confirmation_source": "human_screening" if stratum == "confirmed_failure" else None,
            "domain_bucket_matches": source["domain_bucket_matches"],
            "domain_bucket_guess": source["domain_bucket_guess"],
            "domain_bucket_status": source["domain_bucket_status"],
            "source_hash": source["source_hash"],
        })

    for paper_id in confirmed_failure_ids:
        append_row(paper_id, "confirmed_failure")
    for paper_id in diversity_ids:
        append_row(paper_id, "product_local_diversity")
    for paper_id in local_edge_ids:
        append_row(paper_id, "local_edge_case")

    final_strata = ("confirmed_failure", "product_local_diversity", "local_edge_case")

    def spread_pick(count: int, excluded: set[str]) -> list[str]:
        pools = {
            stratum: sorted(row["paper_id"] for row in selected_rows if row["stratum"] == stratum and row["paper_id"] not in excluded)
            for stratum in final_strata
        }
        result: list[str] = []
        while len(result) < count:
            progressed = False
            for stratum in final_strata:
                if pools[stratum]:
                    result.append(pools[stratum].pop(0))
                    progressed = True
                    if len(result) == count:
                        break
            if not progressed:
                break
        return result

    pilot_ids = spread_pick(PILOT_COUNT, set())
    double_review_ids = spread_pick(DOUBLE_REVIEW_COUNT, set(pilot_ids))
    for row in selected_rows:
        row["metrics_role"] = "pilot_only" if row["paper_id"] in pilot_ids else "headline"
        row["is_double_reviewed"] = row["paper_id"] in double_review_ids
    return selected_rows, shortfalls, pilot_ids, double_review_ids


def _jsonl_bytes(records: list[dict[str, Any]]) -> bytes:
    return "".join(json.dumps(record, sort_keys=True) + "\n" for record in records).encode("utf-8")


def _resolve_source_snapshot(
    selected: list[dict[str, Any]], snapshot_at: str,
) -> tuple[list[dict[str, Any]], list[str]]:
    live_records = _collect_local_papers()
    violations: list[str] = []
    snapshot: list[dict[str, Any]] = []
    for selected_row in sorted(selected, key=lambda item: item["paper_id"]):
        paper_id = selected_row["paper_id"]
        local = live_records.get(paper_id)
        if local is None:
            violations.append(f"{paper_id}: selected source record cannot be resolved")
            continue
        paper = local.paper
        actual_hash = _source_hash(paper.title, paper.abstract)
        if actual_hash != selected_row["source_hash"]:
            violations.append(f"{paper_id}: resolved source hash does not match frozen inventory")
            continue
        snapshot.append({
            "schema_version": SOURCE_SNAPSHOT_SCHEMA_VERSION,
            "paper_id": paper_id,
            "title": paper.title,
            "abstract": paper.abstract,
            "source": paper.source,
            "external_id": _external_id(paper),
            "source_hash": actual_hash,
            "snapshotted_at": snapshot_at,
        })
    return snapshot, violations


def _write_atomic(path: Path, data: bytes) -> None:
    temporary = path.with_name(f".{path.name}.tmp")
    temporary.write_bytes(data)
    os.replace(temporary, path)


def freeze_sample(replace: bool = False) -> int:
    if not INVENTORY_PATH.exists():
        print("error: current inventory is missing; run inventory first.", file=sys.stderr)
        return 1
    if any(path.exists() for path in (SAMPLE_PATH, RULES_PATH, SOURCE_SNAPSHOT_PATH)) and not replace:
        print("refused: sample output exists; pass --replace to overwrite deliberately.", file=sys.stderr)
        return 3

    inventory = load_inventory()
    contract_violations = _inventory_contract_violations(inventory)
    if contract_violations:
        print("refused: inventory does not satisfy the current K5B.1a contract.", file=sys.stderr)
        return 4

    screening_violations, confirmed_ids = validate_failure_screening(require_complete=True, inventory=inventory)
    screening_violations.extend(_canonical_screening_violations(confirmed_ids))
    if screening_violations:
        print("refused: frozen screening snapshot failed validation; sample artifacts were not replaced.", file=sys.stderr)
        return 4
    selected, shortfalls, pilot_ids, double_review_ids = select_sample(
        inventory,
        seed=K5_RANDOM_SEED,
        confirmed_screening_ids=confirmed_ids,
    )
    if shortfalls or len(selected) != sum(STRATA_TARGETS.values()):
        print("refused: bounded sample targets cannot be satisfied; sample artifacts were not replaced.", file=sys.stderr)
        return 4
    selected_at = _now_iso()

    source_snapshot, snapshot_violations = _resolve_source_snapshot(selected, selected_at)
    if snapshot_violations or len(source_snapshot) != sum(STRATA_TARGETS.values()):
        print("refused: a selected paper could not be resolved with its expected source hash; nothing was replaced.", file=sys.stderr)
        return 4

    strata_actual = dict(Counter(row["stratum"] for row in selected))
    observed_domain_composition = dict(Counter(
        _local_diversity_group(row)
        for row in selected
        if row["stratum"] == "product_local_diversity"
    ))
    for row in selected:
        row["schema_version"] = SAMPLE_SCHEMA_VERSION
        row["selection_rule_version"] = SELECTION_RULE_VERSION
        row["status"] = "frozen_bounded_sample"
        row["selected_at"] = selected_at
    selected_sorted = sorted(selected, key=lambda item: item["paper_id"])
    sample_bytes = _jsonl_bytes(selected_sorted)
    snapshot_bytes = _jsonl_bytes(source_snapshot)
    screening_summary = _screening_summary()
    limitations = [
        "product-local corpus",
        "only 10 papers",
        "pilot papers excluded from headline metrics",
        "not an external benchmark",
        "not evidence of universal superiority",
    ]
    rules = {
        "schema_version": SAMPLE_SCHEMA_VERSION,
        "selection_rule_version": SELECTION_RULE_VERSION,
        "status": "frozen_bounded_sample",
        "random_seed": K5_RANDOM_SEED,
        "selection_timestamp": selected_at,
        "strata_targets": STRATA_TARGETS,
        "strata_actual": strata_actual,
        "observed_product_local_domain_composition": observed_domain_composition,
        "total_selected": len(selected),
        "pilot_ids": sorted(pilot_ids),
        "double_review_ids": sorted(double_review_ids),
        "paper_ids": sorted(row["paper_id"] for row in selected),
        "source_hashes": {row["paper_id"]: row["source_hash"] for row in selected},
        "sample_sha256": hashlib.sha256(sample_bytes).hexdigest(),
        "screening_workbook_sha256": _file_sha256(SCREENING_WORKBOOK_PATH),
        "screening_manifest_sha256": _file_sha256(SCREENING_MANIFEST_PATH),
        "inventory_sha256": _file_sha256(INVENTORY_PATH),
        "bounded_source_snapshot_sha256": hashlib.sha256(snapshot_bytes).hexdigest(),
        "screening_snapshot": screening_summary,
        "informational_live_drift": audit_live_screening_drift(),
        "limitations": limitations,
    }
    rules["manifest_sha256"] = _sha256_payload(rules)
    EVAL_WORKING_DIR.mkdir(parents=True, exist_ok=True)
    rules_bytes = (json.dumps(rules, sort_keys=True, indent=2) + "\n").encode("utf-8")
    _write_atomic(SOURCE_SNAPSHOT_PATH, snapshot_bytes)
    _write_atomic(SAMPLE_PATH, sample_bytes)
    _write_atomic(RULES_PATH, rules_bytes)
    print(f"sample status: frozen_bounded_sample; selected {len(selected)} paper(s)")
    return 0


def validate_frozen_sample() -> list[str]:
    violations: list[str] = []
    if not RULES_PATH.exists() or not SAMPLE_PATH.exists() or not SOURCE_SNAPSHOT_PATH.exists():
        return ["no sample found"]
    try:
        rules = json.loads(RULES_PATH.read_text(encoding="utf-8"))
        sample = [json.loads(line) for line in SAMPLE_PATH.read_text(encoding="utf-8").splitlines() if line.strip()]
        source_snapshot = [
            json.loads(line) for line in SOURCE_SNAPSHOT_PATH.read_text(encoding="utf-8").splitlines() if line.strip()
        ]
        inventory = load_inventory()
    except (OSError, json.JSONDecodeError) as exc:
        return [f"sample evidence could not be read: {type(exc).__name__}"]

    violations.extend(_inventory_contract_violations(inventory))
    if rules.get("schema_version") != SAMPLE_SCHEMA_VERSION:
        violations.append("sample rules schema_version is not current")
    if rules.get("selection_rule_version") != SELECTION_RULE_VERSION:
        violations.append("sample rules selection_rule_version is not current")
    rules_copy = dict(rules)
    claimed_hash = rules_copy.pop("manifest_sha256", None)
    if claimed_hash != _sha256_payload(rules_copy):
        violations.append("sample manifest_sha256 mismatch")
    if rules.get("status") != "frozen_bounded_sample":
        violations.append("sample status is not frozen_bounded_sample")
    if rules.get("strata_targets") != STRATA_TARGETS:
        violations.append("sample strata targets do not match bounded contract")
    if rules.get("total_selected") != sum(STRATA_TARGETS.values()):
        violations.append("sample total does not match bounded contract")
    hash_bindings = (
        ("sample_sha256", SAMPLE_PATH, "sample file"),
        ("bounded_source_snapshot_sha256", SOURCE_SNAPSHOT_PATH, "bounded source snapshot"),
        ("screening_workbook_sha256", SCREENING_WORKBOOK_PATH, "screening workbook"),
        ("screening_manifest_sha256", SCREENING_MANIFEST_PATH, "screening manifest file"),
        ("inventory_sha256", INVENTORY_PATH, "inventory"),
    )
    for field, path, label in hash_bindings:
        if not path.exists():
            violations.append(f"{label} is missing")
        elif rules.get(field) != _file_sha256(path):
            violations.append(f"{label} hash does not match rules manifest")

    ids = [row.get("paper_id") for row in sample]
    if len(ids) != len(set(ids)):
        violations.append("duplicate paper_id in sample")
    if sorted(ids) != rules.get("paper_ids"):
        violations.append("sample paper IDs do not match rules manifest")
    inventory_by_id = {record["paper_id"]: record for record in inventory}
    if len(ids) != sum(STRATA_TARGETS.values()):
        violations.append("sample does not contain exactly 10 rows")
    actual_strata = dict(Counter(row.get("stratum") for row in sample))
    if actual_strata != STRATA_TARGETS or rules.get("strata_actual") != STRATA_TARGETS:
        violations.append("sample does not have exact 4/4/2 composition")
    valid_strata = set(STRATA_TARGETS)
    screening_violations, confirmed_ids = validate_failure_screening(require_complete=True, inventory=inventory)
    screening_violations.extend(_canonical_screening_violations(confirmed_ids))
    violations.extend(f"frozen screening: {item}" for item in screening_violations)

    for row in sample:
        paper_id = row.get("paper_id")
        if row.get("schema_version") != SAMPLE_SCHEMA_VERSION or row.get("selection_rule_version") != SELECTION_RULE_VERSION:
            violations.append(f"{paper_id}: sample row uses an obsolete contract")
        if row.get("stratum") not in valid_strata:
            violations.append(f"{paper_id}: invalid stratum")
        source = inventory_by_id.get(paper_id)
        if source is None:
            violations.append(f"{paper_id}: absent from current inventory")
        elif source.get("source_hash") != row.get("source_hash"):
            violations.append(f"{paper_id}: source_hash does not match inventory")
        if row.get("status") != "frozen_bounded_sample":
            violations.append(f"{paper_id}: row status is not frozen_bounded_sample")
        if row.get("stratum") == "confirmed_failure" and paper_id not in confirmed_ids:
            violations.append(f"{paper_id}: confirmed_failure is not human-confirmed")
        if row.get("stratum") == "local_edge_case" and not row.get("measured_signals"):
            violations.append(f"{paper_id}: local_edge_case lacks measured signals")
        forbidden = {"keywords", "candidates", "abstract", "abstract_text"} & set(row)
        if forbidden:
            violations.append(f"{paper_id}: forbidden embedded candidate/source text")

    snapshot_ids = [row.get("paper_id") for row in source_snapshot]
    if len(snapshot_ids) != len(set(snapshot_ids)) or sorted(snapshot_ids) != sorted(ids):
        violations.append("bounded source snapshot IDs do not exactly match sample IDs")
    sample_by_id = {row["paper_id"]: row for row in sample}
    for row in source_snapshot:
        paper_id = row.get("paper_id")
        if row.get("schema_version") != SOURCE_SNAPSHOT_SCHEMA_VERSION:
            violations.append(f"{paper_id}: source snapshot row uses an obsolete contract")
        actual_hash = _source_hash(str(row.get("title") or ""), row.get("abstract"))
        if actual_hash != row.get("source_hash") or sample_by_id.get(paper_id, {}).get("source_hash") != actual_hash:
            violations.append(f"{paper_id}: bounded source snapshot source_hash mismatch")

    pilot_ids = set(rules.get("pilot_ids", []))
    double_review_ids = set(rules.get("double_review_ids", []))
    if len(pilot_ids) != PILOT_COUNT or not pilot_ids <= set(ids):
        violations.append("pilot role count or membership is invalid")
    if len(double_review_ids) != DOUBLE_REVIEW_COUNT or not double_review_ids <= set(ids):
        violations.append("double-review role count or membership is invalid")
    if pilot_ids & double_review_ids:
        violations.append("pilot and double-review IDs overlap")
    by_sample_id = {row["paper_id"]: row for row in sample}
    if any(by_sample_id[paper_id].get("metrics_role") != "pilot_only" for paper_id in pilot_ids):
        violations.append("pilot metrics roles are inconsistent")
    if any(by_sample_id[paper_id].get("metrics_role") != "headline" for paper_id in double_review_ids):
        violations.append("double-review papers must be headline papers")
    if any(not by_sample_id[paper_id].get("is_double_reviewed") for paper_id in double_review_ids):
        violations.append("double-review flags are inconsistent")
    if len({by_sample_id[paper_id]["stratum"] for paper_id in pilot_ids}) < 2:
        violations.append("pilot papers do not span at least two strata")
    if len({by_sample_id[paper_id]["stratum"] for paper_id in double_review_ids}) < 2:
        violations.append("double-review papers do not span at least two strata")

    observed = dict(Counter(
        _local_diversity_group(row) for row in sample if row.get("stratum") == "product_local_diversity"
    ))
    if rules.get("observed_product_local_domain_composition") != observed:
        violations.append("observed product-local domain composition is inconsistent")
    expected_limitations = {
        "product-local corpus", "only 10 papers", "pilot papers excluded from headline metrics",
        "not an external benchmark", "not evidence of universal superiority",
    }
    if set(rules.get("limitations", [])) != expected_limitations:
        violations.append("bounded-sample limitations are incomplete")
    return violations


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="K5B.1 screening and bounded-sample tooling.")
    subparsers = parser.add_subparsers(dest="command", required=True)
    subparsers.add_parser("inventory")
    screening_parser = subparsers.add_parser("export-failure-screening")
    screening_parser.add_argument("--replace", action="store_true")
    freeze_parser = subparsers.add_parser("freeze-sample")
    freeze_parser.add_argument("--replace", action="store_true")
    subparsers.add_parser("audit-live-drift")
    subparsers.add_parser("validate")
    args = parser.parse_args(argv)

    if args.command == "inventory":
        records = build_inventory()
        write_inventory(records)
        print(f"inventoried {len(records)} unique local paper(s) -> {INVENTORY_PATH}")
        return 0
    if args.command == "export-failure-screening":
        return export_failure_screening(replace=args.replace)
    if args.command == "freeze-sample":
        return freeze_sample(replace=args.replace)
    if args.command == "audit-live-drift":
        audit = audit_live_screening_drift()
        print(json.dumps(audit, sort_keys=True))
        return 0 if audit.get("status") != "unavailable" else 1
    if args.command == "validate":
        violations = validate_frozen_sample()
        if violations:
            for violation in violations:
                print(f"VIOLATION: {violation}", file=sys.stderr)
            return 1
        print("valid: sample evidence passes the current contract.")
        return 0
    return 2


if __name__ == "__main__":
    raise SystemExit(main())
