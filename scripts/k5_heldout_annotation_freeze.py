#!/usr/bin/env python3
"""K5D.1b: freeze the corrected, human-approved held-out annotations.

Reads only the already-validated K5D.1a workbook/mapping/manifest (plus
the frozen held-out selection, for prior-K5-evidence identity binding).
Never calls a provider, never runs Policy C, never computes metrics.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import os
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from scripts import k5_heldout_selection as k5d1

BASE = k5d1.K5D1_DIR
FROZEN_PATH = BASE / "heldout_annotation_frozen.json"
FROZEN_SCHEMA = "k5d1b-frozen-heldout-annotation-v1"
REVIEWER_TYPE = "ai_assisted_human_approved"


def now() -> str:
    return datetime.now(timezone.utc).isoformat()


def canon(value: Any) -> bytes:
    return json.dumps(value, sort_keys=True, separators=(",", ":"), ensure_ascii=False).encode("utf-8")


def payload_hash(value: Any) -> str:
    return hashlib.sha256(canon(value)).hexdigest()


def file_hash(path: Path) -> str:
    return k5d1._file_sha256(path)


def json_bytes(value: Any) -> bytes:
    return (json.dumps(value, sort_keys=True, indent=2, ensure_ascii=False) + "\n").encode("utf-8")


def atomic_write(path: Path, data: bytes) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.tmp")
    temporary.write_bytes(data)
    os.replace(temporary, path)


def self_hash_valid(value: dict[str, Any], field: str) -> bool:
    copy = dict(value)
    claimed = copy.pop(field, None)
    return isinstance(claimed, str) and claimed == payload_hash(copy)


def validate_frozen_payload(data: dict[str, Any]) -> list[str]:
    errors: list[str] = []
    if data.get("schema_version") != FROZEN_SCHEMA:
        errors.append("frozen schema mismatch")
    if not self_hash_valid(data, "frozen_annotation_sha256"):
        errors.append("frozen annotation self-hash mismatch")
    if data.get("reviewer_type") != REVIEWER_TYPE:
        errors.append("reviewer provenance mismatch")

    papers = data.get("papers", [])
    candidates = data.get("candidates", [])
    if len(papers) != k5d1.HELD_OUT_COUNT or len({p.get("paper_code") for p in papers}) != k5d1.HELD_OUT_COUNT:
        errors.append("frozen paper count/identity mismatch")
    if len(candidates) != 36 or len({c.get("candidate_id") for c in candidates}) != 36:
        errors.append("frozen candidate count/identity mismatch")

    concepts_by_code = {paper["paper_code"]: paper["concepts"] for paper in papers if "paper_code" in paper}
    for paper in papers:
        nonblank = sum(value not in (None, "") for value in paper.get("concepts", []))
        if not 3 <= nonblank <= 5:
            errors.append(f"{paper.get('paper_code')}: concept completeness")
    for row in candidates:
        decision = row.get("decision")
        reason = row.get("rejection_reason")
        matched_ids, matched_valid = k5d1._parse_concept_ids(row.get("matched_concept_ids"))
        if decision not in k5d1.DECISION_VALUES or row.get("confidence") not in k5d1.CONFIDENCE_VALUES:
            errors.append(f"{row.get('candidate_id')}: incomplete")
        if bool(reason) != (decision == "reject") or not matched_valid:
            errors.append(f"{row.get('candidate_id')}: semantic mismatch")
        if decision == "accept" and not matched_ids:
            errors.append(f"{row.get('candidate_id')}: accept requires at least one matched concept id")
        if decision == "reject" and matched_ids:
            errors.append(f"{row.get('candidate_id')}: reject must not carry matched concept ids")
        concepts = concepts_by_code.get(row.get("paper_code"), [])
        for concept_id in matched_ids:
            index = int(concept_id[1]) - 1
            if index >= len(concepts) or not concepts[index]:
                errors.append(f"{row.get('candidate_id')}: {concept_id} does not identify a populated concept")
    return errors


def _read_workbook_papers_and_candidates() -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    workbook = load_workbook(k5d1.WORKBOOK_PATH, data_only=False)
    concepts_sheet = workbook["Paper concepts"]
    review_sheet = workbook["Candidate review"]
    papers = []
    for row in range(2, concepts_sheet.max_row + 1):
        papers.append({
            "paper_code": concepts_sheet.cell(row, 1).value,
            "concepts": [concepts_sheet.cell(row, column).value for column in range(4, 9)],
            "reviewer_notes": concepts_sheet.cell(row, 9).value,
        })
    candidates = []
    for row in range(2, review_sheet.max_row + 1):
        candidates.append({
            "candidate_id": review_sheet.cell(row, 1).value,
            "paper_code": review_sheet.cell(row, 2).value,
            "candidate_phrase": review_sheet.cell(row, 3).value,
            "decision": review_sheet.cell(row, 4).value,
            "rejection_reason": review_sheet.cell(row, 5).value,
            "matched_concept_ids": review_sheet.cell(row, 6).value,
            "confidence": review_sheet.cell(row, 7).value,
            "reviewer_notes": review_sheet.cell(row, 8).value,
        })
    return papers, candidates


def freeze_annotations(replace: bool = False) -> dict[str, Any]:
    if FROZEN_PATH.exists() and not replace:
        raise FileExistsError("frozen held-out annotation exists; pass --replace")

    errors = k5d1.validate_annotation_workbook(
        require_all_complete=True,
        workbook_path=k5d1.WORKBOOK_PATH, mapping_path=k5d1.MAPPING_PATH, manifest_path=k5d1.MANIFEST_PATH,
    )
    if errors:
        raise ValueError("; ".join(errors))

    selected_ids, selection = k5d1.load_frozen_selection()
    mapping = json.loads(k5d1.MAPPING_PATH.read_text(encoding="utf-8"))
    if mapping.get("extractor_version") != k5d1.production_keywords.KEYWORD_EXTRACTOR_VERSION:
        raise ValueError("production extractor version has drifted since candidate generation")

    papers, candidates = _read_workbook_papers_and_candidates()

    data = {
        "schema_version": FROZEN_SCHEMA,
        "status": "frozen_complete_heldout_annotation",
        "frozen_at": now(),
        "reviewer_type": REVIEWER_TYPE,
        "held_out_count": k5d1.HELD_OUT_COUNT,
        "paper_codes": [paper["paper_code"] for paper in papers],
        "extractor_version": mapping["extractor_version"],
        "bindings": {
            "workbook_sha256": file_hash(k5d1.WORKBOOK_PATH),
            "candidate_mapping_sha256": file_hash(k5d1.MAPPING_PATH),
            "annotation_manifest_sha256": file_hash(k5d1.MANIFEST_PATH),
            "selection_sha256": file_hash(k5d1.SELECTION_PATH),
            "source_snapshot_sha256": file_hash(k5d1.SOURCE_SNAPSHOT_PATH),
        },
        "prior_k5_evidence": selection["prior_k5_sample"],
        "papers": papers,
        "candidates": candidates,
    }
    data["frozen_annotation_sha256"] = payload_hash(data)

    errors = validate_frozen_payload(data)
    if errors:
        raise ValueError("; ".join(errors))
    atomic_write(FROZEN_PATH, json_bytes(data))
    return data


def validate_reproducibility(data: dict[str, Any] | None = None) -> list[str]:
    """Prove every frozen candidate/paper row is an exact readback of the
    canonical workbook -- not a transcription drift."""
    if data is None:
        if not FROZEN_PATH.exists():
            return ["frozen held-out annotation is missing"]
        data = json.loads(FROZEN_PATH.read_text(encoding="utf-8"))
    errors = validate_frozen_payload(data)
    if errors:
        return errors
    if data.get("bindings", {}).get("workbook_sha256") != file_hash(k5d1.WORKBOOK_PATH):
        return ["frozen workbook binding no longer matches the canonical workbook on disk"]
    papers, candidates = _read_workbook_papers_and_candidates()
    if data.get("papers") != papers:
        errors.append("frozen papers are not an exact readback of the canonical workbook")
    if data.get("candidates") != candidates:
        errors.append("frozen candidates are not an exact readback of the canonical workbook")
    return errors


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("command", choices=("freeze", "validate"))
    parser.add_argument("--replace", action="store_true")
    args = parser.parse_args(argv)
    try:
        if args.command == "freeze":
            data = freeze_annotations(replace=args.replace)
            print(f"frozen held-out annotation: papers={len(data['papers'])} candidates={len(data['candidates'])}")
        else:
            data = json.loads(FROZEN_PATH.read_text(encoding="utf-8"))
            errors = validate_frozen_payload(data) + validate_reproducibility(data)
            if errors:
                for error in errors:
                    print(f"VIOLATION: {error}", file=sys.stderr)
                return 1
            print("valid: frozen held-out annotation reproduces the canonical workbook exactly.")
        return 0
    except Exception as exc:  # noqa: BLE001 -- concise offline CLI failure.
        print(f"error: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
