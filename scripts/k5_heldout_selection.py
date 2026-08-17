#!/usr/bin/env python3
"""K5D.1a: minimal held-out validation set preparation for the frozen,
narrow LLM keyword filter policy (Policy C, defined and evaluated
entirely outside this module -- see ``scripts/k5_llm_filter_policy_analysis.py``).

This checkpoint prepares blind human annotation only. It never calls a
provider, never runs Policy C, never computes precision/coverage, and
never touches ``research_agent/keywords.py``.

Three stages, each consuming only the frozen output of the one before it:

1. ``validate-prior``  -- read-only proof that the existing K5B/K5C/K5C.1
   evidence is intact, plus the ten prior paper IDs that the new sample
   must be disjoint from. Touches no existing artifact.
2. ``select``           -- freezes exactly six new held-out paper IDs
   (H01-H06) and their source snapshot, selected without consulting any
   human label, LLM decision, or documented failure phrase.
3. ``generate``         -- reads only the frozen selection + source
   snapshot (never live sessions) to run the production YAKE-v2
   extractor locally and build the blank annotation workbook.

Held-out artifacts live under ``eval_working/paper_keywords/k5d1/`` --
gitignored, disjoint from every K5B/K5C path.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from random import Random
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import TYPE_ERROR, TYPE_FORMULA
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from research_agent import keywords as production_keywords
from scripts import k5_keyword_annotation as annotation
from scripts import k5_keyword_metrics as k5b_metrics
from scripts import k5_llm_filter_eval as k5c
from scripts import k5_llm_filter_policy_analysis as k5c1
from scripts import k5_paper_inventory


K5D1_DIR = annotation.EVAL_WORKING_DIR / "k5d1"
SELECTION_PATH = K5D1_DIR / "heldout_selection.json"
SOURCE_SNAPSHOT_PATH = K5D1_DIR / "heldout_source_snapshot.jsonl"
MAPPING_PATH = K5D1_DIR / "heldout_candidate_mapping.json"
MANIFEST_PATH = K5D1_DIR / "heldout_annotation_manifest.json"
WORKBOOK_PATH = K5D1_DIR / "heldout_keyword_annotation.xlsx"

SELECTION_SCHEMA_VERSION = "k5d1a-heldout-selection-v1"
SOURCE_SNAPSHOT_SCHEMA_VERSION = "k5d1a-heldout-source-v1"
MAPPING_SCHEMA_VERSION = "k5d1a-candidate-mapping-v1"
MANIFEST_SCHEMA_VERSION = "k5d1a-annotation-manifest-v1"

# Documented, distinct from every K5B/K5C seed (5202608, 5202610). Only
# scopes *which* six eligible papers are drawn -- extractor candidate
# order within a paper is never shuffled (see generate_candidate_bundle).
K5D1_RANDOM_SEED = 5202612
# Scopes only the opaque candidate-ID hash, not selection or ordering.
CANDIDATE_ID_SEED = 5202613
SELECTION_RULE_VERSION = "k5d1a-heldout-v1"
HELD_OUT_COUNT = 6
MAX_CANDIDATES_PER_PAPER = 6

PAPER_HEADERS = (
    "paper_code", "title", "abstract",
    "concept_1", "concept_2", "concept_3", "concept_4", "concept_5",
    "reviewer_notes",
)
CANDIDATE_HEADERS = (
    "candidate_id", "paper_code", "candidate_phrase",
    "decision", "rejection_reason", "matched_concept_ids", "confidence", "reviewer_notes",
)
PAPER_EVIDENCE_HEADERS = PAPER_HEADERS[:3]
CANDIDATE_EVIDENCE_HEADERS = CANDIDATE_HEADERS[:3]

DECISION_VALUES = ("accept", "reject", "uncertain")
REJECTION_REASON_VALUES = (
    "fragment", "malformed", "too_broad", "affiliation_or_entity", "redundant_fragment", "other",
)
CONFIDENCE_VALUES = ("high", "medium", "low")
EXPECTED_SHEETS = ("Instructions", "Paper concepts", "Candidate review")

HEADER_FILL = PatternFill("solid", fgColor="4A1F5D")
HEADER_FONT = Font(color="FFFFFF", bold=True)
EVIDENCE_FILL = PatternFill("solid", fgColor="EDE7F6")
EDITABLE_FILL = PatternFill("solid", fgColor="FFF2CC")
SECTION_FILL = PatternFill("solid", fgColor="E1D5F0")


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _canonical_bytes(payload: Any) -> bytes:
    return json.dumps(payload, sort_keys=True, separators=(",", ":"), ensure_ascii=False).encode("utf-8")


def _sha256_payload(payload: Any) -> str:
    return hashlib.sha256(_canonical_bytes(payload)).hexdigest()


def _file_sha256(path: Path) -> str:
    return annotation._file_sha256(path)


def _json_bytes(payload: dict[str, Any]) -> bytes:
    return (json.dumps(payload, sort_keys=True, indent=2, ensure_ascii=False) + "\n").encode("utf-8")


def _write_atomic(path: Path, data: bytes) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.tmp")
    temporary.write_bytes(data)
    os.replace(temporary, path)


def _manifest_hash_valid(payload: dict[str, Any], field: str) -> bool:
    copy = dict(payload)
    claimed = copy.pop(field, None)
    return isinstance(claimed, str) and claimed == _sha256_payload(copy)


def _load_jsonl(path: Path) -> list[dict[str, Any]]:
    return [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]


# ---------------------------------------------------------------------------
# Part A -- read-only validation of existing K5B/K5C/K5C.1 evidence
# ---------------------------------------------------------------------------

def prior_evidence_bindings() -> dict[str, str]:
    """SHA-256 bindings that pin down every frozen K5B/K5C/K5C.1 artifact
    this checkpoint must prove the new sample is disjoint from."""
    bindings = dict(k5c1.input_bindings())
    bindings["k5c1_analysis_sha256"] = _file_sha256(k5c1.ANALYSIS_PATH)
    return bindings


def validate_prior_evidence() -> tuple[dict[str, str], list[str]]:
    """Read-only proof the frozen K5B/K5C/K5C.1 evidence is intact.

    Never writes, regenerates, or mutates any existing artifact; only
    reads and hashes what is already on disk.
    """
    k5c1.validate_inputs()  # validates K5B frozen annotation/results + K5C prompt/raw/metrics bindings
    analysis = json.loads(k5c1.ANALYSIS_PATH.read_text(encoding="utf-8"))
    analysis_errors = k5c1.validate_analysis(analysis)
    if analysis_errors:
        raise ValueError("K5C.1 analysis invalid: " + "; ".join(analysis_errors))

    rules = json.loads(annotation.RULES_PATH.read_text(encoding="utf-8"))
    prior_paper_ids = sorted(rules["paper_ids"])
    if len(prior_paper_ids) != 10 or len(set(prior_paper_ids)) != 10:
        raise ValueError("expected exactly ten distinct prior paper IDs")
    if not set(rules["pilot_ids"]) <= set(prior_paper_ids):
        raise ValueError("prior pilot IDs are not a subset of the prior sample")
    return prior_evidence_bindings(), prior_paper_ids


# ---------------------------------------------------------------------------
# Part B -- select exactly six held-out papers
# ---------------------------------------------------------------------------

def _domain_label(record: dict[str, Any]) -> str:
    if record.get("domain_bucket_status") == "unambiguous":
        return record.get("domain_bucket_guess") or "unclassified"
    return record.get("domain_bucket_status") or "unclassified"


def select_heldout_papers(
    inventory: list[dict[str, Any]],
    prior_paper_ids: list[str],
    seed: int = K5D1_RANDOM_SEED,
    count: int = HELD_OUT_COUNT,
) -> list[str]:
    """Deterministic, label-blind selection of ``count`` new paper IDs.

    Only ``paper_id`` and ``usable_abstract`` are consulted. Domain
    buckets, failure status, and heuristic hit counts are never read
    here -- they are recorded afterward, for honest reporting only, not
    used to filter or bias the draw.
    """
    prior = set(prior_paper_ids)
    eligible = sorted(
        record["paper_id"]
        for record in inventory
        if record["usable_abstract"] and record["paper_id"] not in prior
    )
    if len(eligible) < count:
        raise ValueError(f"requested {count} held-out papers, found {len(eligible)} eligible")
    return sorted(Random(seed).sample(eligible, count))


def _resolve_heldout_source_snapshot(
    selected_ids: list[str], inventory_by_id: dict[str, dict[str, Any]], snapshot_at: str,
) -> list[dict[str, Any]]:
    live_records = k5_paper_inventory._collect_local_papers()
    snapshot: list[dict[str, Any]] = []
    for paper_id in selected_ids:
        local = live_records.get(paper_id)
        if local is None:
            raise ValueError(f"{paper_id}: selected source record cannot be resolved")
        paper = local.paper
        actual_hash = k5_paper_inventory._source_hash(paper.title, paper.abstract)
        if actual_hash != inventory_by_id[paper_id]["source_hash"]:
            raise ValueError(f"{paper_id}: resolved source hash does not match inventory")
        snapshot.append({
            "schema_version": SOURCE_SNAPSHOT_SCHEMA_VERSION,
            "paper_id": paper_id,
            "title": paper.title,
            "abstract": paper.abstract,
            "source": paper.source,
            "external_id": k5_paper_inventory._external_id(paper),
            "source_hash": actual_hash,
            "snapshotted_at": snapshot_at,
        })
    return snapshot


def freeze_heldout_selection(replace: bool = False) -> int:
    outputs = (SELECTION_PATH, SOURCE_SNAPSHOT_PATH)
    if any(path.exists() for path in outputs) and not replace:
        print("refused: held-out selection already exists; pass --replace to overwrite deliberately.", file=sys.stderr)
        return 3

    prior_bindings, prior_paper_ids = validate_prior_evidence()
    inventory = k5_paper_inventory.build_inventory()
    inventory_by_id = {record["paper_id"]: record for record in inventory}

    selected_ids = select_heldout_papers(inventory, prior_paper_ids)
    overlap = sorted(set(selected_ids) & set(prior_paper_ids))
    if overlap:
        print(f"refused: selected papers overlap the prior sample: {overlap}", file=sys.stderr)
        return 4
    if len(selected_ids) != HELD_OUT_COUNT or len(set(selected_ids)) != HELD_OUT_COUNT:
        print("refused: held-out selection is not exactly six distinct papers.", file=sys.stderr)
        return 4

    paper_codes = {paper_id: f"H{index:02d}" for index, paper_id in enumerate(selected_ids, start=1)}
    observed_domain_composition = dict(Counter(_domain_label(inventory_by_id[pid]) for pid in selected_ids))
    selected_at = _now_iso()

    source_snapshot = _resolve_heldout_source_snapshot(selected_ids, inventory_by_id, selected_at)
    snapshot_bytes = "".join(json.dumps(row, sort_keys=True) + "\n" for row in source_snapshot).encode("utf-8")
    _write_atomic(SOURCE_SNAPSHOT_PATH, snapshot_bytes)

    limitations = [
        "product-local corpus", "only six held-out papers",
        "not an external benchmark", "not evidence of Policy C generalization",
        "not a statistical claim",
    ]
    selection = {
        "schema_version": SELECTION_SCHEMA_VERSION,
        "status": "frozen_heldout_selection",
        "selection_rule_version": SELECTION_RULE_VERSION,
        "random_seed": K5D1_RANDOM_SEED,
        "selected_at": selected_at,
        "held_out_count": HELD_OUT_COUNT,
        "eligible_pool_size": len({
            record["paper_id"] for record in inventory
            if record["usable_abstract"] and record["paper_id"] not in set(prior_paper_ids)
        }),
        "paper_ids": selected_ids,
        "paper_codes": paper_codes,
        "source_hashes": {pid: inventory_by_id[pid]["source_hash"] for pid in selected_ids},
        "observed_domain_composition": observed_domain_composition,
        "selection_inputs_used": ["paper_id", "usable_abstract"],
        "selection_inputs_forbidden": [
            "human_labels", "llm_decisions", "failure_status", "heuristic_screening_hit_count", "stress_signals",
        ],
        "prior_k5_sample": {
            "paper_ids": prior_paper_ids,
            "paper_count": len(prior_paper_ids),
            "pilot_paper_codes": ["P01", "P02"],
            "headline_paper_codes": [f"P{number:02d}" for number in range(3, 11)],
            "bindings": prior_bindings,
        },
        "disjoint_from_prior_sample": True,
        "source_snapshot_sha256": _file_sha256(SOURCE_SNAPSHOT_PATH),
        "limitations": limitations,
    }
    selection["manifest_sha256"] = _sha256_payload(selection)
    _write_atomic(SELECTION_PATH, _json_bytes(selection))
    print(f"held-out selection frozen: {len(selected_ids)} paper(s) -> {', '.join(paper_codes[p] for p in selected_ids)}")
    return 0


def load_frozen_selection() -> tuple[list[str], dict[str, Any]]:
    for path, label in ((SELECTION_PATH, "held-out selection"), (SOURCE_SNAPSHOT_PATH, "held-out source snapshot")):
        if not path.exists():
            raise ValueError(f"{label} is missing")
    selection = json.loads(SELECTION_PATH.read_text(encoding="utf-8"))
    if not _manifest_hash_valid(selection, "manifest_sha256"):
        raise ValueError("held-out selection manifest_sha256 mismatch")
    if selection.get("schema_version") != SELECTION_SCHEMA_VERSION:
        raise ValueError("held-out selection schema_version is not current")
    if selection.get("status") != "frozen_heldout_selection":
        raise ValueError("held-out selection status is not frozen_heldout_selection")
    if selection.get("held_out_count") != HELD_OUT_COUNT or len(selection.get("paper_ids", [])) != HELD_OUT_COUNT:
        raise ValueError("held-out selection does not contain exactly six papers")
    if selection.get("source_snapshot_sha256") != _file_sha256(SOURCE_SNAPSHOT_PATH):
        raise ValueError("held-out source snapshot hash does not match the frozen selection")
    prior_ids = set(selection.get("prior_k5_sample", {}).get("paper_ids", []))
    if set(selection["paper_ids"]) & prior_ids:
        raise ValueError("held-out selection overlaps the prior K5 sample")
    if selection.get("prior_k5_sample", {}).get("bindings") != prior_evidence_bindings():
        raise ValueError("prior K5 evidence has drifted since the held-out selection was frozen")

    snapshot = _load_jsonl(SOURCE_SNAPSHOT_PATH)
    snapshot_ids = [row["paper_id"] for row in snapshot]
    if sorted(snapshot_ids) != sorted(selection["paper_ids"]) or len(snapshot_ids) != len(set(snapshot_ids)):
        raise ValueError("held-out source snapshot paper IDs do not match the frozen selection")
    for row in snapshot:
        if row.get("schema_version") != SOURCE_SNAPSHOT_SCHEMA_VERSION:
            raise ValueError(f"{row.get('paper_id')}: source snapshot row uses an obsolete schema")
        actual_hash = k5_paper_inventory._source_hash(str(row.get("title") or ""), row.get("abstract"))
        if actual_hash != row.get("source_hash") or selection["source_hashes"].get(row["paper_id"]) != actual_hash:
            raise ValueError(f"{row.get('paper_id')}: source snapshot hash mismatch")
    return selection["paper_ids"], selection


# ---------------------------------------------------------------------------
# Part C -- candidate generation (frozen selection + snapshot only)
# ---------------------------------------------------------------------------

def _canonical_key(phrase: str) -> str:
    return " ".join(production_keywords._canonical_tokens(phrase))


def _candidate_id(paper_id: str, canonical_key: str, seed: int = CANDIDATE_ID_SEED) -> str:
    digest = hashlib.sha256(f"{seed}\x00{paper_id}\x00{canonical_key}".encode("utf-8")).hexdigest()
    return f"HC-{digest[:12].upper()}"


def generate_candidate_bundle(selection: dict[str, Any], source_by_id: dict[str, dict[str, Any]]) -> dict[str, Any]:
    """Run production YAKE-v2 locally; preserve its exact ordered output.

    No shuffling, no method union, no tuning -- unlike K5B/K5C, exactly
    one extractor is evaluated here so there is nothing to blind between
    methods; the only blinding this workbook needs is that no future LLM
    provenance is ever recorded in it, which is trivially true since no
    LLM is invoked anywhere in this checkpoint.
    """
    if production_keywords.KEYWORD_EXTRACTOR_VERSION != "yake-v2":
        raise RuntimeError("production keyword extractor is no longer the frozen yake-v2 implementation")

    papers: list[dict[str, Any]] = []
    candidates: list[dict[str, Any]] = []
    seen_candidate_ids: set[str] = set()
    for paper_id in selection["paper_ids"]:
        code = selection["paper_codes"][paper_id]
        source = source_by_id[paper_id]
        title = str(source.get("title") or "")
        abstract = source.get("abstract")
        outputs = list(production_keywords.extract_keywords(title, abstract))[:MAX_CANDIDATES_PER_PAPER]
        for rank, phrase in enumerate(outputs, start=1):
            key = _canonical_key(phrase)
            candidate_id = _candidate_id(paper_id, key)
            if candidate_id in seen_candidate_ids:
                raise RuntimeError("opaque candidate ID collision")
            seen_candidate_ids.add(candidate_id)
            candidates.append({
                "candidate_id": candidate_id,
                "paper_id": paper_id,
                "paper_code": code,
                "rank": rank,
                "canonical_key": key,
                "display_phrase": phrase,
            })
        papers.append({
            "paper_id": paper_id,
            "paper_code": code,
            "title": title,
            "abstract": abstract or "",
            "source_hash": source["source_hash"],
        })
    return {
        "extractor_version": production_keywords.KEYWORD_EXTRACTOR_VERSION,
        "extractor_source_sha256": _file_sha256(Path(production_keywords.__file__)),
        "maximum_candidates_per_paper": MAX_CANDIDATES_PER_PAPER,
        "papers": papers,
        "candidates": candidates,
    }


# ---------------------------------------------------------------------------
# Part D -- blank annotation workbook
# ---------------------------------------------------------------------------

def _evidence_hash(headers: tuple[str, ...], values: list[Any]) -> str:
    return _sha256_payload(dict(zip(headers, values)))


def _build_protected_evidence(bundle: dict[str, Any]) -> dict[str, Any]:
    paper_rows = []
    for row_number, paper in enumerate(bundle["papers"], start=2):
        values = [paper["paper_code"], paper["title"], paper["abstract"]]
        paper_rows.append({"row_number": row_number, "paper_code": paper["paper_code"], "sha256": _evidence_hash(PAPER_EVIDENCE_HEADERS, values)})
    candidate_rows = []
    for row_number, candidate in enumerate(bundle["candidates"], start=2):
        values = [candidate["candidate_id"], candidate["paper_code"], candidate["display_phrase"]]
        candidate_rows.append({"row_number": row_number, "candidate_id": candidate["candidate_id"], "sha256": _evidence_hash(CANDIDATE_EVIDENCE_HEADERS, values)})
    payload = {"paper_rows": paper_rows, "candidate_rows": candidate_rows}
    payload["overall_sha256"] = _sha256_payload(payload)
    return payload


def _style_header(sheet, cell_range: str) -> None:
    for row in sheet[cell_range]:
        for cell in row:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _add_list_validation(sheet, cell_range: str, values: tuple[str, ...], prompt: str) -> None:
    validation = DataValidation(type="list", formula1='"' + ",".join(values) + '"', allow_blank=True)
    validation.error = "Choose a value from the dropdown."
    validation.errorTitle = "Invalid value"
    validation.prompt = prompt
    validation.promptTitle = "Allowed values"
    validation.showErrorMessage = True
    validation.showInputMessage = True
    sheet.add_data_validation(validation)
    validation.add(cell_range)


def _validation_specs(candidate_row_count: int) -> set[tuple[str, str]]:
    return {
        (f"D2:D{candidate_row_count + 1}", '"' + ",".join(DECISION_VALUES) + '"'),
        (f"E2:E{candidate_row_count + 1}", '"' + ",".join(REJECTION_REASON_VALUES) + '"'),
        (f"G2:G{candidate_row_count + 1}", '"' + ",".join(CONFIDENCE_VALUES) + '"'),
    }


def create_annotation_workbook(bundle: dict[str, Any], destination: Path) -> None:
    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "Instructions"
    concepts = workbook.create_sheet("Paper concepts")
    review = workbook.create_sheet("Candidate review")

    instruction_rows = [
        ("K5D.1a held-out keyword annotation", "Use only the title and abstract shown in this workbook."),
        ("Order of work", "First complete 3-5 important scientific concepts for a paper. Then review its candidates."),
        ("Concepts", "Concepts are the important topics, tasks, or methods actually represented by the paper -- not restatements of the title."),
        ("Concept IDs", "The five concept cells correspond to C1, C2, C3, C4, and C5. Enter matched IDs as comma-separated values, for example C1, C3."),
        ("Decision", "accept = useful scientific keyword; reject = not useful; uncertain = cannot confidently decide."),
        ("Rejection reason", "For reject only: fragment, malformed, too_broad, affiliation_or_entity, redundant_fragment, or other."),
        ("Confidence", "For every completed candidate decision choose high, medium, or low."),
        ("Editable cells", "Pale yellow cells are for human entry. Lavender cells are frozen evidence and hash-checked."),
        ("Notes", "Use notes for concise rationale or ambiguity. Do not add, delete, sort, or reorder rows."),
    ]
    instructions.append(["Topic", "Guidance"])
    for row in instruction_rows:
        instructions.append(list(row))
    _style_header(instructions, "A1:B1")
    for row_number, row in enumerate(
        instructions.iter_rows(min_row=2, max_row=instructions.max_row, min_col=1, max_col=2), start=2,
    ):
        row[0].fill = SECTION_FILL
        row[0].font = Font(bold=True, color="4A1F5D", size=10)
        row[1].font = Font(size=10)
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        guidance_length = len(str(row[1].value or ""))
        instructions.row_dimensions[row_number].height = min(72, 30 + 12 * (guidance_length // 70))
    instructions.column_dimensions["A"].width = 21
    instructions.column_dimensions["B"].width = 58
    instructions.freeze_panes = "A2"
    instructions.sheet_view.showGridLines = False
    instructions.print_area = f"A1:B{instructions.max_row}"
    instructions.sheet_properties.pageSetUpPr.fitToPage = True
    instructions.page_setup.orientation = "landscape"
    instructions.page_setup.fitToWidth = 1
    instructions.page_setup.fitToHeight = 1
    instructions.protection.sheet = True

    concepts.append(list(PAPER_HEADERS))
    for paper in bundle["papers"]:
        concepts.append([paper["paper_code"], paper["title"], paper["abstract"], None, None, None, None, None, None])
    _style_header(concepts, "A1:I1")
    concepts.freeze_panes = "A2"
    concepts.auto_filter.ref = f"A1:I{concepts.max_row}"
    concepts.sheet_view.showGridLines = False
    for row in concepts.iter_rows(min_row=2, max_row=concepts.max_row, min_col=1, max_col=9):
        for cell in row[:3]:
            cell.fill = EVIDENCE_FILL
            cell.protection = Protection(locked=True)
        for cell in row[3:]:
            cell.fill = EDITABLE_FILL
            cell.protection = Protection(locked=False)
        for index in (1, 2, 3, 4, 5, 6, 7, 8):
            row[index].alignment = Alignment(vertical="top", wrap_text=True)
    concepts.column_dimensions["A"].width = 12
    concepts.column_dimensions["B"].width = 44
    concepts.column_dimensions["C"].width = 80
    for column in ("D", "E", "F", "G", "H"):
        concepts.column_dimensions[column].width = 24
    concepts.column_dimensions["I"].width = 36
    concepts.row_dimensions[1].height = 32
    for row_number in range(2, concepts.max_row + 1):
        concepts.row_dimensions[row_number].height = 105
    concepts.protection.sheet = True
    concepts.protection.autoFilter = False
    concepts.protection.sort = False

    review.append(list(CANDIDATE_HEADERS))
    for candidate in bundle["candidates"]:
        review.append([candidate["candidate_id"], candidate["paper_code"], candidate["display_phrase"], None, None, None, None, None])
    _style_header(review, "A1:H1")
    review.freeze_panes = "A2"
    review.auto_filter.ref = f"A1:H{review.max_row}"
    review.sheet_view.showGridLines = False
    for row in review.iter_rows(min_row=2, max_row=review.max_row, min_col=1, max_col=8):
        for cell in row[:3]:
            cell.fill = EVIDENCE_FILL
            cell.protection = Protection(locked=True)
        for cell in row[3:]:
            cell.fill = EDITABLE_FILL
            cell.protection = Protection(locked=False)
        for index in (2, 5, 7):
            row[index].alignment = Alignment(vertical="top", wrap_text=True)
    review.column_dimensions["A"].width = 18
    review.column_dimensions["B"].width = 12
    review.column_dimensions["C"].width = 42
    review.column_dimensions["D"].width = 14
    review.column_dimensions["E"].width = 24
    review.column_dimensions["F"].width = 24
    review.column_dimensions["G"].width = 14
    review.column_dimensions["H"].width = 38
    review.row_dimensions[1].height = 32
    for row_number in range(2, review.max_row + 1):
        review.row_dimensions[row_number].height = 42
    _add_list_validation(review, f"D2:D{review.max_row}", DECISION_VALUES, "accept, reject, or uncertain")
    _add_list_validation(review, f"E2:E{review.max_row}", REJECTION_REASON_VALUES, "Required for reject only")
    _add_list_validation(review, f"G2:G{review.max_row}", CONFIDENCE_VALUES, "high, medium, or low")
    review.protection.sheet = True
    review.protection.autoFilter = False
    review.protection.sort = False

    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)


def _build_mapping(bundle: dict[str, Any]) -> dict[str, Any]:
    mapping = {
        "schema_version": MAPPING_SCHEMA_VERSION,
        "extractor_version": bundle["extractor_version"],
        "extractor_source_sha256": bundle["extractor_source_sha256"],
        "maximum_candidates_per_paper": bundle["maximum_candidates_per_paper"],
        "paper_count": len(bundle["papers"]),
        "candidate_count": len(bundle["candidates"]),
        "papers": [
            {"paper_id": paper["paper_id"], "paper_code": paper["paper_code"], "source_hash": paper["source_hash"]}
            for paper in bundle["papers"]
        ],
        "candidates": [
            {
                "candidate_id": candidate["candidate_id"], "paper_id": candidate["paper_id"],
                "paper_code": candidate["paper_code"], "rank": candidate["rank"],
                "canonical_key": candidate["canonical_key"], "display_phrase": candidate["display_phrase"],
            }
            for candidate in bundle["candidates"]
        ],
    }
    mapping["mapping_sha256"] = _sha256_payload(mapping)
    return mapping


def _build_manifest(bundle: dict[str, Any], mapping_path: Path, protected: dict[str, Any]) -> dict[str, Any]:
    manifest = {
        "schema_version": MANIFEST_SCHEMA_VERSION,
        "status": "awaiting_human_annotation",
        "creation_timestamp": _now_iso(),
        "workbook_filename": WORKBOOK_PATH.name,
        "mapping_filename": MAPPING_PATH.name,
        "expected_sheets": list(EXPECTED_SHEETS),
        "frozen_inputs": {
            "selection_sha256": _file_sha256(SELECTION_PATH),
            "source_snapshot_sha256": _file_sha256(SOURCE_SNAPSHOT_PATH),
        },
        "candidate_mapping_sha256": _file_sha256(mapping_path),
        "paper_count": len(bundle["papers"]),
        "candidate_count": len(bundle["candidates"]),
        "paper_row_order": [paper["paper_code"] for paper in bundle["papers"]],
        "candidate_row_order": [candidate["candidate_id"] for candidate in bundle["candidates"]],
        "protected_evidence": protected,
        "editable_columns_excluded_from_protected_hashes": {
            "Paper concepts": list(PAPER_HEADERS[3:]),
            "Candidate review": list(CANDIDATE_HEADERS[3:]),
        },
    }
    manifest["manifest_sha256"] = _sha256_payload(manifest)
    return manifest


def _cell_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _parse_concept_ids(value: Any) -> tuple[list[str], bool]:
    return annotation._parse_concept_ids(value)


def generate_annotation_artifacts(replace: bool = False) -> int:
    outputs = (WORKBOOK_PATH, MAPPING_PATH, MANIFEST_PATH)
    if any(path.exists() for path in outputs) and not replace:
        print("refused: held-out annotation output exists; pass --replace to overwrite deliberately.", file=sys.stderr)
        return 3

    try:
        selected_ids, selection = load_frozen_selection()
    except ValueError as exc:
        print(f"VIOLATION: {exc}", file=sys.stderr)
        return 4
    snapshot = _load_jsonl(SOURCE_SNAPSHOT_PATH)
    source_by_id = {row["paper_id"]: row for row in snapshot}

    bundle = generate_candidate_bundle(selection, source_by_id)
    mapping = _build_mapping(bundle)
    protected = _build_protected_evidence(bundle)
    K5D1_DIR.mkdir(parents=True, exist_ok=True)
    temporary_workbook = WORKBOOK_PATH.with_name(f".{WORKBOOK_PATH.name}.tmp.xlsx")
    temporary_mapping = MAPPING_PATH.with_name(f".{MAPPING_PATH.name}.tmp")
    temporary_manifest = MANIFEST_PATH.with_name(f".{MANIFEST_PATH.name}.tmp")
    try:
        create_annotation_workbook(bundle, temporary_workbook)
        temporary_mapping.write_bytes(_json_bytes(mapping))
        manifest = _build_manifest(bundle, temporary_mapping, protected)
        temporary_manifest.write_bytes(_json_bytes(manifest))
        validation = validate_annotation_workbook(
            workbook_path=temporary_workbook, mapping_path=temporary_mapping, manifest_path=temporary_manifest,
        )
        if validation:
            for violation in validation:
                print(f"VIOLATION: {violation}", file=sys.stderr)
            return 4
        os.replace(temporary_workbook, WORKBOOK_PATH)
        os.replace(temporary_mapping, MAPPING_PATH)
        os.replace(temporary_manifest, MANIFEST_PATH)
    finally:
        for temporary in (temporary_workbook, temporary_mapping, temporary_manifest):
            if temporary.exists():
                temporary.unlink()

    print(f"created blank held-out workbook: papers={len(bundle['papers'])}, candidates={len(bundle['candidates'])}")
    return 0


# ---------------------------------------------------------------------------
# Part E -- validation
# ---------------------------------------------------------------------------

def validate_annotation_workbook(
    require_all_complete: bool = False,
    workbook_path: Path = WORKBOOK_PATH,
    mapping_path: Path = MAPPING_PATH,
    manifest_path: Path = MANIFEST_PATH,
) -> list[str]:
    violations: list[str] = []
    for path, label in ((workbook_path, "held-out workbook"), (mapping_path, "held-out candidate mapping"), (manifest_path, "held-out annotation manifest")):
        if not path.exists():
            violations.append(f"{label} is missing")
    if violations:
        return violations

    try:
        selected_ids, selection = load_frozen_selection()
    except ValueError as exc:
        violations.append(str(exc))
        return violations

    try:
        mapping = json.loads(mapping_path.read_text(encoding="utf-8"))
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        workbook = load_workbook(workbook_path, data_only=False)
    except (OSError, json.JSONDecodeError, ValueError) as exc:
        return violations + [f"held-out annotation evidence could not be read: {type(exc).__name__}"]

    if mapping.get("schema_version") != MAPPING_SCHEMA_VERSION:
        violations.append("held-out candidate mapping schema is not current")
    if not _manifest_hash_valid(mapping, "mapping_sha256"):
        violations.append("held-out candidate mapping self-hash mismatch")
    if manifest.get("schema_version") != MANIFEST_SCHEMA_VERSION:
        violations.append("held-out annotation manifest schema is not current")
    if not _manifest_hash_valid(manifest, "manifest_sha256"):
        violations.append("held-out annotation manifest self-hash mismatch")
    if manifest.get("candidate_mapping_sha256") != _file_sha256(mapping_path):
        violations.append("held-out candidate mapping file hash mismatch")
    if manifest.get("frozen_inputs", {}).get("selection_sha256") != _file_sha256(SELECTION_PATH):
        violations.append("manifest selection_sha256 binding mismatch")
    if manifest.get("frozen_inputs", {}).get("source_snapshot_sha256") != _file_sha256(SOURCE_SNAPSHOT_PATH):
        violations.append("manifest source_snapshot_sha256 binding mismatch")

    if workbook.sheetnames != list(EXPECTED_SHEETS):
        violations.append("workbook sheet names or order changed")
        return violations
    concepts = workbook["Paper concepts"]
    review = workbook["Candidate review"]
    paper_count = manifest.get("paper_count")
    candidate_count = manifest.get("candidate_count")
    if paper_count != HELD_OUT_COUNT or mapping.get("paper_count") != HELD_OUT_COUNT:
        violations.append(f"held-out evidence does not describe exactly {HELD_OUT_COUNT} papers")
    if candidate_count != mapping.get("candidate_count"):
        violations.append("candidate count differs between mapping and manifest")
    if tuple(cell.value for cell in concepts[1]) != PAPER_HEADERS:
        violations.append("Paper concepts headers changed")
    if tuple(cell.value for cell in review[1]) != CANDIDATE_HEADERS:
        violations.append("Candidate review headers changed")
    if concepts.max_row != (paper_count or 0) + 1 or concepts.max_column != len(PAPER_HEADERS):
        violations.append("Paper concepts rows or columns were added or removed")
    if review.max_row != (candidate_count or 0) + 1 or review.max_column != len(CANDIDATE_HEADERS):
        violations.append("Candidate review rows or columns were added or removed")

    paper_order = [_cell_text(concepts.cell(row=row, column=1).value) for row in range(2, concepts.max_row + 1)]
    candidate_order = [_cell_text(review.cell(row=row, column=1).value) for row in range(2, review.max_row + 1)]
    if paper_order != manifest.get("paper_row_order"):
        violations.append("Paper concepts rows are missing, added, or reordered")
    if candidate_order != manifest.get("candidate_row_order"):
        violations.append("Candidate review rows are missing, added, or reordered")
    if sorted(paper_order) != sorted(selection["paper_codes"].values()):
        violations.append("workbook paper codes do not match the frozen selection")

    protected = manifest.get("protected_evidence", {})
    for record in protected.get("paper_rows", []):
        row_number = record.get("row_number")
        if not isinstance(row_number, int) or row_number > concepts.max_row:
            violations.append("Paper concepts protected row binding is invalid")
            continue
        values = [concepts.cell(row=row_number, column=column).value for column in range(1, 4)]
        if record.get("sha256") != _evidence_hash(PAPER_EVIDENCE_HEADERS, values):
            violations.append(f"{record.get('paper_code')}: protected paper evidence hash mismatch")
    for record in protected.get("candidate_rows", []):
        row_number = record.get("row_number")
        if not isinstance(row_number, int) or row_number > review.max_row:
            violations.append("Candidate review protected row binding is invalid")
            continue
        values = [review.cell(row=row_number, column=column).value for column in range(1, 4)]
        if record.get("sha256") != _evidence_hash(CANDIDATE_EVIDENCE_HEADERS, values):
            violations.append(f"{record.get('candidate_id')}: protected candidate evidence hash mismatch")
    protected_copy = {"paper_rows": protected.get("paper_rows", []), "candidate_rows": protected.get("candidate_rows", [])}
    if protected.get("overall_sha256") != _sha256_payload(protected_copy):
        violations.append("protected evidence aggregate hash mismatch")

    mapped_order = [row.get("candidate_id") for row in mapping.get("candidates", [])]
    if mapped_order != candidate_order:
        violations.append("candidate mapping and workbook row order differ")
    mapped_paper_order = [row.get("paper_code") for row in mapping.get("papers", [])]
    if mapped_paper_order != paper_order:
        violations.append("paper mapping and workbook row order differ")

    if concepts.freeze_panes != "A2" or review.freeze_panes != "A2":
        violations.append("required frozen header rows changed")
    if concepts.auto_filter.ref != f"A1:I{(paper_count or 0) + 1}" or review.auto_filter.ref != f"A1:H{(candidate_count or 0) + 1}":
        violations.append("required worksheet filters changed")
    observed_validations = {(str(item.sqref), item.formula1) for item in review.data_validations.dataValidation}
    if observed_validations != _validation_specs(candidate_count or 0):
        violations.append("candidate dropdown validations changed")

    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == TYPE_FORMULA:
                    violations.append(f"unexpected formula in {sheet.title}!{cell.coordinate}")
                elif cell.data_type == TYPE_ERROR:
                    violations.append(f"spreadsheet error in {sheet.title}!{cell.coordinate}")

    concept_values_by_code: dict[str, list[str]] = {}
    for row_number in range(2, concepts.max_row + 1):
        code = _cell_text(concepts.cell(row=row_number, column=1).value)
        entered = [_cell_text(concepts.cell(row=row_number, column=column).value) for column in range(4, 9)]
        concept_values_by_code[code] = entered
        nonblank = [value for value in entered if value]
        if len(nonblank) != len({value.casefold() for value in nonblank}):
            violations.append(f"{code}: concept entries must be distinct")
        if require_all_complete and not 3 <= len(nonblank) <= 5:
            violations.append(f"{code}: paper requires 3-5 completed concepts")

    for row_number in range(2, review.max_row + 1):
        candidate_id = _cell_text(review.cell(row=row_number, column=1).value)
        code = _cell_text(review.cell(row=row_number, column=2).value)
        decision = _cell_text(review.cell(row=row_number, column=4).value).casefold()
        reason = _cell_text(review.cell(row=row_number, column=5).value).casefold()
        matched, matched_valid = _parse_concept_ids(review.cell(row=row_number, column=6).value)
        confidence = _cell_text(review.cell(row=row_number, column=7).value).casefold()
        if decision and decision not in DECISION_VALUES:
            violations.append(f"{candidate_id}: invalid decision")
        if reason and reason not in REJECTION_REASON_VALUES:
            violations.append(f"{candidate_id}: invalid rejection reason")
        if confidence and confidence not in CONFIDENCE_VALUES:
            violations.append(f"{candidate_id}: invalid confidence")
        if not matched_valid:
            violations.append(f"{candidate_id}: invalid or duplicate matched concept IDs")
        if not decision:
            if reason or matched or confidence:
                violations.append(f"{candidate_id}: annotation fields entered without a decision")
            if require_all_complete:
                violations.append(f"{candidate_id}: candidate decision is incomplete")
            continue
        if not confidence:
            violations.append(f"{candidate_id}: completed decision requires confidence")
        if decision == "reject" and not reason:
            violations.append(f"{candidate_id}: reject requires a rejection reason")
        if decision != "reject" and reason:
            violations.append(f"{candidate_id}: rejection reason is only valid for reject")
        if decision == "accept" and not matched:
            violations.append(f"{candidate_id}: accept requires at least one matched concept ID")
        if decision == "reject" and matched:
            violations.append(f"{candidate_id}: reject must not contain matched concept IDs")
        entered = concept_values_by_code.get(code, [])
        for concept_id in matched:
            if not entered[int(concept_id[1]) - 1]:
                violations.append(f"{candidate_id}: {concept_id} does not identify a completed concept")

    return violations


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="K5D.1a held-out annotation preparation tooling.")
    subparsers = parser.add_subparsers(dest="command", required=True)
    subparsers.add_parser("validate-prior")
    select_parser = subparsers.add_parser("select")
    select_parser.add_argument("--replace", action="store_true")
    generate_parser = subparsers.add_parser("generate")
    generate_parser.add_argument("--replace", action="store_true")
    validate_parser = subparsers.add_parser("validate")
    validate_parser.add_argument("--require-all-complete", action="store_true")
    args = parser.parse_args(argv)

    if args.command == "validate-prior":
        bindings, prior_ids = validate_prior_evidence()
        print(json.dumps({"prior_paper_count": len(prior_ids), "bindings": bindings}, sort_keys=True))
        return 0
    if args.command == "select":
        return freeze_heldout_selection(replace=args.replace)
    if args.command == "generate":
        return generate_annotation_artifacts(replace=args.replace)
    if args.command == "validate":
        violations = validate_annotation_workbook(
            require_all_complete=args.require_all_complete,
            workbook_path=WORKBOOK_PATH, mapping_path=MAPPING_PATH, manifest_path=MANIFEST_PATH,
        )
        if violations:
            for violation in violations:
                print(f"VIOLATION: {violation}", file=sys.stderr)
            return 1
        print("valid: held-out selection, workbook, mapping, and manifest pass the K5D.1a contract.")
        return 0
    return 2


if __name__ == "__main__":
    raise SystemExit(main())
