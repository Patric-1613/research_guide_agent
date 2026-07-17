#!/usr/bin/env python3
"""One-time/occasional conversion: reference_topics.xlsx -> reference_topics.json.

Reads the "Reference Topics" sheet of the human-maintained reference
spreadsheet and produces a grouped JSON file that scripts/eval_retrieval.py
consumes. Run manually whenever the spreadsheet is updated -- NOT run
automatically as part of the eval script itself, and not part of
pytest/CI.

Reads cells via openpyxl directly rather than pandas.read_excel: pandas
auto-infers the arxiv_id column as float64 (since the IDs look numeric),
which silently corrupts any ID with a trailing zero (e.g. "2106.09680"
becomes 2106.0968) or a non-numeric suffix (e.g. "2506.06962v3" would
break pandas' float coercion entirely). openpyxl's raw cell values
preserve the original string exactly.

Usage:
    python scripts/convert_reference_sheet.py
    python scripts/convert_reference_sheet.py --input eval_data/reference_topics.xlsx --output eval_data/reference_topics.json
"""

from __future__ import annotations

import argparse
import json
import logging
import os

import openpyxl

logger = logging.getLogger(__name__)

SHEET_NAME = "Reference Topics"
REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEFAULT_INPUT = os.path.join(REPO_ROOT, "eval_data", "reference_topics.xlsx")
DEFAULT_OUTPUT = os.path.join(REPO_ROOT, "eval_data", "reference_topics.json")


def _clean(value) -> str | None:
    """Blank cells come back as None; stray whitespace-only strings should
    also collapse to None so downstream code has one consistent "empty"
    representation instead of two."""
    if value is None:
        return None
    text = str(value).strip()
    return text or None


# Reviewers sometimes mark a topic as "no real paper yet" with an explicit
# sentinel phrase rather than leaving paper_title blank (seen in practice:
# "(no single landmark paper identified)") -- treated identically to an
# empty cell, not ingested as a literal expected paper title.
_NO_PAPER_SENTINELS = {"(no single landmark paper identified)"}


def _is_placeholder_title(title: str | None) -> bool:
    return title is None or title.strip().lower() in _NO_PAPER_SENTINELS


def convert(input_path: str) -> dict:
    wb = openpyxl.load_workbook(input_path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet {SHEET_NAME!r} not found in {input_path!r} (found: {wb.sheetnames})")
    ws = wb[SHEET_NAME]

    rows = list(ws.iter_rows(values_only=True))
    header = [_clean(h) for h in rows[0]]
    required = {"topic_id", "topic", "difficulty", "paper_title", "arxiv_id", "doi", "has_abstract", "confidence"}
    missing = required - set(header)
    if missing:
        raise ValueError(f"Reference sheet is missing required column(s): {sorted(missing)}")

    topics: dict[str, dict] = {}
    blank_rows = 0
    skipped_topic_ids: set[str] = set()

    for raw_row in rows[1:]:
        record = dict(zip(header, (_clean(v) for v in raw_row)))
        topic_id = record.get("topic_id")

        if topic_id is None:
            # Fully blank spacer row (no topic_id at all) -- not a real
            # topic, just Excel formatting noise. Counted, not warned on
            # individually: there's no topic_id to name in a warning.
            blank_rows += 1
            continue

        if topic_id not in topics:
            topics[topic_id] = {
                "topic": record.get("topic"),
                "difficulty": record.get("difficulty"),
                "expected_papers": [],
            }

        if _is_placeholder_title(record.get("paper_title")):
            # A real topic_id with no usable paper_title -- either the cell
            # is empty, or it's an explicit "no landmark paper yet" sentinel
            # (see _NO_PAPER_SENTINELS). Either way: skip it, but track it so
            # main() can report exactly which topic_ids lost a row this way.
            skipped_topic_ids.add(topic_id)
            continue

        topics[topic_id]["expected_papers"].append({
            "title": record.get("paper_title"),
            "arxiv_id": record.get("arxiv_id"),
            "doi": record.get("doi"),
            "has_abstract": record.get("has_abstract"),
            "confidence": record.get("confidence"),
        })

    if blank_rows:
        logger.info("Skipped %d fully blank row(s) (no topic_id)", blank_rows)

    # A topic_id can end up with zero expected_papers if EVERY row for it
    # had an empty paper_title -- worth a distinct, louder warning from the
    # per-row skip above, since it means the whole topic is currently unusable.
    empty_topics = sorted(tid for tid, t in topics.items() if not t["expected_papers"])
    if skipped_topic_ids:
        logger.warning(
            "Skipped %d row(s) with empty paper_title, for topic_id(s): %s",
            len(skipped_topic_ids), sorted(skipped_topic_ids),
        )
    if empty_topics:
        logger.warning(
            "%d topic(s) have ZERO usable expected_papers after skipping empty rows: %s",
            len(empty_topics), empty_topics,
        )

    return topics


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")

    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--input", default=DEFAULT_INPUT, help=f"Path to reference_topics.xlsx (default: {DEFAULT_INPUT})")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help=f"Path to write the converted JSON (default: {DEFAULT_OUTPUT})")
    args = parser.parse_args()

    topics = convert(args.input)

    os.makedirs(os.path.dirname(args.output), exist_ok=True)
    with open(args.output, "w") as f:
        json.dump(topics, f, indent=2, ensure_ascii=False)
        f.write("\n")

    total_papers = sum(len(t["expected_papers"]) for t in topics.values())
    print(f"\nWrote {len(topics)} topics ({total_papers} expected papers total) to {args.output}")


if __name__ == "__main__":
    main()
