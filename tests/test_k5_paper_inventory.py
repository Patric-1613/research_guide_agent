"""Focused K5B.1a methodology, screening-workbook, and read-only tests."""

from __future__ import annotations

import json
import os
import sys
from pathlib import Path
from unittest.mock import patch

import pytest
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import scripts.k5_paper_inventory as k5
from research_agent.curation_session import save_curation_session
from research_agent.qa import QA_CHECKPOINT_DB_PATH, sqlite_checkpointer
from research_agent.query_expansion import PaperPoolSession
from research_agent.schema import Paper
from research_agent.telemetry import USAGE_DB_PATH
from tests._usage_db_fingerprint import fingerprint_usage_db

_ABSTRACT = (
    "We present a comprehensive study of graph neural networks for molecular "
    "property prediction. Our method combines message passing with attention "
    "mechanisms to capture long range dependencies between atoms across five "
    "benchmark datasets while remaining computationally efficient at inference."
)
_SHORT_ABSTRACT = "A brief study of anomaly detection methods for industrial sensor deployments today."


def _paper(pid: str, title: str, abstract: str | None = _ABSTRACT, keywords: list[str] | None = None) -> Paper:
    return Paper(
        title=title,
        authors=["A"],
        year=2024,
        venue="X",
        abstract=abstract,
        url=None,
        doi=None,
        citation_count=None,
        source="arxiv",
        paper_id=pid,
        keywords=keywords or [],
    )


@pytest.fixture
def working_dirs(tmp_path, monkeypatch):
    db_path = tmp_path / "checkpoints.sqlite"
    working_dir = tmp_path / "working"
    monkeypatch.setattr(k5, "QA_CHECKPOINT_DB_PATH", db_path)
    monkeypatch.setattr(k5, "EVAL_WORKING_DIR", working_dir)
    monkeypatch.setattr(k5, "INVENTORY_PATH", working_dir / "inventory.jsonl")
    monkeypatch.setattr(k5, "SAMPLE_PATH", working_dir / "proposed_sample.jsonl")
    monkeypatch.setattr(k5, "RULES_PATH", working_dir / "selection_rules.json")
    monkeypatch.setattr(k5, "SCREENING_WORKBOOK_PATH", working_dir / "failure_screening.xlsx")
    monkeypatch.setattr(k5, "SCREENING_MANIFEST_PATH", working_dir / "failure_screening_manifest.json")
    return db_path, working_dir


def _save_session(db_path: Path, session_id: str, session: PaperPoolSession) -> None:
    with sqlite_checkpointer(db_path) as checkpointer:
        save_curation_session(session, session_id, checkpointer)


def _build_full_pool(db_path: Path) -> None:
    heuristic_papers = [
        (_paper(f"heuristic-{index}", f"Heuristic paper {index}", keywords=[f"causing tokens {index}"]), 0.9)
        for index in range(8)
    ]
    _save_session(
        db_path,
        "heuristic-session",
        PaperPoolSession(topic="quality review", reserve=heuristic_papers, cursor=len(heuristic_papers)),
    )

    domain_topics = (
        ("nlp", "natural language retrieval"),
        ("vision", "image segmentation"),
        ("security", "software vulnerability detection"),
        ("applied", "clinical chemistry"),
        ("ambiguous", "natural language privacy"),
        ("unknown", "operations research"),
    )
    for prefix, topic in domain_topics:
        papers = [(_paper(f"{prefix}-{index}", f"{prefix} paper {index}"), 0.8) for index in range(4)]
        _save_session(db_path, f"domain-{prefix}", PaperPoolSession(topic=topic, reserve=papers, cursor=len(papers)))

    acronym = _paper("edge-acronym", "RAG NLP LLM GPU CPU TPU FPGA Study")
    noisy = _paper("edge-noisy", "Noisy prose", _ABSTRACT + " %%%^^^&&&***(((())))")
    short = _paper("edge-short", "Short abstract", _SHORT_ABSTRACT)
    cross = _paper("edge-cross", "Cross topic paper")
    _save_session(db_path, "edge-1", PaperPoolSession(topic="language model", reserve=[(acronym, 0.9), (noisy, 0.8), (short, 0.7), (cross, 0.6)], cursor=4))
    _save_session(db_path, "edge-2", PaperPoolSession(topic="clinical education", reserve=[(cross, 0.9)], cursor=1))


def _complete_screening(working_dir: Path, decision: str = "yes") -> None:
    path = working_dir / "failure_screening.xlsx"
    workbook = load_workbook(path)
    sheet = workbook["Failure screening"]
    for row_number in range(2, sheet.max_row + 1):
        sheet[f"K{row_number}"] = decision
        if decision == "yes":
            sheet[f"L{row_number}"] = "fragment"
    workbook.save(path)


def test_heuristic_hits_are_suspected_and_never_automatically_confirmed(working_dirs):
    db_path, _ = working_dirs
    paper = _paper("heuristic", "Heuristic", keywords=["causing all tokens"])
    _save_session(db_path, "s1", PaperPoolSession(topic="review", reserve=[(paper, 0.9)], cursor=1))

    inventory = k5.build_inventory()
    assert inventory[0]["failure_status"] == "suspected_failure"
    assert inventory[0]["failure_evidence"] is None
    selected, shortfalls, _, _ = k5.select_sample(inventory)
    assert [row["stratum"] for row in selected] == ["suspected_failure"]
    assert "known_failure" in shortfalls


def test_unavailable_documented_ids_do_not_count_as_local_failures(working_dirs):
    db_path, _ = working_dirs
    local = _paper("ordinary-local-id", "Ordinary")
    _save_session(db_path, "s1", PaperPoolSession(topic="review", reserve=[(local, 0.9)], cursor=1))

    inventory = k5.build_inventory()
    assert not ({*k5.DOCUMENTED_FAILURE_EVIDENCE} & {row["paper_id"] for row in inventory})
    selected, shortfalls, _, _ = k5.select_sample(inventory, screening_complete=True)
    assert not [row for row in selected if row["stratum"] == "known_failure"]
    assert "known_failure" in shortfalls


def test_ambiguous_domain_matches_are_exposed_without_a_guess(working_dirs):
    db_path, _ = working_dirs
    paper = _paper("ambiguous", "Ambiguous")
    _save_session(db_path, "s1", PaperPoolSession(topic="natural language privacy", reserve=[(paper, 0.9)], cursor=1))

    record = k5.build_inventory()[0]
    assert record["domain_bucket_matches"] == ["security_software", "ml_nlp"]
    assert record["domain_bucket_guess"] is None
    assert record["domain_bucket_status"] == "ambiguous"


def test_vulnerability_detection_is_security_not_computer_vision(working_dirs):
    db_path, _ = working_dirs
    paper = _paper("vuln", "Vulnerability Detection")
    _save_session(db_path, "s1", PaperPoolSession(topic="vulnerability detection", reserve=[(paper, 0.9)], cursor=1))

    record = k5.build_inventory()[0]
    assert record["domain_bucket_matches"] == ["security_software"]
    assert record["domain_bucket_guess"] == "security_software"
    assert record["domain_bucket_status"] == "unambiguous"


def test_each_reserve_paper_is_recorded_once_per_session(working_dirs):
    db_path, _ = working_dirs
    paper = _paper("reserve", "Reserve")
    _save_session(db_path, "s1", PaperPoolSession(topic="review", reserve=[(paper, 0.9)], cursor=1))

    record = k5.build_inventory()[0]
    assert len(record["provenance"]) == 1


def test_inventory_keeps_abstract_and_stored_phrase_values_out(working_dirs):
    db_path, _ = working_dirs
    phrase = "causing private phrase"
    paper = _paper("private", "Private", _ABSTRACT, [phrase])
    _save_session(db_path, "s1", PaperPoolSession(topic="review", reserve=[(paper, 0.9)], cursor=1))

    payload = json.dumps(k5.build_inventory())
    assert _ABSTRACT not in payload
    assert phrase not in payload


def test_product_local_diversity_includes_all_observed_status_groups(working_dirs):
    db_path, _ = working_dirs
    _build_full_pool(db_path)
    inventory = k5.build_inventory()
    selected, _, _, _ = k5.select_sample(inventory)
    diversity = [row for row in selected if row["stratum"] == "product_local_diversity"]
    observed = {
        row["domain_bucket_guess"] if row["domain_bucket_status"] == "unambiguous" else row["domain_bucket_status"]
        for row in diversity
    }
    assert {"security_software", "computer_vision", "ml_nlp", "applied_domain", "ambiguous", "unclassified"} <= observed


def test_local_edge_cases_retain_subtype_and_measured_signals(working_dirs):
    db_path, _ = working_dirs
    _build_full_pool(db_path)
    selected, _, _, _ = k5.select_sample(k5.build_inventory())
    edges = [row for row in selected if row["stratum"] == "local_edge_case"]
    assert len(edges) == k5.STRATA_TARGETS["local_edge_case"]
    assert all(row["local_edge_subtype"].startswith("local_") for row in edges)
    assert all(row["measured_signals"] for row in edges)
    assert all("stress_case" not in row["stratum"] for row in edges)


def test_screening_workbook_values_layout_and_validation_lists(working_dirs):
    db_path, working_dir = working_dirs
    phrase = "causing all tokens"
    paper = _paper("heuristic", "Screening title", keywords=[phrase])
    _save_session(db_path, "s1", PaperPoolSession(topic="review", reserve=[(paper, 0.9)], cursor=1))

    assert k5.export_failure_screening() == 0
    workbook = load_workbook(working_dir / "failure_screening.xlsx")
    assert workbook.sheetnames == ["Instructions", "Failure screening"]
    sheet = workbook["Failure screening"]
    assert tuple(cell.value for cell in sheet[1]) == k5.SCREENING_HEADERS
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == "A1:M2"
    assert sheet["A2"].value == "heuristic"
    assert sheet["B2"].value == "Screening title"
    assert sheet["C2"].value == _ABSTRACT
    assert sheet["F2"].value == phrase
    assert sheet["G2"].value == k5.FRAGMENT_HEURISTIC_VERSION
    assert sheet["I2"].value == 1
    assert sheet["J2"].value is None
    assert sheet["K2"].value is None and sheet["L2"].value is None and sheet["M2"].value is None
    assert sheet["B2"].alignment.wrap_text and sheet["C2"].alignment.wrap_text
    assert sheet["F2"].alignment.wrap_text and sheet["M2"].alignment.wrap_text
    validations = {(item.formula1, str(item.sqref)) for item in sheet.data_validations.dataValidation}
    assert ('"yes,no,uncertain"', "K2") in validations
    assert ('"fragment,affiliation_or_entity,redundant_fragment,too_broad,malformed,other"', "L2") in validations

    manifest = json.loads((working_dir / "failure_screening_manifest.json").read_text())
    assert manifest["row_count"] == 1
    assert manifest["paper_count"] == 1
    assert manifest["source_records"]["heuristic"]["source_record_sha256"]
    assert manifest["rows"][0]["evidence_sha256"]


def test_screening_export_refuses_overwrite_without_replace(working_dirs):
    db_path, working_dir = working_dirs
    paper = _paper("heuristic", "Screening title", keywords=["causing all tokens"])
    _save_session(db_path, "s1", PaperPoolSession(topic="review", reserve=[(paper, 0.9)], cursor=1))
    assert k5.export_failure_screening() == 0
    workbook_hash = k5._file_sha256(working_dir / "failure_screening.xlsx")
    manifest_hash = k5._file_sha256(working_dir / "failure_screening_manifest.json")

    assert k5.export_failure_screening(replace=False) == 3
    assert k5._file_sha256(working_dir / "failure_screening.xlsx") == workbook_hash
    assert k5._file_sha256(working_dir / "failure_screening_manifest.json") == manifest_hash
    assert k5.export_failure_screening(replace=True) == 0


def test_incomplete_screening_blocks_approval_and_never_uses_suspected_as_known(working_dirs):
    db_path, working_dir = working_dirs
    _build_full_pool(db_path)
    k5.write_inventory(k5.build_inventory())
    assert k5.export_failure_screening() == 0

    assert k5.freeze_sample() == 4
    rules = json.loads((working_dir / "selection_rules.json").read_text())
    sample = [json.loads(line) for line in (working_dir / "proposed_sample.jsonl").read_text().splitlines()]
    assert rules["status"] == "awaiting_human_screening"
    assert rules["strata_actual"].get("known_failure", 0) == 0
    assert rules["strata_actual"]["suspected_failure"] == 8
    assert all(row["status"] != "approved" for row in sample)
    assert "known_failure" in rules["shortfalls"]


def test_completed_screening_with_required_distinct_failures_allows_approval(working_dirs):
    db_path, working_dir = working_dirs
    _build_full_pool(db_path)
    k5.write_inventory(k5.build_inventory())
    assert k5.export_failure_screening() == 0
    _complete_screening(working_dir)

    assert k5.freeze_sample() == 0
    rules = json.loads((working_dir / "selection_rules.json").read_text())
    assert rules["status"] == "approved"
    assert rules["strata_actual"] == k5.STRATA_TARGETS
    assert rules["shortfalls"] == {}
    assert rules["observed_product_local_domain_composition"]
    assert k5.validate_frozen_sample() == []


def test_tampered_workbook_evidence_blocks_final_freezing(working_dirs):
    db_path, working_dir = working_dirs
    _build_full_pool(db_path)
    k5.write_inventory(k5.build_inventory())
    assert k5.export_failure_screening() == 0
    _complete_screening(working_dir)
    workbook_path = working_dir / "failure_screening.xlsx"
    workbook = load_workbook(workbook_path)
    workbook["Failure screening"]["B2"] = "tampered title"
    workbook.save(workbook_path)

    assert k5.freeze_sample() == 4
    rules = json.loads((working_dir / "selection_rules.json").read_text())
    assert rules["status"] == "awaiting_human_screening"
    assert any("hash mismatch" in violation for violation in rules["screening_violations"])


def test_tampered_screening_manifest_blocks_final_freezing(working_dirs):
    db_path, working_dir = working_dirs
    _build_full_pool(db_path)
    k5.write_inventory(k5.build_inventory())
    assert k5.export_failure_screening() == 0
    _complete_screening(working_dir)
    path = working_dir / "failure_screening_manifest.json"
    manifest = json.loads(path.read_text())
    manifest["row_count"] += 1
    path.write_text(json.dumps(manifest))

    assert k5.freeze_sample() == 4
    rules = json.loads((working_dir / "selection_rules.json").read_text())
    assert any("manifest_sha256 mismatch" in violation for violation in rules["screening_violations"])


def test_completed_screening_shortfall_does_not_promote_no_or_uncertain_rows(working_dirs):
    db_path, working_dir = working_dirs
    _build_full_pool(db_path)
    k5.write_inventory(k5.build_inventory())
    k5.export_failure_screening()
    workbook_path = working_dir / "failure_screening.xlsx"
    workbook = load_workbook(workbook_path)
    sheet = workbook["Failure screening"]
    for row_number in range(2, sheet.max_row + 1):
        sheet[f"K{row_number}"] = "yes" if row_number < 5 else "no"
        if row_number < 5:
            sheet[f"L{row_number}"] = "fragment"
    workbook.save(workbook_path)

    assert k5.freeze_sample() == 4
    rules = json.loads((working_dir / "selection_rules.json").read_text())
    assert rules["status"] == "awaiting_human_screening"
    assert rules["strata_actual"]["known_failure"] == 3
    assert rules["strata_actual"].get("suspected_failure", 0) == 0
    assert "known_failure" in rules["shortfalls"]


def test_old_k5b1_v1_artifacts_fail_current_validation(working_dirs):
    _, working_dir = working_dirs
    working_dir.mkdir(parents=True)
    (working_dir / "inventory.jsonl").write_text(json.dumps({
        "schema_version": "k5b1-inventory-v1", "paper_id": "old", "failure_status": None,
    }) + "\n")
    (working_dir / "proposed_sample.jsonl").write_text(json.dumps({
        "schema_version": "k5b1-sample-v1", "selection_rule_version": "k5b1-v1",
        "paper_id": "old", "stratum": "cross_domain", "source_hash": "x",
    }) + "\n")
    (working_dir / "selection_rules.json").write_text(json.dumps({
        "schema_version": "k5b1-sample-v1", "selection_rule_version": "k5b1-v1",
        "status": "approved", "paper_ids": ["old"], "manifest_sha256": "old",
    }))

    violations = k5.validate_frozen_sample()
    assert any("not current" in violation or "obsolete" in violation for violation in violations)


def test_terminal_output_never_contains_abstract_or_keyword_phrase(working_dirs, capsys):
    db_path, _ = working_dirs
    secret_abstract = "SECRET_ABSTRACT evidence words enough to satisfy every local usable source requirement today."
    secret_phrase = "causing SECRET_PHRASE"
    paper = _paper("heuristic", "Safe title", secret_abstract, [secret_phrase])
    _save_session(db_path, "s1", PaperPoolSession(topic="review", reserve=[(paper, 0.9)], cursor=1))

    assert k5.main(["export-failure-screening"]) == 0
    output = capsys.readouterr().out + capsys.readouterr().err
    assert secret_abstract not in output
    assert secret_phrase not in output
    assert "SECRET_ABSTRACT" not in output and "SECRET_PHRASE" not in output


def test_no_fresh_keyword_extraction_in_inventory_screening_or_freeze(working_dirs):
    db_path, working_dir = working_dirs
    _build_full_pool(db_path)
    with patch("research_agent.keywords.extract_keywords") as extract:
        inventory = k5.build_inventory()
        k5.write_inventory(inventory)
        k5.export_failure_screening()
        k5.freeze_sample()
        extract.assert_not_called()
    assert (working_dir / "failure_screening.xlsx").exists()


def test_real_inventory_is_read_only_for_checkpoint_and_telemetry_databases():
    checkpoint_before = fingerprint_usage_db(QA_CHECKPOINT_DB_PATH)
    telemetry_before = fingerprint_usage_db(USAGE_DB_PATH)

    k5.build_inventory()

    assert fingerprint_usage_db(QA_CHECKPOINT_DB_PATH) == checkpoint_before
    assert fingerprint_usage_db(USAGE_DB_PATH) == telemetry_before
