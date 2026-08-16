"""Focused K5B.2 blind annotation workbook tests."""

from __future__ import annotations

import json
import os
import socket
import sys
from pathlib import Path
from unittest.mock import Mock, patch

import pytest
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import scripts.k5_keyword_annotation as annotation
from research_agent.qa import QA_CHECKPOINT_DB_PATH
from research_agent.telemetry import USAGE_DB_PATH
from tests._usage_db_fingerprint import fingerprint_usage_db


class FakeReferenceExtractor:
    def extract_keywords(self, text: str) -> list[tuple[str, float]]:
        paper_number = text.split("Synthetic title ", 1)[1].split()[0]
        return [
            ("Shared Term", 0.01),
            (f"reference concept {paper_number}", 0.02),
        ]


@pytest.fixture
def artifact_paths(tmp_path: Path, monkeypatch):
    working = tmp_path / "paper_keywords"
    paths = {
        "working": working,
        "sample": working / "proposed_sample.jsonl",
        "rules": working / "selection_rules.json",
        "source": working / "bounded_source_snapshot.jsonl",
        "workbook": working / "keyword_annotation_primary.xlsx",
        "mapping": working / "keyword_candidate_mapping.json",
        "manifest": working / "keyword_annotation_manifest.json",
    }
    monkeypatch.setattr(annotation, "EVAL_WORKING_DIR", working)
    monkeypatch.setattr(annotation, "SAMPLE_PATH", paths["sample"])
    monkeypatch.setattr(annotation, "RULES_PATH", paths["rules"])
    monkeypatch.setattr(annotation, "SOURCE_SNAPSHOT_PATH", paths["source"])
    monkeypatch.setattr(annotation, "WORKBOOK_PATH", paths["workbook"])
    monkeypatch.setattr(annotation, "MAPPING_PATH", paths["mapping"])
    monkeypatch.setattr(annotation, "MANIFEST_PATH", paths["manifest"])
    return paths


def _write_frozen_inputs(paths: dict[str, Path], count: int = 10) -> None:
    paths["working"].mkdir(parents=True, exist_ok=True)
    paper_ids = [f"paper-{index:02d}" for index in range(1, count + 1)]
    pilot_ids = paper_ids[:2]
    sample_rows = []
    source_rows = []
    for index, paper_id in enumerate(paper_ids, start=1):
        title = f"Synthetic title {index}"
        abstract = (
            f"This synthetic abstract {index} describes graph learning methods for scientific data, "
            "including evaluation protocols, reproducible experiments, and robust predictive models."
        )
        source_hash = annotation._source_hash(title, abstract)
        sample_rows.append({
            "schema_version": annotation.FROZEN_SAMPLE_SCHEMA_VERSION,
            "selection_rule_version": annotation.FROZEN_SELECTION_RULE_VERSION,
            "status": "frozen_bounded_sample",
            "paper_id": paper_id,
            "source_hash": source_hash,
            "metrics_role": "pilot_only" if paper_id in pilot_ids else "headline",
            "stratum": "confirmed_failure" if index <= 4 else "product_local_diversity",
        })
        source_rows.append({
            "schema_version": annotation.FROZEN_SOURCE_SCHEMA_VERSION,
            "paper_id": paper_id,
            "title": title,
            "abstract": abstract,
            "source": "synthetic",
            "external_id": paper_id,
            "source_hash": source_hash,
        })
    paths["sample"].write_text("".join(json.dumps(row, sort_keys=True) + "\n" for row in sample_rows))
    paths["source"].write_text("".join(json.dumps(row, sort_keys=True) + "\n" for row in source_rows))
    rules = {
        "schema_version": annotation.FROZEN_SAMPLE_SCHEMA_VERSION,
        "selection_rule_version": annotation.FROZEN_SELECTION_RULE_VERSION,
        "status": "frozen_bounded_sample",
        "paper_ids": paper_ids,
        "pilot_ids": pilot_ids,
        "sample_sha256": annotation._file_sha256(paths["sample"]),
        "bounded_source_snapshot_sha256": annotation._file_sha256(paths["source"]),
    }
    rules["manifest_sha256"] = annotation._sha256_payload(rules)
    paths["rules"].write_text(json.dumps(rules, sort_keys=True, indent=2) + "\n")


def _validate(paths: dict[str, Path], require_pilots_complete: bool = False) -> list[str]:
    return annotation.validate_annotation_workbook(
        require_pilots_complete=require_pilots_complete,
        workbook_path=paths["workbook"],
        mapping_path=paths["mapping"],
        manifest_path=paths["manifest"],
        sample_path=paths["sample"],
        rules_path=paths["rules"],
        source_snapshot_path=paths["source"],
    )


def _generate(paths: dict[str, Path], monkeypatch) -> None:
    _write_frozen_inputs(paths)
    monkeypatch.setattr(annotation.yake, "KeywordExtractor", lambda **_kwargs: FakeReferenceExtractor())
    monkeypatch.setattr(
        annotation.production_keywords,
        "extract_keywords",
        lambda title, _abstract: ["shared-term", f"production concept {title.rsplit(' ', 1)[-1]}"],
    )
    assert annotation.generate_annotation_artifacts(replace=True) == 0


def test_frozen_input_contract_requires_exactly_ten_papers(artifact_paths):
    _write_frozen_inputs(artifact_paths, count=10)
    violations, frozen = annotation.validate_frozen_inputs(
        artifact_paths["sample"], artifact_paths["rules"], artifact_paths["source"],
    )
    assert violations == []
    assert frozen is not None and len(frozen["sample"]) == 10

    _write_frozen_inputs(artifact_paths, count=9)
    violations, _ = annotation.validate_frozen_inputs(
        artifact_paths["sample"], artifact_paths["rules"], artifact_paths["source"],
    )
    assert any("exactly 10" in violation for violation in violations)


def test_candidate_union_and_order_are_deterministic(artifact_paths, monkeypatch):
    _write_frozen_inputs(artifact_paths)
    violations, frozen = annotation.validate_frozen_inputs(
        artifact_paths["sample"], artifact_paths["rules"], artifact_paths["source"],
    )
    assert violations == [] and frozen is not None
    monkeypatch.setattr(
        annotation.production_keywords,
        "extract_keywords",
        lambda title, _abstract: ["shared-term", f"production concept {title.rsplit(' ', 1)[-1]}"],
    )
    first = annotation.generate_candidate_bundle(frozen, reference_extractor_factory=FakeReferenceExtractor)
    second = annotation.generate_candidate_bundle(frozen, reference_extractor_factory=FakeReferenceExtractor)
    assert first["candidates"] == second["candidates"]
    assert first["method_counts"] == second["method_counts"]


def test_blind_workbook_exposes_no_method_rank_or_source_stratum(artifact_paths, monkeypatch):
    _generate(artifact_paths, monkeypatch)
    workbook = load_workbook(artifact_paths["workbook"], data_only=False)
    visible_text = "\n".join(
        str(cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    ).casefold()
    for forbidden in (
        annotation.REFERENCE_METHOD_ID,
        annotation.PRODUCTION_METHOD_ID,
        "confirmed_failure",
        "product_local_diversity",
        "local_edge_case",
        "method rank",
    ):
        assert forbidden.casefold() not in visible_text


def test_mapping_preserves_many_to_many_method_provenance(artifact_paths, monkeypatch):
    _generate(artifact_paths, monkeypatch)
    mapping = json.loads(artifact_paths["mapping"].read_text())
    shared = [row for row in mapping["candidates"] if row["canonical_key"] == "shared term"]
    assert len(shared) == 10
    assert all(
        [(item["method_id"], item["rank"]) for item in row["methods"]]
        == [(annotation.REFERENCE_METHOD_ID, 1), (annotation.PRODUCTION_METHOD_ID, 1)]
        for row in shared
    )
    assert mapping["methods"][annotation.REFERENCE_METHOD_ID]["parameters"]["top"] == 6
    assert mapping["methods"][annotation.PRODUCTION_METHOD_ID]["parameters"]["maximum_candidates"] == 6


def test_existing_production_extractor_receives_unchanged_title_and_abstract(artifact_paths):
    _write_frozen_inputs(artifact_paths)
    violations, frozen = annotation.validate_frozen_inputs(
        artifact_paths["sample"], artifact_paths["rules"], artifact_paths["source"],
    )
    assert violations == [] and frozen is not None
    extractor = Mock(return_value=[])
    with patch.object(annotation.production_keywords, "extract_keywords", extractor):
        annotation.generate_candidate_bundle(frozen, reference_extractor_factory=FakeReferenceExtractor)
    expected_calls = [
        ((row["title"], row["abstract"]), {})
        for row in sorted(frozen["sources"], key=lambda item: item["paper_id"])
    ]
    assert [(call.args, call.kwargs) for call in extractor.call_args_list] == expected_calls


def test_generation_makes_no_provider_or_network_call(artifact_paths, monkeypatch):
    _write_frozen_inputs(artifact_paths)
    monkeypatch.setattr(annotation.yake, "KeywordExtractor", lambda **_kwargs: FakeReferenceExtractor())
    monkeypatch.setattr(annotation.production_keywords, "extract_keywords", lambda _title, _abstract: [])
    with patch.object(socket, "create_connection", side_effect=AssertionError("network call")):
        assert annotation.generate_annotation_artifacts(replace=True) == 0


def test_workbook_dropdowns_structure_and_protected_hashes(artifact_paths, monkeypatch):
    _generate(artifact_paths, monkeypatch)
    assert _validate(artifact_paths) == []
    workbook = load_workbook(artifact_paths["workbook"], data_only=False)
    concepts = workbook["Paper concepts"]
    review = workbook["Candidate review"]
    assert workbook.sheetnames == list(annotation.EXPECTED_SHEETS)
    assert concepts.freeze_panes == review.freeze_panes == "A2"
    observed = {(str(item.sqref), item.formula1) for item in review.data_validations.dataValidation}
    assert observed == annotation._validation_specs(review.max_row - 1)
    assert concepts["A2"].protection.locked is True
    assert concepts["E2"].protection.locked is False
    assert review["D2"].protection.locked is True
    assert review["E2"].protection.locked is False


def test_pilot_completion_can_be_required_while_headline_remains_blank(artifact_paths, monkeypatch):
    _generate(artifact_paths, monkeypatch)
    assert any("pilot" in violation for violation in _validate(artifact_paths, require_pilots_complete=True))
    workbook = load_workbook(artifact_paths["workbook"])
    concepts = workbook["Paper concepts"]
    pilot_codes = {
        concepts.cell(row=row, column=1).value
        for row in range(2, concepts.max_row + 1)
        if concepts.cell(row=row, column=2).value == "pilot"
    }
    for row in range(2, concepts.max_row + 1):
        if concepts.cell(row=row, column=1).value in pilot_codes:
            concepts.cell(row=row, column=5).value = "graph learning"
            concepts.cell(row=row, column=6).value = "scientific data"
            concepts.cell(row=row, column=7).value = "evaluation protocols"
    review = workbook["Candidate review"]
    for row in range(2, review.max_row + 1):
        if review.cell(row=row, column=2).value in pilot_codes:
            review.cell(row=row, column=5).value = "reject"
            review.cell(row=row, column=6).value = "other"
            review.cell(row=row, column=8).value = "medium"
    workbook.save(artifact_paths["workbook"])
    assert _validate(artifact_paths, require_pilots_complete=True) == []
    assert all(
        review.cell(row=row, column=5).value is None
        for row in range(2, review.max_row + 1)
        if review.cell(row=row, column=2).value not in pilot_codes
    )


def test_protected_workbook_and_mapping_tampering_are_rejected(artifact_paths, monkeypatch):
    _generate(artifact_paths, monkeypatch)
    workbook = load_workbook(artifact_paths["workbook"])
    workbook["Paper concepts"]["C2"] = "tampered title"
    workbook.save(artifact_paths["workbook"])
    assert any("protected paper evidence" in violation for violation in _validate(artifact_paths))

    _generate(artifact_paths, monkeypatch)
    mapping = json.loads(artifact_paths["mapping"].read_text())
    mapping["candidates"][0]["canonical_key"] = "tampered"
    artifact_paths["mapping"].write_text(json.dumps(mapping))
    violations = _validate(artifact_paths)
    assert any("mapping self-hash" in violation or "mapping file hash" in violation for violation in violations)


def test_generation_refuses_overwrite_without_replace(artifact_paths, monkeypatch):
    _generate(artifact_paths, monkeypatch)
    before = annotation._file_sha256(artifact_paths["workbook"])
    assert annotation.generate_annotation_artifacts(replace=False) == 3
    assert annotation._file_sha256(artifact_paths["workbook"]) == before


def test_generation_leaves_real_checkpoint_and_telemetry_unchanged(artifact_paths, monkeypatch):
    _write_frozen_inputs(artifact_paths)
    monkeypatch.setattr(annotation.yake, "KeywordExtractor", lambda **_kwargs: FakeReferenceExtractor())
    monkeypatch.setattr(annotation.production_keywords, "extract_keywords", lambda _title, _abstract: [])
    checkpoint_before = fingerprint_usage_db(QA_CHECKPOINT_DB_PATH)
    telemetry_before = fingerprint_usage_db(USAGE_DB_PATH)

    assert annotation.generate_annotation_artifacts(replace=True) == 0

    assert fingerprint_usage_db(QA_CHECKPOINT_DB_PATH) == checkpoint_before
    assert fingerprint_usage_db(USAGE_DB_PATH) == telemetry_before
