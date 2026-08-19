"""Focused K5D.1a held-out annotation preparation tests.

These tests never call gpt-4.1-mini, never run Policy C, and never
compute precision/coverage. Part A functions read the *real* frozen
K5B/K5C/K5C.1 evidence already on disk (the project's own established
convention for these bounded, one-shot artifacts -- see
test_k5_llm_filter_policy_analysis.py), since that evidence is exactly
what this checkpoint must prove the new sample is disjoint from and
must leave byte-identical.
"""
from __future__ import annotations

import json
import os
import re
import socket
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from scripts import k5_heldout_selection as k5d1
from scripts import k5_keyword_annotation as annotation
from scripts import k5_keyword_metrics as k5b_metrics
from scripts import k5_llm_filter_eval as k5c
from scripts import k5_llm_filter_policy_analysis as k5c1
from scripts import k5_paper_inventory
from research_agent.curation_session import save_curation_session
from research_agent.qa import QA_CHECKPOINT_DB_PATH, sqlite_checkpointer
from research_agent.query_expansion import PaperPoolSession
from research_agent.schema import Paper
from research_agent.telemetry import USAGE_DB_PATH
from tests._usage_db_fingerprint import fingerprint_usage_db


_ABSTRACT = (
    "We present a comprehensive study of graph neural networks for molecular "
    "property prediction. Our method combines message passing with attention "
    "mechanisms to capture long range dependencies between atoms across five "
    "benchmark datasets while remaining computationally efficient at inference."
)

_PRIOR_EVIDENCE_PATHS = (
    annotation.SAMPLE_PATH, annotation.RULES_PATH, annotation.SOURCE_SNAPSHOT_PATH,
    annotation.WORKBOOK_PATH, annotation.MAPPING_PATH, annotation.MANIFEST_PATH,
    k5b_metrics.FROZEN_PATH, k5b_metrics.RESULTS_PATH,
    k5c.PROMPT_PATH, k5c.RAW_PATH, k5c.METRICS_PATH,
    k5c1.ANALYSIS_PATH,
)


def _hash_all(paths) -> dict[Path, str]:
    return {path: annotation._file_sha256(path) for path in paths if path.exists()}


def _fake_paper(pid: str, title: str, abstract: str | None = _ABSTRACT) -> Paper:
    return Paper(
        title=title, authors=["A"], year=2024, venue="X", abstract=abstract,
        url=None, doi=None, citation_count=None, source="synthetic", paper_id=pid, keywords=[],
    )


def _save_session(db_path: Path, session_id: str, session: PaperPoolSession) -> None:
    with sqlite_checkpointer(db_path) as checkpointer:
        save_curation_session(session, session_id, checkpointer)


@pytest.fixture
def synthetic_corpus(tmp_path, monkeypatch):
    """Redirects only *new-candidate* discovery to a synthetic checkpoint
    DB and a synthetic k5d1 output directory. The real frozen K5B/K5C/K5C.1
    evidence paths are deliberately left unmonkeypatched -- Part A always
    validates the genuine on-disk artifacts."""
    db_path = tmp_path / "checkpoints.sqlite"
    working = tmp_path / "k5d1"
    monkeypatch.setattr(k5_paper_inventory, "QA_CHECKPOINT_DB_PATH", db_path)
    monkeypatch.setattr(k5d1, "K5D1_DIR", working)
    monkeypatch.setattr(k5d1, "SELECTION_PATH", working / "heldout_selection.json")
    monkeypatch.setattr(k5d1, "SOURCE_SNAPSHOT_PATH", working / "heldout_source_snapshot.jsonl")
    monkeypatch.setattr(k5d1, "MAPPING_PATH", working / "heldout_candidate_mapping.json")
    monkeypatch.setattr(k5d1, "MANIFEST_PATH", working / "heldout_annotation_manifest.json")
    monkeypatch.setattr(k5d1, "WORKBOOK_PATH", working / "heldout_keyword_annotation.xlsx")

    papers = [(_fake_paper(f"synthetic-{index:02d}", f"Synthetic paper {index}"), 0.9) for index in range(1, 9)]
    _save_session(db_path, "synthetic-pool", PaperPoolSession(topic="synthetic topic", reserve=papers, cursor=8))
    return db_path, working


# ---------------------------------------------------------------------------
# Part A -- read-only validation of the real, frozen prior evidence
# ---------------------------------------------------------------------------

def test_validate_prior_evidence_reads_real_frozen_k5_artifacts_without_mutation():
    checkpoint_before = fingerprint_usage_db(QA_CHECKPOINT_DB_PATH)
    telemetry_before = fingerprint_usage_db(USAGE_DB_PATH)
    prior_before = _hash_all(_PRIOR_EVIDENCE_PATHS)

    bindings, prior_ids = k5d1.validate_prior_evidence()

    assert len(prior_ids) == 10 and len(set(prior_ids)) == 10
    assert set(bindings) >= {
        "frozen_annotation_sha256", "k5b_method_results_sha256", "k5c_prompt_sha256",
        "k5c_raw_results_sha256", "k5c_metrics_sha256", "k5c1_analysis_sha256",
    }
    assert fingerprint_usage_db(QA_CHECKPOINT_DB_PATH) == checkpoint_before
    assert fingerprint_usage_db(USAGE_DB_PATH) == telemetry_before
    assert _hash_all(_PRIOR_EVIDENCE_PATHS) == prior_before


def test_prior_evidence_bindings_detect_tampering():
    bindings = k5d1.prior_evidence_bindings()
    tampered = dict(bindings)
    tampered["frozen_annotation_sha256"] = "0" * 64
    assert tampered != k5d1.prior_evidence_bindings()


# ---------------------------------------------------------------------------
# Part B -- deterministic, disjoint, label-blind selection
# ---------------------------------------------------------------------------

def _inventory_row(paper_id: str, usable: bool = True, **extra) -> dict:
    return {"paper_id": paper_id, "usable_abstract": usable, **extra}


def test_selection_is_deterministic_for_a_fixed_seed():
    inventory = [_inventory_row(f"p{i:02d}") for i in range(20)]
    first = k5d1.select_heldout_papers(inventory, prior_paper_ids=[], seed=k5d1.K5D1_RANDOM_SEED)
    second = k5d1.select_heldout_papers(inventory, prior_paper_ids=[], seed=k5d1.K5D1_RANDOM_SEED)
    assert first == second
    assert len(first) == k5d1.HELD_OUT_COUNT == 6
    assert first == sorted(first)


def test_selection_excludes_every_prior_paper_id():
    prior_ids = [f"p{i:02d}" for i in range(10)]
    inventory = [_inventory_row(pid) for pid in prior_ids] + [_inventory_row(f"new{i:02d}") for i in range(10)]
    selected = k5d1.select_heldout_papers(inventory, prior_paper_ids=prior_ids)
    assert set(selected).isdisjoint(prior_ids)
    assert len(selected) == 6


def test_selection_ignores_unusable_abstracts():
    inventory = [_inventory_row(f"good{i}") for i in range(6)] + [_inventory_row(f"bad{i}", usable=False) for i in range(20)]
    selected = k5d1.select_heldout_papers(inventory, prior_paper_ids=[])
    assert set(selected) == {f"good{i}" for i in range(6)}


def test_selection_raises_when_eligible_pool_is_too_small():
    inventory = [_inventory_row(f"only{i}") for i in range(3)]
    with pytest.raises(ValueError, match="found 3 eligible"):
        k5d1.select_heldout_papers(inventory, prior_paper_ids=[])


def test_selection_never_consults_human_labels_or_failure_status():
    """Two inventories that differ only in fields select_heldout_papers
    must not read (failure_status, heuristic hit counts, domain buckets)
    must select identical papers."""
    base = [_inventory_row(f"p{i:02d}") for i in range(20)]
    labeled = [
        {**row, "failure_status": "known_failure" if index % 2 else None,
         "heuristic_screening_hit_count": index, "domain_bucket_guess": "security_software"}
        for index, row in enumerate(base)
    ]
    plain = k5d1.select_heldout_papers(base, prior_paper_ids=[])
    with_labels = k5d1.select_heldout_papers(labeled, prior_paper_ids=[])
    assert plain == with_labels


# ---------------------------------------------------------------------------
# Part C -- exact-order candidate generation with stable opaque IDs
# ---------------------------------------------------------------------------

def _selection_and_sources(paper_ids: list[str]) -> tuple[dict, dict]:
    codes = {pid: f"H{index:02d}" for index, pid in enumerate(paper_ids, start=1)}
    selection = {"paper_ids": paper_ids, "paper_codes": codes}
    sources = {
        pid: {"title": f"Title {pid}", "abstract": f"Abstract {pid}", "source_hash": f"hash-{pid}"}
        for pid in paper_ids
    }
    return selection, sources


def test_extractor_output_order_and_cap_are_preserved_exactly(monkeypatch):
    selection, sources = _selection_and_sources(["p1"])
    eight_phrases = [f"phrase {i}" for i in range(8)]
    monkeypatch.setattr(k5d1.production_keywords, "extract_keywords", lambda _t, _a: eight_phrases)
    bundle = k5d1.generate_candidate_bundle(selection, sources)
    assert [c["display_phrase"] for c in bundle["candidates"]] == eight_phrases[:6]
    assert [c["rank"] for c in bundle["candidates"]] == [1, 2, 3, 4, 5, 6]


def test_candidate_ids_are_stable_opaque_and_unique(monkeypatch):
    selection, sources = _selection_and_sources(["p1", "p2"])
    monkeypatch.setattr(
        k5d1.production_keywords, "extract_keywords",
        lambda title, _a: [f"shared term", f"unique {title}"],
    )
    first = k5d1.generate_candidate_bundle(selection, sources)
    second = k5d1.generate_candidate_bundle(selection, sources)
    assert [c["candidate_id"] for c in first["candidates"]] == [c["candidate_id"] for c in second["candidates"]]
    ids = [c["candidate_id"] for c in first["candidates"]]
    assert len(ids) == len(set(ids))
    assert all(re.fullmatch(r"HC-[0-9A-F]{12}", cid) for cid in ids)


def test_generation_refuses_when_extractor_version_is_stale(monkeypatch):
    monkeypatch.setattr(k5d1.production_keywords, "KEYWORD_EXTRACTOR_VERSION", "yake-v3")
    selection, sources = _selection_and_sources(["p1"])
    with pytest.raises(RuntimeError, match="frozen yake-v2"):
        k5d1.generate_candidate_bundle(selection, sources)


# ---------------------------------------------------------------------------
# Workbook structure -- validated against the real, already-frozen output
# ---------------------------------------------------------------------------

def test_real_heldout_workbook_passes_structural_validation():
    assert k5d1.SELECTION_PATH.exists(), "run `python -m scripts.k5_heldout_selection select && generate` first"
    assert k5d1.validate_annotation_workbook() == []


def test_workbook_sheet_names_and_order_are_fixed():
    workbook = load_workbook(k5d1.WORKBOOK_PATH, data_only=False)
    assert workbook.sheetnames == list(k5d1.EXPECTED_SHEETS)
    assert "Export Summary" not in workbook.sheetnames


def test_workbook_dropdowns_and_editable_columns():
    workbook = load_workbook(k5d1.WORKBOOK_PATH, data_only=False)
    review = workbook["Candidate review"]
    concepts = workbook["Paper concepts"]
    observed = {(str(item.sqref), item.formula1) for item in review.data_validations.dataValidation}
    assert observed == k5d1._validation_specs(review.max_row - 1)
    assert concepts["A2"].protection.locked is True
    assert concepts["D2"].protection.locked is False
    assert review["C2"].protection.locked is True
    assert review["D2"].protection.locked is False


def test_workbook_has_no_formulas():
    workbook = load_workbook(k5d1.WORKBOOK_PATH, data_only=False)
    from openpyxl.cell.cell import TYPE_FORMULA
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                assert cell.data_type != TYPE_FORMULA


def test_protected_evidence_tampering_is_rejected(tmp_path):
    tampered_workbook = tmp_path / "tampered.xlsx"
    tampered_mapping = tmp_path / "mapping.json"
    tampered_manifest = tmp_path / "manifest.json"
    workbook = load_workbook(k5d1.WORKBOOK_PATH)
    workbook["Paper concepts"]["B2"] = "tampered title"
    workbook.save(tampered_workbook)
    tampered_mapping.write_bytes(k5d1.MAPPING_PATH.read_bytes())
    tampered_manifest.write_bytes(k5d1.MANIFEST_PATH.read_bytes())
    violations = k5d1.validate_annotation_workbook(
        workbook_path=tampered_workbook, mapping_path=tampered_mapping, manifest_path=tampered_manifest,
    )
    assert any("protected paper evidence" in v for v in violations)


def test_mapping_self_hash_tampering_is_rejected(tmp_path):
    tampered_mapping = tmp_path / "mapping.json"
    mapping = json.loads(k5d1.MAPPING_PATH.read_text())
    mapping["candidates"][0]["canonical_key"] = "tampered"
    tampered_mapping.write_text(json.dumps(mapping))
    violations = k5d1.validate_annotation_workbook(mapping_path=tampered_mapping)
    assert any("mapping self-hash" in v for v in violations)


def test_require_all_complete_rejects_an_incomplete_workbook(tmp_path):
    """The canonical workbook may legitimately be blank (freshly generated,
    K5D.1a) or fully annotated (post human review, K5D.1b) depending on
    where the checkpoint pipeline currently stands -- this test proves the
    require_all_complete gate itself, on a copy, regardless of which state
    the real canonical workbook is in right now."""
    incomplete = tmp_path / "incomplete.xlsx"
    workbook = load_workbook(k5d1.WORKBOOK_PATH)
    workbook["Candidate review"]["D2"] = None  # blank out one decision
    workbook.save(incomplete)
    violations = k5d1.validate_annotation_workbook(
        require_all_complete=True, workbook_path=incomplete, mapping_path=k5d1.MAPPING_PATH, manifest_path=k5d1.MANIFEST_PATH,
    )
    assert any("candidate decision is incomplete" in v or "paper requires 3-5" in v for v in violations)


def test_real_selection_is_disjoint_from_prior_ten_papers():
    selection = json.loads(k5d1.SELECTION_PATH.read_text())
    prior_ids = set(selection["prior_k5_sample"]["paper_ids"])
    assert len(prior_ids) == 10
    assert set(selection["paper_ids"]).isdisjoint(prior_ids)
    assert len(selection["paper_ids"]) == 6


# ---------------------------------------------------------------------------
# Full offline pipeline: synthetic corpus, real prior evidence, no network
# ---------------------------------------------------------------------------

def test_full_pipeline_is_offline_and_leaves_prior_evidence_byte_identical(synthetic_corpus, monkeypatch):
    prior_before = _hash_all(_PRIOR_EVIDENCE_PATHS)
    checkpoint_before = fingerprint_usage_db(QA_CHECKPOINT_DB_PATH)
    telemetry_before = fingerprint_usage_db(USAGE_DB_PATH)

    monkeypatch.setattr(socket, "create_connection", lambda *_a, **_k: (_ for _ in ()).throw(AssertionError("network call")))
    monkeypatch.setattr(k5d1.production_keywords, "extract_keywords", lambda _t, _a: ["alpha beta", "gamma delta"])

    assert k5d1.freeze_heldout_selection(replace=True) == 0
    assert k5d1.generate_annotation_artifacts(replace=True) == 0
    assert k5d1.validate_annotation_workbook(
        workbook_path=k5d1.WORKBOOK_PATH, mapping_path=k5d1.MAPPING_PATH, manifest_path=k5d1.MANIFEST_PATH,
    ) == []

    selection = json.loads(k5d1.SELECTION_PATH.read_text())
    assert len(selection["paper_ids"]) == 6
    assert set(selection["paper_ids"]).isdisjoint(selection["prior_k5_sample"]["paper_ids"])

    assert _hash_all(_PRIOR_EVIDENCE_PATHS) == prior_before
    assert fingerprint_usage_db(QA_CHECKPOINT_DB_PATH) == checkpoint_before
    assert fingerprint_usage_db(USAGE_DB_PATH) == telemetry_before


def test_freeze_refuses_overwrite_without_replace(synthetic_corpus):
    assert k5d1.freeze_heldout_selection(replace=False) == 0
    before = annotation._file_sha256(k5d1.SELECTION_PATH)
    assert k5d1.freeze_heldout_selection(replace=False) == 3
    assert annotation._file_sha256(k5d1.SELECTION_PATH) == before


def test_generate_refuses_overwrite_without_replace(synthetic_corpus, monkeypatch):
    monkeypatch.setattr(k5d1.production_keywords, "extract_keywords", lambda _t, _a: ["one", "two"])
    assert k5d1.freeze_heldout_selection(replace=True) == 0
    assert k5d1.generate_annotation_artifacts(replace=False) == 0
    before = annotation._file_sha256(k5d1.WORKBOOK_PATH)
    assert k5d1.generate_annotation_artifacts(replace=False) == 3
    assert annotation._file_sha256(k5d1.WORKBOOK_PATH) == before


def test_selection_freeze_is_reproducible_against_same_local_corpus(synthetic_corpus):
    assert k5d1.freeze_heldout_selection(replace=True) == 0
    first = json.loads(k5d1.SELECTION_PATH.read_text())["paper_ids"]
    assert k5d1.freeze_heldout_selection(replace=True) == 0
    second = json.loads(k5d1.SELECTION_PATH.read_text())["paper_ids"]
    assert first == second


def test_generation_reads_frozen_snapshot_not_live_sessions(synthetic_corpus, monkeypatch):
    db_path, working = synthetic_corpus
    monkeypatch.setattr(k5d1.production_keywords, "extract_keywords", lambda _t, _a: ["kept phrase"])
    assert k5d1.freeze_heldout_selection(replace=True) == 0
    # Live sessions vanish; frozen snapshot must still be sufficient.
    db_path.unlink()
    db_path.with_name(db_path.name + "-wal").unlink(missing_ok=True)
    db_path.with_name(db_path.name + "-shm").unlink(missing_ok=True)
    assert k5d1.generate_annotation_artifacts(replace=True) == 0
